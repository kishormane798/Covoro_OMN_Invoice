"""Read a Covoro/OMN invoice workbook and compare filled upload columns to a download."""
from __future__ import annotations

import json
import re
import sys
from datetime import date, datetime

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel

HEADER_ROW_TEMPLATE = 4
DATA_ROW_TEMPLATE = 6
INVOICE_NUMBER_HEADER = "invoice number"
SHEET_CANDIDATES = ("E Invoice",)

_DATE_HEADER_RE = re.compile(r"date", re.I)
_NUMBER_RE = re.compile(r"^-?\d+(\.\d+)?$")


def normalize_header(value: object) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).strip())


def _header_key(value: object) -> str:
    return normalize_header(value).lower()


def _looks_like_date_header(header: str) -> bool:
    return bool(_DATE_HEADER_RE.search(header))


def _parse_date(text: str) -> date | None:
    raw = (text or "").strip()
    if not raw:
        return None
    iso = raw.replace("Z", "")
    try:
        return datetime.fromisoformat(iso).date()
    except ValueError:
        pass
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d", "%d.%m.%Y", "%m/%d/%Y"):
        snippet = raw[:10] if len(raw) >= 10 else raw
        try:
            return datetime.strptime(snippet, fmt).date()
        except ValueError:
            continue
    return None


def _parse_number(text: str) -> float | None:
    raw = (text or "").strip().replace(",", "")
    if not raw or not _NUMBER_RE.fullmatch(raw):
        return None
    try:
        return float(raw)
    except ValueError:
        return None


def cell_to_text(value: object, header: str) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, bool):
        return str(value).strip()
    if isinstance(value, (int, float)):
        if _looks_like_date_header(header) and 1 <= float(value) < 100000:
            try:
                parsed = from_excel(float(value))
                as_date = parsed.date() if isinstance(parsed, datetime) else parsed
                return as_date.isoformat()
            except Exception:
                pass
        if float(value).is_integer():
            return str(int(value))
        return format(value, "g")
    return str(value).strip()


def values_match(uploaded: str, downloaded: str, header: str) -> bool:
    left = (uploaded or "").strip()
    right = (downloaded or "").strip()
    if left == right:
        return True
    left_date = _parse_date(left)
    right_date = _parse_date(right)
    if left_date and right_date:
        return left_date == right_date
    if _looks_like_date_header(header) and left_date and right_date:
        return left_date == right_date
    left_num = _parse_number(left)
    right_num = _parse_number(right)
    if left_num is not None and right_num is not None:
        return abs(left_num - right_num) < 1e-9
    return False


def compare_filled_columns(
    uploaded: dict[str, str], downloaded: dict[str, str]
) -> list[str]:
    down_by_key = {_header_key(k): (k, (v or "").strip()) for k, v in downloaded.items()}
    mismatches: list[str] = []
    for header, raw in uploaded.items():
        uploaded_val = (raw or "").strip()
        if not uploaded_val:
            continue
        found = down_by_key.get(_header_key(header))
        if found is None:
            mismatches.append(f'"{header}": column not found')
            continue
        _dl_header, downloaded_val = found
        if not downloaded_val:
            mismatches.append(
                f'"{header}": uploaded "{uploaded_val}", downloaded empty'
            )
            continue
        if not values_match(uploaded_val, downloaded_val, header):
            mismatches.append(
                f'"{header}": uploaded "{uploaded_val}", downloaded "{downloaded_val}"'
            )
    return mismatches


def _pick_sheet(wb):
    for name in SHEET_CANDIDATES:
        if name in wb.sheetnames:
            return wb[name]
    return wb[wb.sheetnames[0]]


def _header_row_index(ws) -> int:
    for row_idx in range(1, 11):
        values = [normalize_header(c.value) for c in ws[row_idx]]
        if any(_header_key(v) == INVOICE_NUMBER_HEADER for v in values if v):
            return row_idx
    return HEADER_ROW_TEMPLATE


def _row_map(ws, header_row: int, data_row: int) -> dict[str, str]:
    headers = [normalize_header(c.value) for c in ws[header_row]]
    cells = list(ws[data_row])
    out: dict[str, str] = {}
    for i, header in enumerate(headers):
        if not header:
            continue
        value = cells[i].value if i < len(cells) else None
        out[header] = cell_to_text(value, header)
    return out


def _row_is_empty(row_map: dict[str, str]) -> bool:
    return not any((v or "").strip() for v in row_map.values())


def read_invoice_rows(file_path: str) -> dict:
    wb = load_workbook(file_path, data_only=True, read_only=False)
    try:
        ws = _pick_sheet(wb)
        header_row = _header_row_index(ws)
        start = (
            DATA_ROW_TEMPLATE if header_row == HEADER_ROW_TEMPLATE else header_row + 1
        )
        rows: list[dict[str, str]] = []
        max_row = ws.max_row or start
        for data_row in range(start, max_row + 1):
            mapped = _row_map(ws, header_row, data_row)
            if _row_is_empty(mapped):
                if rows:
                    break
                continue
            rows.append(mapped)
        invoice_number = ""
        filled: dict[str, str] = {}
        if rows:
            first = rows[0]
            for k, v in first.items():
                if _header_key(k) == INVOICE_NUMBER_HEADER:
                    invoice_number = v
                if (v or "").strip():
                    filled[k] = v
        return {
            "dataRowCount": len(rows),
            "invoiceNumber": invoice_number,
            "filled": filled,
        }
    finally:
        wb.close()


def compare_workbooks(upload_path: str, download_path: str) -> dict:
    uploaded = read_invoice_rows(upload_path)
    downloaded = read_invoice_rows(download_path)
    mismatches = compare_filled_columns(uploaded["filled"], downloaded.get("filled") or {})
    return {
        "invoiceNumber": uploaded["invoiceNumber"],
        "dataRowCount": uploaded["dataRowCount"],
        "mismatches": mismatches,
    }


def main(argv: list[str]) -> int:
    if len(argv) < 2:
        print(
            "Usage: read <xlsx> | compare <upload.xlsx> <download.xlsx>",
            file=sys.stderr,
        )
        return 2
    cmd = argv[0]
    if cmd == "read":
        print(json.dumps(read_invoice_rows(argv[1])))
        return 0
    if cmd == "compare":
        print(json.dumps(compare_workbooks(argv[1], argv[2])))
        return 0
    print(f"Unknown command: {cmd}", file=sys.stderr)
    return 2


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))
