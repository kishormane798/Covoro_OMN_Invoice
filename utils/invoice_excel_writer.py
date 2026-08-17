"""
Small Python writer for invoice template updates.
Used to avoid intermittent ExcelJS hangs while reading/writing template.xlsx.
"""

from __future__ import annotations

import json
import os
import random
import string
import sys
import time
import base64
from datetime import datetime, timedelta

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

_UTILS_DIR = os.path.dirname(os.path.abspath(__file__))
if _UTILS_DIR not in sys.path:
    sys.path.insert(0, _UTILS_DIR)

from app_config import resolve_base_url

# Keep in sync with `utils/invoiceExcel.ts` `INVOICE_TEMPLATE_DATA_ROW`.
INVOICE_TEMPLATE_DATA_ROW = 6
# Submit `clear_row`: wipe this many columns on the data row (A through column letter for 120).
INVOICE_SUBMIT_CLEAR_LAST_COL = 120

# Exact template headers (matches testData FieldValidations / Playwright FV constants).
FIELD_TYPE_OF_GOODS_OR_SERVICES_SUBJECT_TO_RCM = "Type of goods or services subject to RCM"
FIELD_ITEM_TYPE = "Item Type"
FIELD_CHARGES_ON_DOCUMENT_LEVEL = "Charges on document level"
FIELD_ALLOWANCES_ON_DOCUMENT_LEVEL = "Allowances on document level"
FIELD_VAT_CATEGORY_CHARGES = "Vat category - charges"
FIELD_VAT_CATEGORY_ALLOWANCES = "Vat category - allowances"
FIELD_TAX_EXEMPTION_REASON_CHARGES = "Tax exemption reason - charges"
FIELD_TAX_EXEMPTION_REASON_ALLOWANCES = "Tax exemption reason - allowances"
FIELD_TAX_EXEMPTION_REASON_CODE = "Tax exemption reason code"
FIELD_TAX_EXEMPTION_REASON_TEXT = "Tax exemption reason text"
STANDARD_TAX_CATEGORY = "Standard rate."
ZERO_RATED_TAX_CATEGORY = "Zero rated"
EXEMPT_FROM_TAX_TAX_CATEGORY = "Exempt from tax"
NOT_SUBJECT_TO_VAT_TAX_CATEGORY = "Services outside scope of tax / Not subject to tax"
DOCUMENT_CHARGES_SAMPLE_AMOUNT = "100"
DOCUMENT_ALLOWANCES_SAMPLE_AMOUNT = "50"
EXEMPTION_REASON_CERTAIN_FINANCIAL_SERVICES = "Certain financial services"
EXEMPTION_REASON_SUPPLY_RESIDENTIAL = "Supply of residential units (lease or sale)"
EXEMPTION_REASON_BARE_LAND = "Bare land"
EXEMPTION_REASON_LOCAL_PASSENGER_TRANSPORT = "Local passenger transport"
FIELD_INVOICE_TYPE_CODE = "Invoice Type Code"
FIELD_PAYMENT_MEANS_TYPE_CODE = "Payment means type code"
FIELD_TAX_CATEGORY = "Tax Category"
FIELD_TAX_RATE = "Tax Rate"
FIELD_LINE_ITEM_VAT_AMOUNT = "Line Item VAT Amount"
FIELD_TOTAL_AMOUNT_INCLUDING_VAT = "Total Amount Including VAT"
FIELD_TOTAL_AMOUNT_DUE_PROFIT_MARGIN = "Total Amount Due (Profit Margin)"
OMAN_HOME_CURRENCY = "OMR"
# Covoro / PINT-OM labels that require BTOM-020 (IBR-082-OM). Keep in sync with
# utils/invoiceExcel.ts `isProfitMarginTransactionType`.
_PROFIT_MARGIN_TRANSACTION_TYPE_LABELS = frozenset(
    {
        "profit margin invoice",
        "profit margin self-invoice",
    }
)
FIELD_SELLER_NAME = "Seller name"
FIELD_CREDIT_NOTE_REASON_CODE = "Credit note reason code"
FIELD_SELLER_PASSPORT_ISSUING_COUNTRY_CODE = "Seller - Passport issuing Country code"
FIELD_BUYER_PASSPORT_ISSUING_COUNTRY_CODE = "Buyer - Passport issuing Country code"
FIELD_SELLER_LEGAL_REGISTRATION_IDENTIFIER_TYPE = "Seller legal registration identifier type"
FIELD_BUYER_LEGAL_REGISTRATION_IDENTIFIER_TYPE = "Buyer legal registration identifier type"
FIELD_SCHEME_IDENTIFIER = "Scheme identifier"
FIELD_BUYER_LEGAL_REGISTRATION_IDENTIFIER = "Buyer legal registration identifier"
FIELD_BUYER_ELECTRONIC_ADDRESS = "Buyer electronic address"
FIELD_BUYER_ELECTRONIC_ADDRESS_SCHEME = "Buyer electronic address Scheme"
UAE_TIN_SCHEME_IDENTIFIER = "UAE Tax Identification Number (TIN)"
BUYER_LEGAL_REG_IDENTIFIER_FOR_DROPDOWN_BATCH = "100123456700003"
# Field-validation dropdown batch for Buyer legal registration identifier type only.
BUYER_ELECTRONIC_ADDRESS_FOR_LEGAL_REG_TYPE_DROPDOWN = "8900000099"
LEGAL_REGISTRATION_IDENTIFIER_TYPE_PASSPORT = "Passport"
FIELD_PRECEDING_INVOICE_REFERENCE = "Preceding Invoice reference"
FIELD_PRECEDING_INVOICE_ISSUE_DATE = "Preceding Invoice issue date"
# Exact sheet value (matches `REVERSE_CHARGE_TYPE_GOODS_SCENARIOS` / Playwright FV).
VAT_REVERSE_CHARGE_TAX_CATEGORY = "VAT Reverse Charge"
# Kept on sheet for RCM batch rows; `effective_tax_rate` ignores % unless category is Standard rate.
TAX_RATE_SHEET_DISPLAY_VALUE = "5"
CREDIT_NOTE_REASON_CODE_VOLUME_DISCOUNT = "Volume Discount."
INVOICE_TYPE_CODE_CREDIT_NOTE = "Credit note"
INVOICE_TYPE_CODE_CREDIT_NOTE_RELATED = "Credit note related to goods or services"
# Commercial invoice / credit note cannot contain only Exempt / Not subject line tax categories (dropdown batches).
INVOICE_TYPE_CODE_OUT_OF_SCOPE_OF_TAX = "Invoice out of scope of tax"
PAYMENT_MEANS_TYPE_CODE_INSTRUMENT_NOT_DEFINED = "Instrument not defined"

# Cleared when Tax Category is VAT Reverse Charge (keep in sync with `invoiceExcel.ts`).
DOCUMENT_LEVEL_CLEARED_FOR_VAT_REVERSE_CHARGE = (
    "Charges on document level",
    "Vat category - charges",
    "Tax exemption reason - charges",
    "Allowances on document level",
    "Vat category - allowances",
    "Tax exemption reason - allowances",
)
PAYMENT_FIELDS_CLEARED_FOR_CREDIT_NOTE_DROPDOWN = (
    "Payment means type code",
    "Scheme Identifier",
    "Payment account identifier",
    "Payment account name",
    "Payment service provider identifier",
    "Payment card primary account number",
)


def fail(message: str) -> None:
    print(message, file=sys.stderr)
    sys.exit(1)


def normalize(text: object) -> str:
    return str(text or "").strip().lower()


def is_profit_margin_transaction_type(value: object) -> bool:
    """True when Invoice Transaction Type Code requires Total Amount Due (Profit Margin)."""
    normalized = " ".join(str(value or "").split()).strip().lower()
    return normalized in _PROFIT_MARGIN_TRANSACTION_TYPE_LABELS


def normalize_invoice_header_label(name: str) -> str:
    """Match TS `normalizeInvoiceHeader`: collapse whitespace, trim; case preserved for first/last disambiguation."""
    return " ".join(str(name or "").split()).strip()


def _pick_duplicate_header_column(field_name: str, cols: list[int]) -> int:
    """
    Oman Title Case template duplicates `Scheme Identifier` and `Custom 1`/`Custom 2`.
    Legacy request-key casing selects the column: sentence/lower → first, Title Case → last.
    """
    if len(cols) == 1:
        return cols[0]
    wanted = str(field_name or "").strip()
    # Python str.title() lowercases letters after the first in each word (VAT→Vat), so
    # equality with title() is a good proxy for "request already looks Title Case".
    if wanted == wanted.title():
        return cols[-1]
    return cols[0]


def resolve_header_column(ws, header_row: int, field_name: str) -> int:
    """
    Resolve a row-4 column by header text.
    Prefer exact match; fall back to case-insensitive. When duplicates exist (Oman Title Case
    template), pick first vs last from the request key casing (see `_pick_duplicate_header_column`).
    """
    wanted = str(field_name).strip()
    if not wanted:
        fail("Empty field name for column resolution")
    wanted_lower = wanted.lower()
    exact_cols: list[int] = []
    ci_cols: list[int] = []
    for cell in ws[header_row]:
        raw = str(cell.value if cell.value is not None else "").strip()
        if not raw:
            continue
        if raw == wanted:
            exact_cols.append(cell.column)
        elif raw.lower() == wanted_lower:
            ci_cols.append(cell.column)
    if exact_cols:
        return _pick_duplicate_header_column(wanted, exact_cols)
    if ci_cols:
        return _pick_duplicate_header_column(wanted, ci_cols)
    fail(
        f"Column not found: {field_name!r}. "
        f"Use the exact row-{header_row} title from the template (matching case and spelling)."
    )


def read_header_row_cells(ws, header_row: int) -> list[str]:
    """Row 4-style headers; strip trailing empty cells (same spirit as ExcelJS reader)."""
    headers: list[str] = []
    for cell in ws[header_row]:
        val = cell.value
        text = "" if val is None else str(val).strip()
        headers.append(text)
    while headers and headers[-1] == "":
        headers.pop()
    return headers


def resolve_write_row_output_dir() -> str:
    """Default `testData/generated/excel` (uploads/ keeps template .xlsx only). Override with `INVOICE_EXCEL_OUTPUT_DIR`.

    With Playwright parallel workers, `TEST_PARALLEL_INDEX` is set per process; append `pw-<n>` so one worker's
    cleanup does not unlink another worker's generated .xlsx (matches `getGeneratedInvoiceExcelDir` in TS).
    """
    override = os.environ.get("INVOICE_EXCEL_OUTPUT_DIR", "").strip()
    if not override:
        base = os.path.join(os.getcwd(), "testData", "generated", "excel")
    elif os.path.isabs(override):
        base = override
    else:
        base = os.path.abspath(os.path.join(os.getcwd(), override))
    raw = os.environ.get("TEST_PARALLEL_INDEX", "").strip()
    if raw != "":
        try:
            n = int(raw)
            if n >= 0:
                return os.path.join(base, f"pw-{n}")
        except ValueError:
            pass
    return base


def random_string(length: int) -> str:
    chars = string.ascii_uppercase + string.digits
    return "".join(random.choice(chars) for _ in range(length))


def get_header_map(sheet, header_row: int) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for cell in sheet[header_row]:
        key = normalize(cell.value)
        if key:
            mapping[key] = cell.column
    return mapping


def get_col(header_map: dict[str, int], header_name: str) -> int:
    col = header_map.get(normalize(header_name))
    if not col:
        fail(f"Column not found: {header_name}")
    return col


def today_offset_date() -> datetime:
    days = random.randint(1, 10)
    dt = datetime.now() - timedelta(days=days)
    return datetime(dt.year, dt.month, dt.day)


def write_common(sheet, data_row: int, date_col: int, invoice_col: int) -> str:
    invoice_number = f"INV-{int(time.time() * 1000)}"
    sheet.cell(row=data_row, column=date_col).value = today_offset_date()
    sheet.cell(row=data_row, column=date_col).number_format = "yyyy-mm-dd"
    sheet.cell(row=data_row, column=invoice_col).value = invoice_number
    return invoice_number


def _safe_invoice_workbook_filename(invoice_number: str) -> str:
    """
    Disk name for the saved clone. Empty invoice # must not become ``.xlsx`` (invalid / breaks openpyxl on Windows).
    Row cell still uses the real value (including empty string).
    """
    stem = str(invoice_number).strip()
    if stem:
        return f"{stem}.xlsx"
    return f"INV-EMPTY-{int(time.time() * 1000)}.xlsx"


def to_number(value: object, fallback: float = 0.0) -> float:
    if value is None:
        return fallback
    s = str(value).strip()
    if s == "":
        return fallback
    try:
        return float(s)
    except ValueError:
        return fallback


def fix6(num: float) -> float:
    """Intermediate calculation precision (6 decimal places)."""
    return round(float(num), 6)


def ceil2(num: float) -> float:
    """Monetary outputs: round up to 2 decimal places."""
    import math

    x = float(num)
    if not math.isfinite(x):
        return 0.0
    return math.ceil(x * 100.0 - 1e-12) / 100.0


def normalize_category(value: object) -> str:
    return " ".join(str(value or "").split()).strip().lower()


def is_exempt_from_tax_tax_category(tax_category: object) -> bool:
    return normalize_category(tax_category) == normalize_category(EXEMPT_FROM_TAX_TAX_CATEGORY)


EXEMPT_BLANK_TAX_FIELD_HEADERS = (
    FIELD_TAX_RATE,
    "Standard Tax Rate",
    FIELD_LINE_ITEM_VAT_AMOUNT,
    "Line item VAT amount",
)


def _apply_exempt_from_tax_blank_tax_fields(
    ws, header_row: int, data_row: int, header_map: dict[str, int]
) -> None:
    """IBG-30: exempt lines keep tax-rate and VAT-line columns unset (not 0 / not '')."""
    if not is_exempt_from_tax_tax_category(
        cell_value(ws, data_row, header_map, FIELD_TAX_CATEGORY)
    ):
        return
    for header in EXEMPT_BLANK_TAX_FIELD_HEADERS:
        clear_cell_optional(ws, data_row, header_map, header)


def effective_tax_rate(tax_category: object, raw_rate: float) -> float:
    """
    VAT % applies only when Tax Category is Standard rate (with or without trailing period).
    Any other value (including empty) treats the sheet Tax Rate as null/0 for calculation.
    """
    cat = normalize_category(tax_category)
    if cat in ("standard rate", "standard rate."):
        return raw_rate
    return 0.0


def cell_value(ws, data_row: int, header_map: dict[str, int], *header_names: str) -> object:
    for name in header_names:
        col = header_map.get(normalize(name))
        if col is None:
            continue
        return ws.cell(row=data_row, column=col).value
    return None


def read_number(
    ws, data_row: int, header_map: dict[str, int], *header_names: str, default: float = 0.0
) -> float:
    return to_number(cell_value(ws, data_row, header_map, *header_names), default)


def set_cell_optional(
    ws, data_row: int, header_map: dict[str, int], header: str, value: object
) -> None:
    col = header_map.get(normalize(header))
    if not col:
        return
    ws.cell(row=data_row, column=col).value = value


def clear_cell_optional(
    ws, data_row: int, header_map: dict[str, int], header: str
) -> None:
    """
    Unset the cell value (no stored empty string). Unlike set_text_value / force_text_cell,
    we never write '' — backend/validation can treat that differently from a missing value.
    """
    col = header_map.get(normalize(header))
    if not col:
        return
    cell = ws.cell(row=data_row, column=col)
    cell.value = None
    if getattr(cell, "hyperlink", None) is not None:
        cell.hyperlink = None


def apply_invoice_calculations_to_data_row(ws, header_row: int, data_row: int) -> None:
    """
    Mirror utils/invoiceExcel.ts calculateInvoiceValues for one data row; no save.
    Used by apply_calculations CLI and write_dropdown_batch (single save after all rows).
    """
    header_map = get_header_map(ws, header_row)

    invoice_currency = str(
        cell_value(ws, data_row, header_map, "Invoice Currency Code") or OMAN_HOME_CURRENCY
    ).strip() or OMAN_HOME_CURRENCY
    currency_rate = read_number(ws, data_row, header_map, "Currency Exchange Rate", default=1.0)
    if currency_rate <= 0:
        currency_rate = 1.0

    item_price_base_qty = read_number(ws, data_row, header_map, "Item price base quantity", default=1.0)
    item_gross_price = read_number(ws, data_row, header_map, "Item gross price", default=0.0)
    item_price_discount = read_number(ws, data_row, header_map, "Item price discount", default=0.0)
    invoiced_qty = read_number(ws, data_row, header_map, "Invoiced quantity", default=0.0)
    line_charge = read_number(ws, data_row, header_map, "Invoice line charge amount", default=0.0)
    line_allowance = read_number(ws, data_row, header_map, "Invoice line allowance amount", default=0.0)
    doc_charges = read_number(ws, data_row, header_map, "Charges on document level", default=0.0)
    doc_allowances = read_number(ws, data_row, header_map, "Allowances on document level", default=0.0)
    paid_amount = read_number(ws, data_row, header_map, "Paid amount", default=0.0)
    rounding_amount = read_number(ws, data_row, header_map, "Rounding amount", default=0.0)

    tax_cat = cell_value(ws, data_row, header_map, "Tax Category")
    raw_tax = read_number(ws, data_row, header_map, "Tax Rate", "Standard Tax Rate", default=0.0)
    tax_rate = effective_tax_rate(tax_cat, raw_tax)

    item_net_price_raw = fix6(item_gross_price - item_price_discount)
    # Base quantity <= 0 is invalid for per-unit scaling; avoid ZeroDivisionError (negative formula tests).
    if item_price_base_qty > 0:
        line_net_raw = fix6(
            (item_net_price_raw * invoiced_qty) / item_price_base_qty + line_charge - line_allowance
        )
    else:
        line_net_raw = fix6(line_charge - line_allowance)

    vat_base_raw = fix6(line_net_raw * (tax_rate / 100.0))
    doc_charge_tax = fix6(doc_charges * (tax_rate / 100.0))
    doc_allowance_tax = fix6(doc_allowances * (tax_rate / 100.0))
    invoice_total_tax_raw = fix6(vat_base_raw + doc_charge_tax - doc_allowance_tax)

    line_plus_vat_raw = fix6(line_net_raw + vat_base_raw)

    item_net_price = ceil2(item_net_price_raw)
    invoice_line_net_amount = ceil2(line_net_raw)
    line_item_vat_amount = None if is_exempt_from_tax_tax_category(tax_cat) else ceil2(vat_base_raw)
    total_amount_including_vat = ceil2(line_plus_vat_raw)
    sum_invoice_line_net = ceil2(line_net_raw)

    total_without_raw = fix6(line_net_raw + doc_charges - doc_allowances)
    invoice_total_without_tax = ceil2(total_without_raw)
    invoice_total_tax = ceil2(invoice_total_tax_raw)
    if invoice_currency == OMAN_HOME_CURRENCY:
        invoice_total_tax_accounting = None
    else:
        invoice_total_tax_accounting = ceil2(fix6(invoice_total_tax_raw * currency_rate))

    total_with_raw = fix6(total_without_raw + invoice_total_tax_raw)
    invoice_total_with_tax = ceil2(total_with_raw)
    amount_due = ceil2(fix6(total_with_raw - paid_amount + rounding_amount))
    txn_type = cell_value(ws, data_row, header_map, "Invoice Transaction Type Code")
    # IBR-082-OM: fill only for Profit Margin Invoice / Profit Margin Self-Invoice.
    if is_profit_margin_transaction_type(txn_type):
        total_amount_due_profit_margin = total_amount_including_vat
    else:
        total_amount_due_profit_margin = None

    set_cell_optional(ws, data_row, header_map, "Item net price", item_net_price)
    set_cell_optional(ws, data_row, header_map, "Invoice line net amount", invoice_line_net_amount)
    if is_exempt_from_tax_tax_category(tax_cat):
        _apply_exempt_from_tax_blank_tax_fields(ws, header_row, data_row, header_map)
    else:
        set_cell_optional(
            ws, data_row, header_map, FIELD_LINE_ITEM_VAT_AMOUNT, line_item_vat_amount
        )
        set_cell_optional(
            ws, data_row, header_map, "Line item VAT amount", line_item_vat_amount
        )
    set_cell_optional(
        ws, data_row, header_map, FIELD_TOTAL_AMOUNT_INCLUDING_VAT, total_amount_including_vat
    )
    set_cell_optional(
        ws, data_row, header_map, "Total amount including VAT", total_amount_including_vat
    )
    set_cell_optional(ws, data_row, header_map, "Sum of Invoice line net amount", sum_invoice_line_net)
    set_cell_optional(ws, data_row, header_map, "Invoice total amount without tax", invoice_total_without_tax)
    set_cell_optional(ws, data_row, header_map, "Invoice total tax amount", invoice_total_tax)
    set_cell_optional(ws, data_row, header_map, "Invoice total amount with tax", invoice_total_with_tax)
    set_cell_optional(ws, data_row, header_map, "Amount due for payment", amount_due)
    if total_amount_due_profit_margin is None:
        clear_cell_optional(ws, data_row, header_map, FIELD_TOTAL_AMOUNT_DUE_PROFIT_MARGIN)
        clear_cell_optional(ws, data_row, header_map, "Total amount due (profit margin)")
    else:
        set_cell_optional(
            ws, data_row, header_map, FIELD_TOTAL_AMOUNT_DUE_PROFIT_MARGIN, total_amount_due_profit_margin
        )
        set_cell_optional(
            ws, data_row, header_map, "Total amount due (profit margin)", total_amount_due_profit_margin
        )
    if invoice_total_tax_accounting is None:
        set_cell_optional(ws, data_row, header_map, "Invoice total tax amount in tax accounting currency", None)
    else:
        set_cell_optional(
            ws,
            data_row,
            header_map,
            "Invoice total tax amount in tax accounting currency",
            invoice_total_tax_accounting,
        )


def cmd_apply_calculations(args: list[str]) -> None:
    """
    Read numeric inputs from data row, mirror utils/invoiceExcel.ts (calculateInvoiceValues), write calculated columns.
    """
    if len(args) < 4:
        fail("Usage: apply_calculations <filePath> <sheetName> <headerRow> <dataRow>")
    file_path, sheet_name, header_row_s, data_row_s = args[:4]
    header_row = int(header_row_s)
    data_row = int(data_row_s)

    if not os.path.exists(file_path):
        fail(f"File not found: {file_path}")

    wb = load_workbook(file_path)
    if sheet_name not in wb.sheetnames:
        fail(f"Sheet '{sheet_name}' not found")
    ws = wb[sheet_name]
    apply_invoice_calculations_to_data_row(ws, header_row, data_row)

    wb.save(file_path)
    print(json.dumps({"ok": True, "filePath": os.path.abspath(file_path), "dataRow": data_row}))


def save_result(wb, output_dir: str, file_name: str, invoice_number: str) -> None:
    os.makedirs(output_dir, exist_ok=True)
    file_path = os.path.abspath(os.path.join(output_dir, file_name))
    wb.save(file_path)
    print(json.dumps({"filePath": file_path, "invoiceNumber": invoice_number}))


def strip_cell_validation(ws, row_number: int, column_number: int) -> None:
    """Remove data-validation rules that cover this cell so arbitrary values can be written."""
    target = f"{get_column_letter(column_number)}{row_number}"

    for dv in list(ws.data_validations.dataValidation):
        kept_ranges = []
        for cell_range in list(dv.ranges.ranges):
            if target not in cell_range:
                kept_ranges.append(str(cell_range))
        if kept_ranges:
            dv.ranges = " ".join(kept_ranges)
        else:
            ws.data_validations.dataValidation.remove(dv)


def strip_column_validation(ws, column_number: int) -> None:
    """Remove list/dropdown data-validation from an entire column.

    Covoro template dropdowns use errorStyle=stop over the whole data column
    (e.g. BM6:BM18 plus BM20:BM1048576). Stripping only the current cell leaves
    the rest of the column as a list, so Excel still refuses values that are not
    in the dedicated master. Invalid dropdown tests must clear the column list
    first, then write the junk label as plain text.
    """
    for dv in list(ws.data_validations.dataValidation):
        kept_ranges = []
        for cell_range in list(dv.ranges.ranges):
            if cell_range.min_col <= column_number <= cell_range.max_col:
                continue
            kept_ranges.append(str(cell_range))
        if kept_ranges:
            dv.ranges = " ".join(kept_ranges)
        else:
            ws.data_validations.dataValidation.remove(dv)


def force_text_cell(ws, row_number: int, column_number: int) -> None:
    strip_column_validation(ws, column_number)
    strip_cell_validation(ws, row_number, column_number)
    cell = ws.cell(row=row_number, column=column_number)
    cell.value = ""
    cell.number_format = "@"


def set_numeric_cell_value(ws, row_number: int, column_number: int, value: float) -> None:
    strip_cell_validation(ws, row_number, column_number)
    cell = ws.cell(row=row_number, column=column_number)
    cell.value = value
    cell.number_format = "0.00"


def set_text_value(ws, row_number: int, column_number: int, value: object) -> None:
    force_text_cell(ws, row_number, column_number)
    ws.cell(row=row_number, column=column_number).value = "" if value is None else str(value)


def cmd_patch_invoice_cell(args: list[str]) -> None:
    """
    Patch one cell on an existing workbook (e.g. break a calculated total after generateInvoiceExcel).
    Preserves invoice # and dates. Writes a numeric value; saves in place.
    """
    if len(args) < 6:
        fail(
            "Usage: patch_invoice_cell <filePath> <sheetName> <headerRow> <dataRow> "
            "<fieldNameExact> <value>"
        )
    file_path, sheet_name, header_row_s, data_row_s, field_name, value_s = args[:6]
    if not os.path.isfile(file_path):
        fail(f"File not found: {file_path}")
    header_row = int(header_row_s)
    data_row = int(data_row_s)
    try:
        num = float(str(value_s).strip().replace(",", ""))
    except ValueError:
        fail(f"Invalid numeric value for patch_invoice_cell: {value_s!r}")

    wb = load_workbook(file_path)
    if sheet_name not in wb.sheetnames:
        fail(f"Sheet '{sheet_name}' not found")
    ws = wb[sheet_name]
    target_col = resolve_header_column(ws, header_row, field_name)
    set_numeric_cell_value(ws, data_row, target_col, num)
    wb.save(file_path)
    print(json.dumps({"ok": True, "filePath": os.path.abspath(file_path)}))


def cmd_patch_invoice_text_cell(args: list[str]) -> None:
    """
    Patch one text cell on an existing workbook by exact/normalized header.
    Preserves invoice # and dates. Writes text (including blank) and saves in place.
    """
    if len(args) < 6:
        fail(
            "Usage: patch_invoice_text_cell <filePath> <sheetName> <headerRow> <dataRow> "
            "<fieldNameExact> <value>"
        )
    file_path, sheet_name, header_row_s, data_row_s, field_name, value = args[:6]
    if not os.path.isfile(file_path):
        fail(f"File not found: {file_path}")
    header_row = int(header_row_s)
    data_row = int(data_row_s)

    wb = load_workbook(file_path)
    if sheet_name not in wb.sheetnames:
        fail(f"Sheet '{sheet_name}' not found")
    ws = wb[sheet_name]
    target_col = resolve_header_column(ws, header_row, field_name)
    strip_column_validation(ws, target_col)
    set_text_value(ws, data_row, target_col, value)
    wb.save(file_path)
    print(json.dumps({"ok": True, "filePath": os.path.abspath(file_path)}))


def cmd_read_invoice_text_cell(args: list[str]) -> None:
    """Read one data-row cell by header. Used to verify invalid dropdown values before upload."""
    if len(args) < 5:
        fail(
            "Usage: read_invoice_text_cell <filePath> <sheetName> <headerRow> <dataRow> "
            "<fieldNameExact>"
        )
    file_path, sheet_name, header_row_s, data_row_s, field_name = args[:5]
    if not os.path.isfile(file_path):
        fail(f"File not found: {file_path}")
    header_row = int(header_row_s)
    data_row = int(data_row_s)

    wb = load_workbook(file_path)
    if sheet_name not in wb.sheetnames:
        fail(f"Sheet '{sheet_name}' not found")
    ws = wb[sheet_name]
    target_col = resolve_header_column(ws, header_row, field_name)
    cell = ws.cell(row=data_row, column=target_col)
    value = "" if cell.value is None else str(cell.value)
    dropdown_present = False
    for dv in ws.data_validations.dataValidation:
        for cell_range in list(dv.ranges.ranges):
            if cell_range.min_col <= target_col <= cell_range.max_col:
                dropdown_present = True
                break
        if dropdown_present:
            break
    wb.close()
    print(
        json.dumps(
            {
                "ok": True,
                "value": value,
                "dropdownPresent": dropdown_present,
            }
        )
    )


def cmd_update_field(args: list[str]) -> None:
    if len(args) < 6:
        fail("Usage: update_field <templatePath> <sheetName> <headerRow> <dataRow> <fieldName> <length>")
    template_path, sheet_name, header_row_s, data_row_s, field_name, length_s = args[:6]
    header_row = int(header_row_s)
    data_row = int(data_row_s)
    length = int(length_s)

    wb = load_workbook(template_path)
    if sheet_name not in wb.sheetnames:
        fail(f"Sheet '{sheet_name}' not found")
    ws = wb[sheet_name]
    header_map = get_header_map(ws, header_row)
    date_col = get_col(header_map, "Invoice Issue Date")
    invoice_col = get_col(header_map, "Invoice Number")
    target_col = resolve_header_column(ws, header_row, field_name)
    invoice_number = write_common(ws, data_row, date_col, invoice_col)

    if length == 0:
        set_text_value(ws, data_row, target_col, "")
    elif length == -1:
        set_text_value(ws, data_row, target_col, " ")
    elif length == -2:
        set_text_value(ws, data_row, target_col, "   ")
    else:
        set_text_value(ws, data_row, target_col, random_string(length))

    _maybe_apply_parallel_worker_identity_row(
        ws, header_row, data_row, skip_if_field=field_name
    )
    save_result(wb, resolve_write_row_output_dir(), f"{invoice_number}.xlsx", invoice_number)


def cmd_update_number_field(args: list[str]) -> None:
    if len(args) < 6:
        fail("Usage: update_number_field <templatePath> <sheetName> <headerRow> <dataRow> <fieldName> <value>")
    template_path, sheet_name, header_row_s, data_row_s, field_name, value = args[:6]
    header_row = int(header_row_s)
    data_row = int(data_row_s)

    wb = load_workbook(template_path)
    if sheet_name not in wb.sheetnames:
        fail(f"Sheet '{sheet_name}' not found")
    ws = wb[sheet_name]
    header_map = get_header_map(ws, header_row)
    date_col = get_col(header_map, "Invoice Issue Date")
    invoice_col = get_col(header_map, "Invoice Number")
    target_col = resolve_header_column(ws, header_row, field_name)
    invoice_number = write_common(ws, data_row, date_col, invoice_col)
    set_text_value(ws, data_row, target_col, value)
    _maybe_apply_parallel_worker_identity_row(
        ws, header_row, data_row, skip_if_field=field_name
    )
    save_result(wb, resolve_write_row_output_dir(), f"{invoice_number}.xlsx", invoice_number)


def cmd_update_field_with_invoice(args: list[str]) -> None:
    if len(args) < 6:
        fail("Usage: update_field_with_invoice <templatePath> <sheetName> <headerRow> <dataRow> <fieldName> <length>")
    template_path, sheet_name, header_row_s, data_row_s, field_name, length_s = args[:6]
    header_row = int(header_row_s)
    data_row = int(data_row_s)
    length = int(length_s)

    wb = load_workbook(template_path)
    if sheet_name not in wb.sheetnames:
        fail(f"Sheet '{sheet_name}' not found")
    ws = wb[sheet_name]
    header_map = get_header_map(ws, header_row)
    date_col = get_col(header_map, "Invoice Issue Date")
    invoice_col = get_col(header_map, "Invoice Number")
    target_col = resolve_header_column(ws, header_row, field_name)
    invoice_number = write_common(ws, data_row, date_col, invoice_col)
    set_text_value(ws, data_row, target_col, random_string(length))
    _maybe_apply_parallel_worker_identity_row(
        ws, header_row, data_row, skip_if_field=field_name
    )
    save_result(wb, resolve_write_row_output_dir(), f"{invoice_number}.xlsx", invoice_number)


def cmd_update_issue_date(args: list[str]) -> None:
    if len(args) < 8:
        fail(
            "Usage: update_issue_date <templatePath> <sheetName> <headerRow> <dataRow> "
            "<invoiceNumber> <valueType> <value> <numberFormat>"
        )
    (
        template_path,
        sheet_name,
        header_row_s,
        data_row_s,
        invoice_number,
        value_type,
        raw_value,
        number_format,
    ) = args[:8]
    header_row = int(header_row_s)
    data_row = int(data_row_s)

    wb = load_workbook(template_path)
    if sheet_name not in wb.sheetnames:
        fail(f"Sheet '{sheet_name}' not found")
    ws = wb[sheet_name]
    header_map = get_header_map(ws, header_row)
    date_col = get_col(header_map, "Invoice Issue Date")
    invoice_col = get_col(header_map, "Invoice Number")

    ws.cell(row=data_row, column=invoice_col).value = invoice_number

    normalized_type = value_type.strip().lower()
    if normalized_type == "date":
        # Accept ISO values from TS (e.g. 2026-03-27T00:00:00.000Z).
        parsed = datetime.fromisoformat(raw_value.replace("Z", "+00:00"))
        ws.cell(row=data_row, column=date_col).value = datetime(parsed.year, parsed.month, parsed.day)
    elif normalized_type == "number":
        ws.cell(row=data_row, column=date_col).value = to_number(raw_value, 0.0)
    else:
        stripped = str(raw_value).strip() if raw_value is not None else ""
        ws.cell(row=data_row, column=date_col).value = (
            None if stripped == "" else raw_value
        )

    ws.cell(row=data_row, column=date_col).number_format = number_format
    _maybe_apply_parallel_worker_identity_row(ws, header_row, data_row)
    save_result(
        wb,
        resolve_write_row_output_dir(),
        _safe_invoice_workbook_filename(invoice_number),
        invoice_number,
    )


def _set_by_header_name(ws, header_row: int, data_row: int, header_name: str, value: object) -> None:
    # Prefer exact text match; fall back to case-insensitive. Duplicate Title Case headers
    # (Scheme Identifier / Custom 1/2) use request-key casing: sentence/lower → first, Title Case → last.
    target_exact = str(header_name).strip()
    target_lower = target_exact.lower()
    exact_cols: list[int] = []
    ci_cols: list[int] = []

    for cell in ws[header_row]:
        raw = str(cell.value or "").strip()
        if not raw:
            continue
        if raw == target_exact:
            exact_cols.append(cell.column)
        elif raw.lower() == target_lower:
            ci_cols.append(cell.column)

    cols = exact_cols or ci_cols
    if not cols:
        return
    col = _pick_duplicate_header_column(target_exact, cols)
    # For text/dropdown columns (including explicit blanks), remove data-validation and write as text.
    # This guarantees fields like "Payment means type code" are truly cleared when value is "".
    if isinstance(value, str):
        set_text_value(ws, data_row, col, value)
        return
    ws.cell(row=data_row, column=col).value = value


def cmd_write_row_json(args: list[str]) -> None:
    if len(args) < 11:
        fail(
            "Usage: write_row_json <templatePath> <sheetName> <headerRow> <dataRow> <invoiceNumber> "
            "<issueDateIso> <issueDateFormat> <fileName> <clearRow(0|1)> <rowJsonBase64> <strictHeaders(0|1)>"
        )
    (
        template_path,
        sheet_name,
        header_row_s,
        data_row_s,
        invoice_number,
        issue_date_iso,
        issue_date_format,
        file_name,
        clear_row_s,
        row_json_base64,
        strict_headers_s,
    ) = args[:11]
    header_row = int(header_row_s)
    data_row = int(data_row_s)
    clear_row = clear_row_s == "1"
    strict_headers = strict_headers_s == "1"

    wb = load_workbook(template_path)
    if sheet_name not in wb.sheetnames:
        fail(f"Sheet '{sheet_name}' not found")
    ws = wb[sheet_name]
    header_map = get_header_map(ws, header_row)
    date_col = get_col(header_map, "Invoice Issue Date")
    invoice_col = get_col(header_map, "Invoice Number")

    if clear_row:
        if data_row != INVOICE_TEMPLATE_DATA_ROW:
            fail(
                "write_row_json with clear_row=1 is submit-only: "
                f"data_row must be {INVOICE_TEMPLATE_DATA_ROW}, got {data_row}"
            )
        for col in range(1, INVOICE_SUBMIT_CLEAR_LAST_COL + 1):
            c = ws.cell(row=data_row, column=col)
            c.value = None
            c.number_format = "General"

    ws.cell(row=data_row, column=invoice_col).value = invoice_number
    parsed_date = datetime.fromisoformat(issue_date_iso.replace("Z", "+00:00"))
    ws.cell(row=data_row, column=date_col).value = datetime(parsed_date.year, parsed_date.month, parsed_date.day)
    ws.cell(row=data_row, column=date_col).number_format = issue_date_format

    try:
        row_json = base64.b64decode(row_json_base64.encode("utf-8")).decode("utf-8")
        values = json.loads(row_json)
    except json.JSONDecodeError:
        fail("Invalid rowJson: expected valid JSON object")
    if not isinstance(values, dict):
        fail("Invalid rowJson: expected JSON object")

    for key, value in values.items():
        # JSON null: do not write this column (submit row was cleared; avoids forcing '' on optional fields).
        if value is None:
            continue
        _set_by_header_name(ws, header_row, data_row, str(key), value)
        if strict_headers:
            # Strict mode verifies at least one matching column exists for the header.
            has_exact = any(str(c.value or "").strip() == str(key).strip() for c in ws[header_row])
            has_normalized = any(normalize(c.value) == normalize(key) for c in ws[header_row])
            if not has_exact and not has_normalized:
                fail(f"Column not found: {key}")

    tax_after_write = cell_value(ws, data_row, header_map, FIELD_TAX_CATEGORY)
    if normalize_category(tax_after_write) == normalize_category(VAT_REVERSE_CHARGE_TAX_CATEGORY):
        for hdr in DOCUMENT_LEVEL_CLEARED_FOR_VAT_REVERSE_CHARGE:
            _set_by_header_name(ws, header_row, data_row, hdr, "")
    if is_exempt_from_tax_tax_category(tax_after_write):
        _apply_exempt_from_tax_blank_tax_fields(ws, header_row, data_row, header_map)

    _maybe_apply_parallel_worker_identity_row(ws, header_row, data_row)
    save_result(wb, resolve_write_row_output_dir(), file_name, invoice_number)


def cmd_write_rows_json(args: list[str]) -> None:
    """
    Like `write_row_json` but writes multiple consecutive rows (data_row..data_row+n-1) in one save.
    Intended for **multi-line submit invoices** where one invoice number groups several line items.

    Usage:
      write_rows_json <templatePath> <sheetName> <headerRow> <dataRow> <invoiceNumber>
        <issueDateIso> <issueDateFormat> <fileName> <clearRow(0|1)> <rowsJsonBase64> <strictHeaders(0|1)>
    where rowsJsonBase64 decodes to a JSON array of objects (one object per row).
    """
    if len(args) < 11:
        fail(
            "Usage: write_rows_json <templatePath> <sheetName> <headerRow> <dataRow> <invoiceNumber> "
            "<issueDateIso> <issueDateFormat> <fileName> <clearRow(0|1)> <rowsJsonBase64> <strictHeaders(0|1)>"
        )
    (
        template_path,
        sheet_name,
        header_row_s,
        data_row_s,
        invoice_number,
        issue_date_iso,
        issue_date_format,
        file_name,
        clear_row_s,
        rows_json_base64,
        strict_headers_s,
    ) = args[:11]
    header_row = int(header_row_s)
    data_row = int(data_row_s)
    clear_row = clear_row_s == "1"
    strict_headers = strict_headers_s == "1"

    wb = load_workbook(template_path)
    if sheet_name not in wb.sheetnames:
        fail(f"Sheet '{sheet_name}' not found")
    ws = wb[sheet_name]
    header_map = get_header_map(ws, header_row)
    date_col = get_col(header_map, "Invoice Issue Date")
    invoice_col = get_col(header_map, "Invoice Number")

    try:
        rows_json = base64.b64decode(rows_json_base64.encode("utf-8")).decode("utf-8")
        rows_values = json.loads(rows_json)
    except json.JSONDecodeError:
        fail("Invalid rowsJson: expected valid JSON array")
    if not isinstance(rows_values, list) or len(rows_values) == 0:
        fail("Invalid rowsJson: expected non-empty JSON array")

    parsed_date = datetime.fromisoformat(issue_date_iso.replace("Z", "+00:00"))
    issue_date = datetime(parsed_date.year, parsed_date.month, parsed_date.day)

    for i, values in enumerate(rows_values):
        if not isinstance(values, dict):
            fail(f"Invalid rowsJson: expected object at index {i}")
        row_number = data_row + i

        # For multi-line invoices, the template's row-6 formatting/validations should be cloned down,
        # otherwise later line rows may miss data validation / formats and the app parser can behave differently.
        if i > 0:
            for col in range(1, ws.max_column + 1):
                src = ws.cell(row=data_row, column=col)
                dst = ws.cell(row=row_number, column=col)
                dst.value = src.value
                dst.number_format = src.number_format

        if clear_row:
            if data_row != INVOICE_TEMPLATE_DATA_ROW:
                fail(
                    "write_rows_json with clear_row=1 is submit-only: "
                    f"data_row must be {INVOICE_TEMPLATE_DATA_ROW}, got {data_row}"
                )
            for col in range(1, INVOICE_SUBMIT_CLEAR_LAST_COL + 1):
                c = ws.cell(row=row_number, column=col)
                c.value = None
                c.number_format = "General"

        ws.cell(row=row_number, column=invoice_col).value = invoice_number
        ws.cell(row=row_number, column=date_col).value = issue_date
        ws.cell(row=row_number, column=date_col).number_format = issue_date_format

        for key, value in values.items():
            if value is None:
                continue
            _set_by_header_name(ws, header_row, row_number, str(key), value)
            if strict_headers:
                has_exact = any(str(c.value or "").strip() == str(key).strip() for c in ws[header_row])
                has_normalized = any(normalize(c.value) == normalize(key) for c in ws[header_row])
                if not has_exact and not has_normalized:
                    fail(f"Column not found: {key}")

        tax_after_write = cell_value(ws, row_number, header_map, FIELD_TAX_CATEGORY)
        if normalize_category(tax_after_write) == normalize_category(VAT_REVERSE_CHARGE_TAX_CATEGORY):
            for hdr in DOCUMENT_LEVEL_CLEARED_FOR_VAT_REVERSE_CHARGE:
                _set_by_header_name(ws, header_row, row_number, hdr, "")
        if is_exempt_from_tax_tax_category(tax_after_write):
            _apply_exempt_from_tax_blank_tax_fields(ws, header_row, row_number, header_map)

    # IMPORTANT:
    # Do NOT apply `apply_invoice_calculations_to_data_row` here. Multi-line submit cases can include
    # invoice-level totals computed across *all* lines; applying per-row math would overwrite those
    # totals with single-line values and break the intended submit flow.

    if not _parallel_worker_identity_disabled():
        wi = _parallel_worker_index()
        for i in range(len(rows_values)):
            _apply_parallel_worker_identity_to_row(ws, header_row, data_row + i, wi)

    save_result(wb, resolve_write_row_output_dir(), file_name, invoice_number)


def cmd_write_rows_json_from_file(args: list[str]) -> None:
    """
    Same as `write_rows_json` but reads the rows array from a UTF-8 JSON file to avoid
    Windows command-line length limits.

    Usage:
      write_rows_json_from_file <templatePath> <sheetName> <headerRow> <dataRow> <invoiceNumber>
        <issueDateIso> <issueDateFormat> <fileName> <clearRow(0|1)> <rowsJsonPath> <strictHeaders(0|1)>
    where rowsJsonPath contains a JSON array of objects (one object per row).
    """
    if len(args) < 11:
        fail(
            "Usage: write_rows_json_from_file <templatePath> <sheetName> <headerRow> <dataRow> <invoiceNumber> "
            "<issueDateIso> <issueDateFormat> <fileName> <clearRow(0|1)> <rowsJsonPath> <strictHeaders(0|1)>"
        )
    (
        template_path,
        sheet_name,
        header_row_s,
        data_row_s,
        invoice_number,
        issue_date_iso,
        issue_date_format,
        file_name,
        clear_row_s,
        rows_json_path,
        strict_headers_s,
    ) = args[:11]
    header_row = int(header_row_s)
    data_row = int(data_row_s)
    clear_row = clear_row_s == "1"
    strict_headers = strict_headers_s == "1"

    if not os.path.isfile(rows_json_path):
        fail(f"Rows JSON file not found: {rows_json_path}")
    try:
        with open(rows_json_path, encoding="utf-8") as f:
            rows_values = json.load(f)
    except json.JSONDecodeError as e:
        fail(f"Invalid rows JSON file: {e}")

    # Reuse the existing implementation by packaging into the same argv shape
    # (keep logic in one place).
    # Note: cmd_write_rows_json expects base64 for the rows payload.
    rows_json_base64 = base64.b64encode(json.dumps(rows_values).encode("utf-8")).decode("utf-8")
    cmd_write_rows_json(
        [
            template_path,
            sheet_name,
            str(header_row),
            str(data_row),
            invoice_number,
            issue_date_iso,
            issue_date_format,
            file_name,
            "1" if clear_row else "0",
            rows_json_base64,
            "1" if strict_headers else "0",
        ]
    )


def _parse_optional_bool_cli_arg(raw: str | None, *, default: bool) -> bool:
    """Optional trailing CLI flag: 1/true/yes vs 0/false/no; missing or unknown → default."""
    if raw is None:
        return default
    s = str(raw).strip().lower()
    if s in ("0", "false", "no", "off"):
        return False
    if s in ("1", "true", "yes", "on"):
        return True
    return default


def _line_exemption_text_for_code(code: str) -> str:
    c = str(code or "").strip()
    if normalize(c) == normalize(EXEMPTION_REASON_BARE_LAND):
        return "Bare land exemption narrative."
    if normalize(c) == normalize(EXEMPTION_REASON_SUPPLY_RESIDENTIAL):
        return "Residential supply exemption narrative."
    if normalize(c) == normalize(EXEMPTION_REASON_LOCAL_PASSENGER_TRANSPORT):
        return "Local passenger transport exemption narrative."
    if normalize(c) == normalize(EXEMPTION_REASON_CERTAIN_FINANCIAL_SERVICES):
        return "Financial service exemption for document charges validation."
    return "Exemption reason text for exempt line."


def _apply_exempt_line_tax_exemption_reason(
    ws, header_row: int, row_number: int, reason_code: str
) -> None:
    """IBG-30: exempt line requires **Tax exemption reason code** and **text**."""
    reason = str(reason_code or "").strip() or EXEMPTION_REASON_CERTAIN_FINANCIAL_SERVICES
    _set_by_header_name(ws, header_row, row_number, FIELD_TAX_EXEMPTION_REASON_CODE, reason)
    _set_by_header_name(
        ws,
        header_row,
        row_number,
        FIELD_TAX_EXEMPTION_REASON_TEXT,
        _line_exemption_text_for_code(reason),
    )


def _clear_line_tax_exemption_reason(ws, header_row: int, row_number: int) -> None:
    _set_by_header_name(ws, header_row, row_number, FIELD_TAX_EXEMPTION_REASON_CODE, "")
    _set_by_header_name(ws, header_row, row_number, FIELD_TAX_EXEMPTION_REASON_TEXT, "")


def _apply_exempt_document_charges_row(
    ws, header_row: int, row_number: int, reason: str
) -> None:
    """Shared exempt document + line setup for exemption-reason and line-code dropdown batches."""
    set_text_value(
        ws, row_number, resolve_header_column(ws, header_row, FIELD_TAX_CATEGORY), EXEMPT_FROM_TAX_TAX_CATEGORY
    )
    _apply_exempt_or_not_subject_line_dropdown_invoice_defaults(ws, header_row, row_number)
    _set_by_header_name(
        ws, header_row, row_number, FIELD_VAT_CATEGORY_CHARGES, EXEMPT_FROM_TAX_TAX_CATEGORY
    )
    _set_by_header_name(
        ws, header_row, row_number, FIELD_VAT_CATEGORY_ALLOWANCES, EXEMPT_FROM_TAX_TAX_CATEGORY
    )
    _set_by_header_name(ws, header_row, row_number, FIELD_TAX_EXEMPTION_REASON_CHARGES, reason)
    _set_by_header_name(ws, header_row, row_number, FIELD_TAX_EXEMPTION_REASON_ALLOWANCES, reason)
    _apply_exempt_line_tax_exemption_reason(ws, header_row, row_number, reason)


def _apply_exempt_or_not_subject_line_dropdown_invoice_defaults(
    ws, header_row: int, row_number: int
) -> None:
    """
    Dropdown rows with line **Exempt from tax** or **Not subject to tax** only: use
    **Invoice out of scope of tax** (Commercial invoice / Credit note cannot contain only those categories).
    """
    set_text_value(
        ws,
        row_number,
        resolve_header_column(ws, header_row, FIELD_INVOICE_TYPE_CODE),
        INVOICE_TYPE_CODE_OUT_OF_SCOPE_OF_TAX,
    )
    _set_by_header_name(ws, header_row, row_number, FIELD_CREDIT_NOTE_REASON_CODE, "")
    _set_by_header_name(
        ws,
        header_row,
        row_number,
        FIELD_PAYMENT_MEANS_TYPE_CODE,
        PAYMENT_MEANS_TYPE_CODE_INSTRUMENT_NOT_DEFINED,
    )
    for hdr in PAYMENT_FIELDS_CLEARED_FOR_CREDIT_NOTE_DROPDOWN:
        if hdr == FIELD_PAYMENT_MEANS_TYPE_CODE:
            continue
        _set_by_header_name(ws, header_row, row_number, hdr, "")


def _line_tax_for_document_vat_category(doc_vat: str) -> tuple[str, str]:
    """Map document VAT category batch value → line Tax Category + Tax Rate."""
    v = str(doc_vat or "").strip()
    if normalize(v) == normalize(ZERO_RATED_TAX_CATEGORY):
        return ZERO_RATED_TAX_CATEGORY, "0"
    if normalize(v) == normalize(EXEMPT_FROM_TAX_TAX_CATEGORY):
        return EXEMPT_FROM_TAX_TAX_CATEGORY, ""
    if normalize(v) == normalize(NOT_SUBJECT_TO_VAT_TAX_CATEGORY):
        return NOT_SUBJECT_TO_VAT_TAX_CATEGORY, ""
    return STANDARD_TAX_CATEGORY, TAX_RATE_SHEET_DISPLAY_VALUE


def _apply_document_charges_allowances_dropdown_row(
    ws,
    header_row: int,
    row_number: int,
    field_name: str,
    batched_value: object,
) -> None:
    """Positive dropdown master: document amounts + line tax aligned with batched document VAT / reason."""
    fn = str(field_name).strip()
    _set_by_header_name(
        ws, header_row, row_number, FIELD_CHARGES_ON_DOCUMENT_LEVEL, DOCUMENT_CHARGES_SAMPLE_AMOUNT
    )
    _set_by_header_name(
        ws,
        header_row,
        row_number,
        FIELD_ALLOWANCES_ON_DOCUMENT_LEVEL,
        DOCUMENT_ALLOWANCES_SAMPLE_AMOUNT,
    )
    _set_by_header_name(ws, header_row, row_number, FIELD_TYPE_OF_GOODS_OR_SERVICES_SUBJECT_TO_RCM, "")

    if fn in (FIELD_VAT_CATEGORY_CHARGES, FIELD_VAT_CATEGORY_ALLOWANCES):
        doc_vat = str(batched_value if batched_value is not None else "").strip()
        line_cat, line_rate = _line_tax_for_document_vat_category(doc_vat)
        set_text_value(ws, row_number, resolve_header_column(ws, header_row, FIELD_TAX_CATEGORY), line_cat)
        set_text_value(ws, row_number, resolve_header_column(ws, header_row, FIELD_TAX_RATE), line_rate)
        _set_by_header_name(ws, header_row, row_number, FIELD_VAT_CATEGORY_CHARGES, doc_vat)
        _set_by_header_name(ws, header_row, row_number, FIELD_VAT_CATEGORY_ALLOWANCES, doc_vat)
        if normalize(doc_vat) == normalize(EXEMPT_FROM_TAX_TAX_CATEGORY):
            _apply_exempt_or_not_subject_line_dropdown_invoice_defaults(ws, header_row, row_number)
            _set_by_header_name(
                ws,
                header_row,
                row_number,
                FIELD_TAX_EXEMPTION_REASON_CHARGES,
                EXEMPTION_REASON_CERTAIN_FINANCIAL_SERVICES,
            )
            _set_by_header_name(
                ws,
                header_row,
                row_number,
                FIELD_TAX_EXEMPTION_REASON_ALLOWANCES,
                EXEMPTION_REASON_CERTAIN_FINANCIAL_SERVICES,
            )
            _apply_exempt_line_tax_exemption_reason(
                ws, header_row, row_number, EXEMPTION_REASON_CERTAIN_FINANCIAL_SERVICES
            )
        elif normalize(doc_vat) == normalize(NOT_SUBJECT_TO_VAT_TAX_CATEGORY):
            _apply_exempt_or_not_subject_line_dropdown_invoice_defaults(ws, header_row, row_number)
            _set_by_header_name(ws, header_row, row_number, FIELD_TAX_EXEMPTION_REASON_CHARGES, "")
            _set_by_header_name(ws, header_row, row_number, FIELD_TAX_EXEMPTION_REASON_ALLOWANCES, "")
            _clear_line_tax_exemption_reason(ws, header_row, row_number)
        else:
            _set_by_header_name(ws, header_row, row_number, FIELD_TAX_EXEMPTION_REASON_CHARGES, "")
            _set_by_header_name(ws, header_row, row_number, FIELD_TAX_EXEMPTION_REASON_ALLOWANCES, "")
            _clear_line_tax_exemption_reason(ws, header_row, row_number)
        return

    if fn == FIELD_TAX_EXEMPTION_REASON_CODE:
        reason = str(batched_value if batched_value is not None else "").strip()
        _apply_exempt_document_charges_row(ws, header_row, row_number, reason)
        return

    if fn in (FIELD_TAX_EXEMPTION_REASON_CHARGES, FIELD_TAX_EXEMPTION_REASON_ALLOWANCES):
        reason = str(batched_value if batched_value is not None else "").strip()
        _apply_exempt_document_charges_row(ws, header_row, row_number, reason)


def _apply_buyer_legal_reg_identifier_type_dropdown_context(
    ws, header_row: int, row_number: int, field_name: str
) -> None:
    """
    IBR-149-AE / IBR-183-AE preview for Buyer legal registration identifier type dropdown batches:
    Buyer electronic address Scheme 0235 (UAE TIN), identifier provided, address 8900000099 (not 1/9).
    Does not set buyer Scheme identifier or Deemed Supply.
    """
    if str(field_name).strip() != FIELD_BUYER_LEGAL_REGISTRATION_IDENTIFIER_TYPE:
        return
    _set_by_header_name(
        ws,
        header_row,
        row_number,
        FIELD_BUYER_LEGAL_REGISTRATION_IDENTIFIER,
        BUYER_LEGAL_REG_IDENTIFIER_FOR_DROPDOWN_BATCH,
    )
    _set_by_header_name(
        ws,
        header_row,
        row_number,
        FIELD_BUYER_ELECTRONIC_ADDRESS_SCHEME,
        UAE_TIN_SCHEME_IDENTIFIER,
    )
    _set_by_header_name(
        ws,
        header_row,
        row_number,
        FIELD_BUYER_ELECTRONIC_ADDRESS,
        BUYER_ELECTRONIC_ADDRESS_FOR_LEGAL_REG_TYPE_DROPDOWN,
    )


def _parallel_worker_skip_headers_for_dropdown_field(field_name: str) -> frozenset[str] | None:
    fn = str(field_name).strip()
    if fn == FIELD_BUYER_LEGAL_REGISTRATION_IDENTIFIER_TYPE:
        return frozenset(
            {
                fn,
                FIELD_BUYER_LEGAL_REGISTRATION_IDENTIFIER,
                FIELD_BUYER_ELECTRONIC_ADDRESS,
                FIELD_BUYER_ELECTRONIC_ADDRESS_SCHEME,
            }
        )
    if fn in _PARALLEL_IDENTITY_FIELD_NAMES:
        return frozenset([fn])
    return None


def _apply_passport_issuing_country_dropdown_context(
    ws, header_row: int, row_number: int, field_name: str
) -> None:
    """Passport country dropdown validation applies only when legal registration type is Passport."""
    fn = str(field_name).strip()
    if fn == FIELD_SELLER_PASSPORT_ISSUING_COUNTRY_CODE:
        _set_by_header_name(
            ws,
            header_row,
            row_number,
            FIELD_SELLER_LEGAL_REGISTRATION_IDENTIFIER_TYPE,
            LEGAL_REGISTRATION_IDENTIFIER_TYPE_PASSPORT,
        )
    elif fn == FIELD_BUYER_PASSPORT_ISSUING_COUNTRY_CODE:
        _set_by_header_name(
            ws,
            header_row,
            row_number,
            FIELD_BUYER_LEGAL_REGISTRATION_IDENTIFIER_TYPE,
            LEGAL_REGISTRATION_IDENTIFIER_TYPE_PASSPORT,
        )


def _run_write_dropdown_batch(
    template_path: str,
    sheet_name: str,
    header_row: int,
    data_row: int,
    field_name: str,
    values: list,
    file_name: str,
    *,
    clear_item_type_for_rcm: bool = True,
    clear_rcm_for_item_type: bool = True,
) -> None:
    if not isinstance(values, list) or len(values) == 0:
        fail("Dropdown batch: expected non-empty JSON array")

    wb = load_workbook(template_path)
    if sheet_name not in wb.sheetnames:
        fail(f"Sheet '{sheet_name}' not found")
    ws = wb[sheet_name]
    header_map = get_header_map(ws, header_row)
    date_col = get_col(header_map, "Invoice Issue Date")
    invoice_col = get_col(header_map, "Invoice Number")
    target_col = resolve_header_column(ws, header_row, field_name)
    # Clear the template list on this column so valid and invalid labels are
    # written as text (non-master values cannot be entered while the dropdown remains).
    strip_column_validation(ws, target_col)

    seed = int(time.time() * 1000)
    apply_credit_note_preceding_rule = (
        str(field_name).strip() == FIELD_CREDIT_NOTE_REASON_CODE
    )
    preceding_invoice_ref_col = None
    preceding_invoice_issue_date_col = None
    invoice_type_col = None
    tax_cat_col_for_credit_note = None
    tax_rate_col_for_credit_note = None
    if apply_credit_note_preceding_rule:
        preceding_invoice_ref_col = resolve_header_column(
            ws, header_row, FIELD_PRECEDING_INVOICE_REFERENCE
        )
        preceding_invoice_issue_date_col = resolve_header_column(
            ws, header_row, FIELD_PRECEDING_INVOICE_ISSUE_DATE
        )
        invoice_type_col = resolve_header_column(ws, header_row, FIELD_INVOICE_TYPE_CODE)
        tax_cat_col_for_credit_note = resolve_header_column(ws, header_row, FIELD_TAX_CATEGORY)
        tax_rate_col_for_credit_note = resolve_header_column(ws, header_row, FIELD_TAX_RATE)

    for i, value in enumerate(values):
        row_number = data_row + i
        if i > 0:
            for col in range(1, ws.max_column + 1):
                src = ws.cell(row=data_row, column=col)
                dst = ws.cell(row=row_number, column=col)
                dst.value = src.value
                dst.number_format = src.number_format

        ws.cell(row=row_number, column=invoice_col).value = f"INV-{seed}-{i}"
        ws.cell(row=row_number, column=date_col).value = today_offset_date()
        ws.cell(row=row_number, column=date_col).number_format = "yyyy-mm-dd"
        set_text_value(ws, row_number, target_col, value)
        # Oman: Source Currency is mandatory — fill OMR unless this batch is testing that field.
        if normalize(str(field_name)) != normalize("Source Currency Code"):
            _set_by_header_name(ws, header_row, row_number, "Source Currency Code", "OMR")
        if normalize(str(field_name)) != normalize("Invoice Currency Code"):
            # Keep currency coherent when template seed left invoice currency blank.
            cur = _read_data_row_text(ws, row_number, header_map, "Invoice Currency Code")
            if not cur:
                _set_by_header_name(ws, header_row, row_number, "Invoice Currency Code", "OMR")
        _apply_passport_issuing_country_dropdown_context(
            ws, header_row, row_number, field_name
        )
        _apply_buyer_legal_reg_identifier_type_dropdown_context(
            ws, header_row, row_number, field_name
        )
        if str(field_name).strip() in (
            FIELD_VAT_CATEGORY_CHARGES,
            FIELD_VAT_CATEGORY_ALLOWANCES,
            FIELD_TAX_EXEMPTION_REASON_CHARGES,
            FIELD_TAX_EXEMPTION_REASON_ALLOWANCES,
            FIELD_TAX_EXEMPTION_REASON_CODE,
        ):
            _apply_document_charges_allowances_dropdown_row(
                ws, header_row, row_number, field_name, value
            )
        if apply_credit_note_preceding_rule:
            # For credit-note-reason dropdown master uploads, validate against both invoice types:
            # even rows -> Credit note, odd rows -> Credit note related to goods or services.
            invoice_type_value = (
                INVOICE_TYPE_CODE_CREDIT_NOTE
                if i % 2 == 0
                else INVOICE_TYPE_CODE_CREDIT_NOTE_RELATED
            )
            set_text_value(ws, row_number, invoice_type_col, invoice_type_value)
            # For this dropdown only, clear document-level allowance/charge and payment blocks.
            for hdr in DOCUMENT_LEVEL_CLEARED_FOR_VAT_REVERSE_CHARGE:
                _set_by_header_name(ws, header_row, row_number, hdr, "")
            for hdr in PAYMENT_FIELDS_CLEARED_FOR_CREDIT_NOTE_DROPDOWN:
                _set_by_header_name(ws, header_row, row_number, hdr, "")
            if invoice_type_value == INVOICE_TYPE_CODE_CREDIT_NOTE_RELATED:
                set_text_value(ws, row_number, tax_cat_col_for_credit_note, "Zero rated")
                set_text_value(ws, row_number, tax_rate_col_for_credit_note, "0")

            reason = str(value if value is not None else "").strip()
            is_volume_discount = (
                normalize(reason) == normalize(CREDIT_NOTE_REASON_CODE_VOLUME_DISCOUNT)
            )
            if reason != "" and not is_volume_discount:
                set_text_value(ws, row_number, preceding_invoice_ref_col, "INV-01")
                today = datetime.now()
                ws.cell(
                    row=row_number, column=preceding_invoice_issue_date_col
                ).value = datetime(today.year, today.month, today.day)
                ws.cell(
                    row=row_number, column=preceding_invoice_issue_date_col
                ).number_format = "yyyy-mm-dd"
            else:
                set_text_value(ws, row_number, preceding_invoice_ref_col, "")
                set_text_value(ws, row_number, preceding_invoice_issue_date_col, "")

    # RCM dropdown values apply under reverse-charge tax category on the line.
    if str(field_name).strip() == FIELD_TYPE_OF_GOODS_OR_SERVICES_SUBJECT_TO_RCM:
        tax_cat_col = resolve_header_column(ws, header_row, FIELD_TAX_CATEGORY)
        tax_rate_col = resolve_header_column(ws, header_row, FIELD_TAX_RATE)
        for i in range(len(values)):
            set_text_value(ws, data_row + i, tax_cat_col, VAT_REVERSE_CHARGE_TAX_CATEGORY)
            set_text_value(ws, data_row + i, tax_rate_col, TAX_RATE_SHEET_DISPLAY_VALUE)
            for hdr in DOCUMENT_LEVEL_CLEARED_FOR_VAT_REVERSE_CHARGE:
                _set_by_header_name(ws, header_row, data_row + i, hdr, "")

    for i in range(len(values)):
        apply_invoice_calculations_to_data_row(ws, header_row, data_row + i)

    if not _parallel_worker_identity_disabled():
        wi = _parallel_worker_index()
        skip = _parallel_worker_skip_headers_for_dropdown_field(field_name)
        for i in range(len(values)):
            _apply_parallel_worker_identity_to_row(
                ws, header_row, data_row + i, wi, skip_headers=skip
            )
            _apply_buyer_legal_reg_identifier_type_dropdown_context(
                ws, header_row, data_row + i, field_name
            )

    # Cross-clear optional fields (unset via clear_cell_optional, never '') so template leftovers
    # do not trigger RCM+Item Type combination rules during single-column master uploads.
    fn = str(field_name).strip()
    if clear_item_type_for_rcm and fn == FIELD_TYPE_OF_GOODS_OR_SERVICES_SUBJECT_TO_RCM:
        for i in range(len(values)):
            clear_cell_optional(ws, data_row + i, header_map, FIELD_ITEM_TYPE)
    if clear_rcm_for_item_type and fn == FIELD_ITEM_TYPE:
        for i in range(len(values)):
            clear_cell_optional(
                ws,
                data_row + i,
                header_map,
                FIELD_TYPE_OF_GOODS_OR_SERVICES_SUBJECT_TO_RCM,
            )

    save_result(wb, resolve_write_row_output_dir(), file_name, f"INV-{seed}-0")


def cmd_write_dropdown_batch(args: list[str]) -> None:
    if len(args) < 7:
        fail(
            "Usage: write_dropdown_batch <templatePath> <sheetName> <headerRow> <dataRow> "
            "<fieldName> <valuesBase64> <fileName> [clearItemTypeForRcm=1] [clearRcmForItemType=1]"
        )
    template_path, sheet_name, header_row_s, data_row_s, field_name, values_b64, file_name = args[:7]
    header_row = int(header_row_s)
    data_row = int(data_row_s)
    clear_item_type_for_rcm = _parse_optional_bool_cli_arg(
        args[7] if len(args) > 7 else None, default=True
    )
    clear_rcm_for_item_type = _parse_optional_bool_cli_arg(
        args[8] if len(args) > 8 else None, default=True
    )

    try:
        values_json = base64.b64decode(values_b64.encode("utf-8")).decode("utf-8")
        values = json.loads(values_json)
    except Exception:
        fail("Invalid valuesBase64: expected base64 JSON array")

    _run_write_dropdown_batch(
        template_path,
        sheet_name,
        header_row,
        data_row,
        field_name,
        values,
        file_name,
        clear_item_type_for_rcm=clear_item_type_for_rcm,
        clear_rcm_for_item_type=clear_rcm_for_item_type,
    )


# Map Playwright worker slot to Oman seller VATINs; keep aligned with Helpers/parallelWorkerSubmitIdentity.ts.
_DEFAULT_OMAN_SELLER_TIN_SLOTS = [
    "OM1108202600",
    "OM1108202601",
    "OM1108202602",
    "OM1108202603",
    "OM1108202604",
]
_PARALLEL_WORKER_TIN_SLOTS = 5
# Default counterparty electronic address (not TRN/TIN); normal buyer / self-billed seller.
_COUNTERPARTY_ELECTRONIC_BY_ENV = {
    "dev": "OM1000091919",
    "preprod": "1008212295",
}


def _resolve_target_env() -> str:
    url = resolve_base_url().lower()
    return "preprod" if "preprod" in url else "dev"


def _seller_tin_slots() -> list[str]:
    """Optional comma/semicolon list from `UAE_EINVOICE_SELLER_TIN_SLOTS` (Oman VATIN slots)."""
    raw = os.environ.get("UAE_EINVOICE_SELLER_TIN_SLOTS", "").strip()
    if not raw:
        return []
    if (raw.startswith('"') and raw.endswith('"')) or (
        raw.startswith("'") and raw.endswith("'")
    ):
        raw = raw[1:-1].strip()
    return [
        part.strip().strip('"').strip("'")
        for part in raw.replace(";", ",").split(",")
        if part.strip().strip('"').strip("'")
    ]


def _electronic_tin_for_worker_index(worker_index: int) -> str:
    """Keep aligned with Helpers/parallelWorkerSubmitIdentity.electronicTinForParallelIndex."""
    slot = int(worker_index) % _PARALLEL_WORKER_TIN_SLOTS
    slots = _seller_tin_slots() or _DEFAULT_OMAN_SELLER_TIN_SLOTS
    return slots[slot % len(slots)]


def _worker_vat_for_electronic(worker_el: str) -> str:
    """Oman VATIN equals electronic address (no UAE `{tin}00003` suffix)."""
    return worker_el


def _counterparty_electronic_address() -> str:
    """Keep aligned with utils/envPartyIdentity.ts `getCounterpartyElectronicAddress`."""
    override = os.environ.get("UAE_EINVOICE_COUNTERPARTY_ELECTRONIC", "").strip()
    if override:
        return override
    return _COUNTERPARTY_ELECTRONIC_BY_ENV[_resolve_target_env()]


def _counterparty_vat() -> str:
    """Keep aligned with utils/envPartyIdentity.ts `getCounterpartyVatIdentifier`."""
    el = _counterparty_electronic_address()
    return el + "00003" if el.isdigit() else el


def _read_data_row_text(
    ws, data_row: int, header_map: dict[str, int], header_label: str
) -> str:
    col = header_map.get(normalize(header_label))
    if not col:
        return ""
    v = ws.cell(row=data_row, column=col).value
    return "" if v is None else str(v).strip()


def _apply_parallel_worker_identity_to_row(
    ws,
    header_row: int,
    data_row: int,
    worker_index: int,
    *,
    skip_headers: frozenset[str] | None = None,
) -> None:
    header_map = get_header_map(ws, header_row)

    def put(header: str, value: object) -> None:
        if skip_headers is not None and header in skip_headers:
            return
        set_cell_optional(ws, data_row, header_map, header, value)

    inv_type = _read_data_row_text(ws, data_row, header_map, "Invoice Type Code")
    txn_type = _read_data_row_text(ws, data_row, header_map, "Invoice Transaction Type Code")

    worker_el = _electronic_tin_for_worker_index(worker_index)
    worker_vat = _worker_vat_for_electronic(worker_el)
    t_inv = " ".join(inv_type.split()).strip().lower()
    t_txn = " ".join(txn_type.split()).strip().lower()
    self_billed = "self-billed" in t_inv or "self billed credit" in t_inv
    deemed = t_txn == "deemed supply"

    counterparty_el = _counterparty_electronic_address()
    if self_billed:
        put("Seller electronic address", counterparty_el)
        put("Seller VAT Identifier (TRN / TIN)", _counterparty_vat())
        put("Buyer electronic address", worker_el)
        put("Buyer VAT identifier", worker_vat)
    elif deemed:
        put("Seller electronic address", worker_el)
        put("Seller VAT Identifier (TRN / TIN)", worker_vat)
        put("Buyer electronic address", counterparty_el)
        # Keep Principle ID / legal registration aligned with patched seller TRN (avoids Submission Error after parallel TIN swap).
        put("Principle ID", worker_vat)
        put("Seller legal registration identifier", worker_vat)
    else:
        put("Seller electronic address", worker_el)
        put("Seller VAT Identifier (TRN / TIN)", worker_vat)
        put("Buyer electronic address", counterparty_el)
        put("Buyer VAT identifier", _counterparty_vat())


_PARALLEL_IDENTITY_FIELD_NAMES = frozenset(
    {
        "Seller electronic address",
        "Seller VAT Identifier (TRN / TIN)",
        "Buyer electronic address",
        "Buyer VAT identifier",
    }
)


def _parallel_worker_identity_disabled() -> bool:
    return os.environ.get("UAE_EINVOICE_DISABLE_WORKER_IDENTITY", "").strip() == "1"


def _parallel_worker_index() -> int:
    """Slot 0–4. Match Helpers/parallelWorkerSubmitIdentity.getParallelWorkerIndex."""
    raw_uae = os.environ.get("UAE_EINVOICE_WORKER_INDEX", "").strip()
    if raw_uae:
        try:
            return max(0, int(raw_uae)) % _PARALLEL_WORKER_TIN_SLOTS
        except ValueError:
            pass
    tp = os.environ.get("TEST_PARALLEL_INDEX", "").strip()
    if tp:
        try:
            return max(0, int(tp)) % _PARALLEL_WORKER_TIN_SLOTS
        except ValueError:
            pass
    return 0


def _maybe_apply_parallel_worker_identity_row(
    ws,
    header_row: int,
    data_row: int,
    *,
    skip_if_field: str | None = None,
) -> None:
    """
    After cloning/writing a row: set seller/buyer TIN cells per worker (env mirrors TS baseTest).

    If skip_if_field is one of _PARALLEL_IDENTITY_FIELD_NAMES, only that column is left unchanged;
    the other identity columns are still patched (so e.g. VAT/TRN is set when testing seller electronic only).
    """
    if _parallel_worker_identity_disabled():
        return
    skip: frozenset[str] | None = None
    if skip_if_field is not None:
        name = str(skip_if_field).strip()
        if name in _PARALLEL_IDENTITY_FIELD_NAMES:
            # Still patch VAT / buyer / other identity columns; only leave the field under test untouched.
            skip = frozenset([name])
    _apply_parallel_worker_identity_to_row(
        ws, header_row, data_row, _parallel_worker_index(), skip_headers=skip
    )


def cmd_apply_parallel_worker_identity(args: list[str]) -> None:
    if len(args) < 5:
        fail(
            "Usage: apply_parallel_worker_identity <filePath> <sheetName> <headerRow> <dataRow> "
            "<workerIndex> [rowCount]"
        )
    file_path, sheet_name, header_row_s, data_row_s, worker_index_s = args[:5]
    header_row = int(header_row_s)
    data_row = int(data_row_s)
    worker_index = int(worker_index_s)
    row_count = int(args[5]) if len(args) > 5 else 1
    row_count = max(1, min(row_count, 500))

    if not os.path.isfile(file_path):
        fail(f"File not found: {file_path}")

    wb = load_workbook(file_path)
    try:
        if sheet_name not in wb.sheetnames:
            fail(f"Sheet '{sheet_name}' not found")
        ws = wb[sheet_name]
        for i in range(row_count):
            _apply_parallel_worker_identity_to_row(ws, header_row, data_row + i, worker_index)
        wb.save(file_path)
    finally:
        wb.close()
    print(json.dumps({"ok": True}))


def cmd_write_dropdown_batch_from_file(args: list[str]) -> None:
    """
    Same as write_dropdown_batch but values come from a UTF-8 JSON file (array of strings).
    Avoids Windows command-line length limits when batching hundreds of long labels.
    """
    if len(args) < 7:
        fail(
            "Usage: write_dropdown_batch_from_file <templatePath> <sheetName> <headerRow> <dataRow> "
            "<fieldName> <valuesJsonPath> <fileName> [clearItemTypeForRcm=1] [clearRcmForItemType=1]"
        )
    template_path, sheet_name, header_row_s, data_row_s, field_name, values_path, file_name = args[:7]
    header_row = int(header_row_s)
    data_row = int(data_row_s)
    clear_item_type_for_rcm = _parse_optional_bool_cli_arg(
        args[7] if len(args) > 7 else None, default=True
    )
    clear_rcm_for_item_type = _parse_optional_bool_cli_arg(
        args[8] if len(args) > 8 else None, default=True
    )

    if not os.path.isfile(values_path):
        fail(f"Values file not found: {values_path}")
    try:
        with open(values_path, encoding="utf-8") as f:
            values = json.load(f)
    except json.JSONDecodeError as e:
        fail(f"Invalid values JSON file: {e}")

    _run_write_dropdown_batch(
        template_path,
        sheet_name,
        header_row,
        data_row,
        field_name,
        values,
        file_name,
        clear_item_type_for_rcm=clear_item_type_for_rcm,
        clear_rcm_for_item_type=clear_rcm_for_item_type,
    )


def cmd_write_currency_exchange_batch_from_file(args: list[str]) -> None:
    """
    Build one workbook with many rows for Invoice Currency Code + Currency Exchange Rate validation.
    mode:
      - allowed: AED -> blank exchange rate, non-AED -> "3.67"
      - invalid_blank_non_aed: every provided currency gets blank exchange rate (use non-AED list)
    """
    if len(args) < 7:
        fail(
            "Usage: write_currency_exchange_batch_from_file <templatePath> <sheetName> <headerRow> <dataRow> "
            "<valuesJsonPath> <fileName> <mode>"
        )
    template_path, sheet_name, header_row_s, data_row_s, values_path, file_name, mode = args[:7]
    header_row = int(header_row_s)
    data_row = int(data_row_s)
    mode_norm = str(mode).strip().lower()
    if mode_norm not in {"allowed", "invalid_blank_non_aed"}:
        fail(f"Invalid mode: {mode}. Expected 'allowed' or 'invalid_blank_non_aed'.")

    if not os.path.isfile(values_path):
        fail(f"Values file not found: {values_path}")
    try:
        with open(values_path, encoding="utf-8") as f:
            currencies = json.load(f)
    except json.JSONDecodeError as e:
        fail(f"Invalid values JSON file: {e}")

    if not isinstance(currencies, list) or len(currencies) == 0:
        fail("Currency batch: expected non-empty JSON array")

    wb = load_workbook(template_path)
    if sheet_name not in wb.sheetnames:
        fail(f"Sheet '{sheet_name}' not found")
    ws = wb[sheet_name]
    header_map = get_header_map(ws, header_row)
    date_col = get_col(header_map, "Invoice Issue Date")
    invoice_col = get_col(header_map, "Invoice Number")
    currency_col = resolve_header_column(ws, header_row, "Invoice Currency Code")
    exchange_col = resolve_header_column(ws, header_row, "Currency Exchange Rate")

    seed = int(time.time() * 1000)
    for i, raw_currency in enumerate(currencies):
        currency = "" if raw_currency is None else str(raw_currency).strip().upper()
        if not currency:
            fail(f"Currency batch: empty currency at index {i}")
        row_number = data_row + i
        if i > 0:
            for col in range(1, ws.max_column + 1):
                src = ws.cell(row=data_row, column=col)
                dst = ws.cell(row=row_number, column=col)
                dst.value = src.value
                dst.number_format = src.number_format

        ws.cell(row=row_number, column=invoice_col).value = f"INV-{seed}-{i}"
        ws.cell(row=row_number, column=date_col).value = today_offset_date()
        ws.cell(row=row_number, column=date_col).number_format = "yyyy-mm-dd"
        set_text_value(ws, row_number, currency_col, currency)

        exchange_value = ""
        if mode_norm == "allowed":
            exchange_value = "" if currency == "AED" else "3.67"
        set_text_value(ws, row_number, exchange_col, exchange_value)

    for i in range(len(currencies)):
        apply_invoice_calculations_to_data_row(ws, header_row, data_row + i)

    if not _parallel_worker_identity_disabled():
        wi = _parallel_worker_index()
        for i in range(len(currencies)):
            _apply_parallel_worker_identity_to_row(ws, header_row, data_row + i, wi)

    save_result(wb, resolve_write_row_output_dir(), file_name, f"INV-{seed}-0")


E_INVOICE_SHEET_NAME = "E Invoice"


def _next_regression_data_row_offset(
    ws, header_map: dict, data_row: int, invoice_col: int
) -> int:
    row = data_row
    while row <= ws.max_row:
        if ws.cell(row=row, column=invoice_col).value in (None, ""):
            break
        row += 1
    return row - data_row


def _run_write_regression_batch(
    template_path: str,
    sheet_name: str,
    header_row: int,
    data_row: int,
    file_name: str,
    units: list,
    *,
    default_issue_date_iso: str,
    default_issue_date_format: str = "yyyy-mm-dd",
    append_workbook_path: str | None = None,
    skip_worker_identity: bool = False,
) -> None:
    if not isinstance(units, list) or len(units) == 0:
        fail("Regression batch: expected non-empty JSON array of units")

    if append_workbook_path and os.path.isfile(append_workbook_path):
        wb = load_workbook(append_workbook_path)
    else:
        wb = load_workbook(template_path)
    if sheet_name not in wb.sheetnames:
        fail(f"Sheet '{sheet_name}' not found")
    ws = wb[sheet_name]
    header_map = get_header_map(ws, header_row)
    date_col = get_col(header_map, "Invoice Issue Date")
    invoice_col = get_col(header_map, "Invoice Number")

    if append_workbook_path and os.path.isfile(append_workbook_path):
        current_row_offset = _next_regression_data_row_offset(
            ws, header_map, data_row, invoice_col
        )
    else:
        current_row_offset = 0
    first_invoice: str | None = None
    chunk_start_offset = current_row_offset

    for unit_idx, unit in enumerate(units):
        if not isinstance(unit, dict):
            fail(f"Regression batch: unit {unit_idx} must be an object")
        rows = unit.get("rows")
        if not isinstance(rows, list) or len(rows) == 0:
            fail(f"Regression batch: unit {unit_idx} rows must be a non-empty array")
        invoice_number = str(unit.get("invoiceNumber") or "").strip()
        if not invoice_number:
            fail(f"Regression batch: unit {unit_idx} missing invoiceNumber")
        if first_invoice is None:
            first_invoice = invoice_number

        skip_calc = bool(unit.get("skipCalc"))
        clear_row = unit.get("clearRow", True)
        issue_date_format = str(unit.get("issueDateFormat") or default_issue_date_format)
        issue_date_iso = unit.get("issueDateIso", default_issue_date_iso)

        for i, values in enumerate(rows):
            if not isinstance(values, dict):
                fail(f"Regression batch: unit {unit_idx} row {i} must be an object")
            row_number = data_row + current_row_offset

            if i > 0:
                for col in range(1, ws.max_column + 1):
                    src = ws.cell(row=data_row, column=col)
                    dst = ws.cell(row=row_number, column=col)
                    dst.value = src.value
                    dst.number_format = src.number_format

            if clear_row and i == 0:
                if data_row != INVOICE_TEMPLATE_DATA_ROW:
                    fail(
                        "Regression batch clear_row requires "
                        f"data_row={INVOICE_TEMPLATE_DATA_ROW}, got {data_row}"
                    )
                for col in range(1, INVOICE_SUBMIT_CLEAR_LAST_COL + 1):
                    c = ws.cell(row=row_number, column=col)
                    c.value = None
                    c.number_format = "General"

            row_invoice = str(values.get("Invoice Number") or invoice_number).strip()
            if not row_invoice:
                row_invoice = invoice_number
            ws.cell(row=row_number, column=invoice_col).value = row_invoice

            if issue_date_iso is None or str(issue_date_iso).strip() == "":
                ws.cell(row=row_number, column=date_col).value = None
            else:
                parsed_date = datetime.fromisoformat(
                    str(issue_date_iso).replace("Z", "+00:00")
                )
                issue_date = datetime(
                    parsed_date.year, parsed_date.month, parsed_date.day
                )
                ws.cell(row=row_number, column=date_col).value = issue_date
            ws.cell(row=row_number, column=date_col).number_format = issue_date_format

            for key, value in values.items():
                if key == "Invoice Number":
                    continue
                if value is None:
                    continue
                _set_by_header_name(ws, header_row, row_number, str(key), value)

            tax_after_write = cell_value(ws, row_number, header_map, FIELD_TAX_CATEGORY)
            if normalize_category(tax_after_write) == normalize_category(
                VAT_REVERSE_CHARGE_TAX_CATEGORY
            ):
                for hdr in DOCUMENT_LEVEL_CLEARED_FOR_VAT_REVERSE_CHARGE:
                    _set_by_header_name(ws, header_row, row_number, hdr, "")
            if is_exempt_from_tax_tax_category(tax_after_write):
                _apply_exempt_from_tax_blank_tax_fields(
                    ws, header_row, row_number, header_map
                )

            if not skip_calc:
                apply_invoice_calculations_to_data_row(ws, header_row, row_number)

            current_row_offset += 1

    if not skip_worker_identity and not _parallel_worker_identity_disabled():
        wi = _parallel_worker_index()
        for off in range(chunk_start_offset, current_row_offset):
            _apply_parallel_worker_identity_to_row(ws, header_row, data_row + off, wi)

    save_result(
        wb,
        resolve_write_row_output_dir(),
        file_name,
        first_invoice or "REG-0",
    )


def cmd_write_regression_batch_from_file(args: list[str]) -> None:
    if len(args) < 7:
        fail(
            "Usage: write_regression_batch_from_file <templatePath> <sheetName> <headerRow> "
            "<dataRow> <fileName> <unitsJsonPath> <defaultIssueDateIso> [defaultIssueDateFormat] "
            "[appendWorkbookPath] [skipWorkerIdentity(0|1)]"
        )
    (
        template_path,
        sheet_name,
        header_row_s,
        data_row_s,
        file_name,
        units_json_path,
        default_issue_date_iso,
    ) = args[:7]
    default_issue_date_format = args[7] if len(args) > 7 else "yyyy-mm-dd"
    append_workbook_path = args[8] if len(args) > 8 and str(args[8]).strip() else None
    skip_worker_identity = len(args) > 9 and str(args[9]).strip() == "1"
    header_row = int(header_row_s)
    data_row = int(data_row_s)

    if not os.path.isfile(units_json_path):
        fail(f"Units JSON file not found: {units_json_path}")
    with open(units_json_path, encoding="utf-8-sig") as f:
        units = json.load(f)

    _run_write_regression_batch(
        template_path,
        sheet_name,
        header_row,
        data_row,
        file_name,
        units,
        default_issue_date_iso=default_issue_date_iso,
        default_issue_date_format=default_issue_date_format,
        append_workbook_path=append_workbook_path,
        skip_worker_identity=skip_worker_identity,
    )


def cmd_read_e_invoice_headers(args: list[str]) -> None:
    """
    Header row on sheet "E Invoice" only. Prints JSON: {"headers":["..."]}.
    openpyxl only (no ExcelJS).
    """
    if len(args) < 2:
        fail("Usage: read_e_invoice_headers <templatePath> <headerRow>")
    template_path, header_row_s = args[0], args[1]
    header_row = int(header_row_s)
    if not os.path.isfile(template_path):
        fail(f"Template not found: {template_path}")

    wb = load_workbook(template_path, data_only=True, read_only=False)
    try:
        if E_INVOICE_SHEET_NAME not in wb.sheetnames:
            fail(
                f"Sheet '{E_INVOICE_SHEET_NAME}' not found. Available: {', '.join(wb.sheetnames)}"
            )
        ws = wb[E_INVOICE_SHEET_NAME]
        headers_out: list[str] = []
        seen: set[str] = set()
        for h in read_header_row_cells(ws, header_row):
            key = normalize_invoice_header_label(h)
            if not key:
                continue
            if key not in seen:
                seen.add(key)
                headers_out.append(h)
    finally:
        wb.close()
    print(json.dumps({"headers": headers_out}))


def main() -> None:
    if len(sys.argv) < 2:
        fail("Missing command")
    command = sys.argv[1].strip().lower()
    args = sys.argv[2:]
    if command == "update_field":
        cmd_update_field(args)
        return
    if command == "update_number_field":
        cmd_update_number_field(args)
        return
    if command == "update_field_with_invoice":
        cmd_update_field_with_invoice(args)
        return
    if command == "update_issue_date":
        cmd_update_issue_date(args)
        return
    if command == "write_row_json":
        cmd_write_row_json(args)
        return
    if command == "write_rows_json":
        cmd_write_rows_json(args)
        return
    if command == "write_rows_json_from_file":
        cmd_write_rows_json_from_file(args)
        return
    if command == "write_regression_batch_from_file":
        cmd_write_regression_batch_from_file(args)
        return
    if command == "write_dropdown_batch":
        cmd_write_dropdown_batch(args)
        return
    if command == "write_dropdown_batch_from_file":
        cmd_write_dropdown_batch_from_file(args)
        return
    if command == "write_currency_exchange_batch_from_file":
        cmd_write_currency_exchange_batch_from_file(args)
        return
    if command == "apply_calculations":
        cmd_apply_calculations(args)
        return
    if command == "read_e_invoice_headers":
        cmd_read_e_invoice_headers(args)
        return
    if command == "apply_parallel_worker_identity":
        cmd_apply_parallel_worker_identity(args)
        return
    if command == "patch_invoice_cell":
        cmd_patch_invoice_cell(args)
        return
    if command == "patch_invoice_text_cell":
        cmd_patch_invoice_text_cell(args)
        return
    if command == "read_invoice_text_cell":
        cmd_read_invoice_text_cell(args)
        return
    fail(f"Unknown command: {command}")


if __name__ == "__main__":
    main()

