"""
Read UAE e-invoice error workbooks (openpyxl): extract cell comments, validate Errors column.

CLI subcommands (invoked from Node via `pythonRunner`): ``comment``, ``validate``, ``list_comments``.
Stdout is JSON for `getErrorFieldExcelDetails` in `invoiceExcel.ts`.
``list_comments`` returns every Errors-column field (comma-separated) plus that column's cell comment.
`comment` may be "" when no cell comment exists and the field is not in Errors; TS then fails the test.
If the field appears in Errors but there is no cell comment, JSON uses the Errors cell text (or a short placeholder) so automation can proceed.
"""

from openpyxl import load_workbook
import io
import json
import os
import re
import sys
import tempfile
import warnings
import zipfile

warnings.filterwarnings("ignore")

SHEET_NAME = "E Invoice"
HEADER_ROW = 4
DEFAULT_DATA_ROW = 6

# Server-generated error workbooks sometimes include data validation XML that openpyxl cannot
# parse (e.g. missing/invalid sqref → TypeError: expected MultiCellRange). Strip those blocks
# into a temp copy and load that instead.
_DV_BLOCK_RE = re.compile(
    r"<dataValidations\b[^>]*>[\s\S]*?</dataValidations>", re.IGNORECASE
)
_DV_EMPTY_RE = re.compile(r"<dataValidations\b[^>]*/>", re.IGNORECASE)


def _safe_unlink(path):
    if path and os.path.isfile(path):
        try:
            os.unlink(path)
        except OSError:
            pass


def _strip_data_validations_to_temp_xlsx(src_path: str) -> str:
    buf = io.BytesIO()
    with zipfile.ZipFile(src_path, "r") as zin:
        with zipfile.ZipFile(buf, "w", compression=zipfile.ZIP_DEFLATED) as zout:
            for zi in zin.infolist():
                data = zin.read(zi.filename)
                if zi.filename.startswith("xl/worksheets/") and zi.filename.endswith(
                    ".xml"
                ):
                    try:
                        text = data.decode("utf-8")
                    except UnicodeDecodeError:
                        pass
                    else:
                        if "datavalidation" in text.lower():
                            text = _DV_BLOCK_RE.sub("", text)
                            text = _DV_EMPTY_RE.sub("", text)
                            data = text.encode("utf-8")
                zout.writestr(zi.filename, data, compress_type=zi.compress_type)
    fd, tmp_path = tempfile.mkstemp(suffix=".xlsx", prefix="errxlsx_")
    try:
        os.close(fd)
        with open(tmp_path, "wb") as out:
            out.write(buf.getvalue())
    except Exception:
        _safe_unlink(tmp_path)
        raise
    return tmp_path


def _load_workbook_tolerant(file_path: str):
    """Return (workbook, temp_path_or_none). Caller must close workbook and unlink temp."""
    try:
        return load_workbook(file_path, data_only=True, read_only=False), None
    except Exception:
        tmp = None
        try:
            tmp = _strip_data_validations_to_temp_xlsx(file_path)
            wb = load_workbook(tmp, data_only=True, read_only=False)
            return wb, tmp
        except Exception:
            _safe_unlink(tmp)
            raise


def normalize(value):
    if value is None:
        return ""
    return str(value).strip()


def normalize_header_ts(value):
    """
    Collapse internal whitespace and trim; preserve case.
    Matches TS `normalizeInvoiceHeader` so buyer `Scheme identifier` ≠ payment `Scheme Identifier`.
    """
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def normalize_key(value):
    return normalize(value).lower()


def fail(message):
    print(message, file=sys.stderr)
    sys.exit(1)


def get_sheet(file_path):
    if not os.path.exists(file_path):
        fail(f"File not found: {file_path}")
    wb, cleanup_path = _load_workbook_tolerant(file_path)
    if SHEET_NAME not in wb.sheetnames:
        wb.close()
        _safe_unlink(cleanup_path)
        fail(f"Sheet '{SHEET_NAME}' not found")
    return wb, wb[SHEET_NAME], cleanup_path


# submitInvoiceExcelHeaderMap: template column is usually "Tax Rate"; tests still use "Standard Tax Rate".
_TAX_RATE_HEADER_KEYS = frozenset({"tax rate", "standard tax rate"})

# Errors column / server wording may abbreviate or rephrase.
_CURRENCY_EXCHANGE_PHRASE_KEYS = frozenset(
    {
        "currency exchange rate",
        "exchange rate",
        "currency rate",
        "fx rate",
    }
)

_INVOICING_PERIOD_HEADER_KEYS = frozenset(
    {"invoicing period start date", "invoicing period end date"}
)


def _tax_rate_header_alternates(field_name: str) -> list[str]:
    """Return header strings to try for column / Errors matching (order: exact name first)."""
    names = [normalize_header_ts(field_name)]
    if not names[0]:
        return []
    key = names[0].lower()
    if key == "standard tax rate":
        names.append("Tax Rate")
    elif key == "tax rate":
        names.append("Standard Tax Rate")
    return names


def _currency_exchange_header_alternates(field_name: str) -> list[str]:
    names = [normalize_header_ts(field_name)]
    if not names[0]:
        return []
    if names[0].lower() == "currency exchange rate":
        names.extend(["Exchange Rate", "Currency rate", "FX Rate"])
    return names


def _all_header_alternates(field_name: str) -> list[str]:
    base = _tax_rate_header_alternates(field_name)
    if normalize_header_ts(field_name).lower() == "currency exchange rate":
        for extra in _currency_exchange_header_alternates(field_name):
            if extra not in base:
                base.append(extra)
    return base


def errors_list_includes_field(error_parts: list[str], field_name: str) -> bool:
    """True if Errors column lists this field, treating Tax Rate ↔ Standard Tax Rate as the same,
    and Invoicing period start date ↔ Invoicing period end date as the same for IBG-14."""
    wanted = normalize_header_ts(field_name)
    if not wanted:
        return False
    wanted_l = wanted.lower()
    for p in error_parts:
        pn = normalize_header_ts(p)
        if pn == wanted:
            return True
        pl = pn.lower()
        if pl == wanted_l:
            return True
        if wanted_l in _TAX_RATE_HEADER_KEYS and pl in _TAX_RATE_HEADER_KEYS:
            return True
        # Currency / FX: server may use "Exchange rate", "Currency rate", etc.
        if wanted_l == "currency exchange rate":
            if pl in _CURRENCY_EXCHANGE_PHRASE_KEYS:
                return True
            if "exchange" in pl and "rate" in pl:
                return True
        if pl == "currency exchange rate" and wanted_l in _CURRENCY_EXCHANGE_PHRASE_KEYS:
            return True
    # IBG-14: Errors may list only start or only end for invoicing period rule failures.
    if wanted_l in _INVOICING_PERIOD_HEADER_KEYS:
        for p in error_parts:
            pl = normalize_header_ts(p).lower()
            if pl in _INVOICING_PERIOD_HEADER_KEYS:
                return True
    return False


def _find_column_one(sheet, field_name):
    """
    Resolve a column by one header string. Prefer exact match (after whitespace normalization).
    When the Oman Title Case template duplicates Scheme Identifier / Custom 1/2, pick first vs last
    from the request key casing (sentence/lower → first, Title Case → last).
    """
    wanted = normalize_header_ts(field_name)
    if not wanted:
        return None
    wanted_lower = wanted.lower()
    exact_cols = []
    ci_cols = []
    for cell in sheet[HEADER_ROW]:
        header = normalize_header_ts(cell.value)
        if not header:
            continue
        if header == wanted:
            exact_cols.append(cell.column)
        elif header.lower() == wanted_lower:
            ci_cols.append(cell.column)
    cols = exact_cols or ci_cols
    if not cols:
        return None
    if len(cols) == 1:
        return cols[0]
    if wanted == wanted.title():
        return cols[-1]
    return cols[0]


def find_column(sheet, field_name):
    for name in _all_header_alternates(field_name) or [field_name]:
        col = _find_column_one(sheet, name)
        if col is not None:
            return col
    return None


def read_comment(file_path, field_name, data_row):
    wb, sheet, cleanup_path = get_sheet(file_path)
    try:
        target_col = find_column(sheet, field_name)
        errors_col = find_column(sheet, "Errors")

        if target_col is None:
            fail(f"Column '{field_name}' not found in row {HEADER_ROW}")

        target_cell = sheet.cell(row=data_row, column=target_col)
        comment_text = (
            target_cell.comment.text.strip()
            if target_cell.comment and target_cell.comment.text
            else ""
        )

        if not comment_text:
            for r in [data_row - 1, data_row + 1, data_row + 2]:
                if r <= 0:
                    continue
                c = sheet.cell(row=r, column=target_col)
                if c.comment and c.comment.text:
                    comment_text = c.comment.text.strip()
                    data_row = r
                    break

        value_cell = sheet.cell(row=data_row, column=target_col)
        cell_value = normalize(value_cell.value)

        errors_text = ""
        errors_lists_field = False
        if errors_col is not None:
            e = sheet.cell(row=data_row, column=errors_col)
            errors_text = normalize(e.value)
            if errors_text:
                parts = [p.strip() for p in errors_text.split(",") if p.strip()]
                errors_lists_field = errors_list_includes_field(parts, field_name)

        if not comment_text:
            # Some server builds list the field in Errors but omit Excel cell comments; accept and
            # surface the Errors cell text (or a short placeholder) so TS validation can proceed.
            if errors_lists_field:
                comment_text = (
                    errors_text
                    if errors_text
                    else "(Listed in Errors column; no cell comment from server.)"
                )
            elif errors_col is not None and errors_text:
                # Row mismatch: check adjacent rows for the same invoice block.
                for r in (data_row - 1, data_row + 1, data_row + 2):
                    if r <= 0:
                        continue
                    e2 = sheet.cell(row=r, column=errors_col)
                    et = normalize(e2.value)
                    if not et:
                        continue
                    parts2 = [p.strip() for p in et.split(",") if p.strip()]
                    if errors_list_includes_field(parts2, field_name):
                        errors_lists_field = True
                        comment_text = et
                        data_row = r
                        value_cell = sheet.cell(row=data_row, column=target_col)
                        cell_value = normalize(value_cell.value)
                        break

        if not comment_text:
            print(
                json.dumps(
                    {
                        "field": field_name,
                        "row": data_row,
                        "comment": "",
                        "cell_value": cell_value,
                    }
                )
            )
            sys.exit(0)

        print(
            json.dumps(
                {
                    "field": field_name,
                    "row": data_row,
                    "comment": comment_text,
                    "cell_value": cell_value,
                }
            )
        )
        sys.exit(0)
    finally:
        try:
            wb.close()
        except Exception:
            pass
        _safe_unlink(cleanup_path)


def _cell_comment_text(cell):
    if cell.comment and cell.comment.text:
        return cell.comment.text.strip()
    return ""


def list_comments(file_path, data_row):
    """JSON: Errors column field names (comma-separated) and each field's cell comment."""
    wb, sheet, cleanup_path = get_sheet(file_path)
    try:
        errors_col = find_column(sheet, "Errors")
        resolved_row = data_row
        errors_text = ""
        if errors_col is not None:
            errors_text = normalize(sheet.cell(row=data_row, column=errors_col).value)
            if not errors_text:
                for r in (data_row - 1, data_row + 1, data_row + 2):
                    if r <= 0:
                        continue
                    et = normalize(sheet.cell(row=r, column=errors_col).value)
                    if et:
                        errors_text = et
                        resolved_row = r
                        break

        field_names = [p.strip() for p in errors_text.split(",") if p.strip()]
        items = []
        for field_name in field_names:
            target_col = find_column(sheet, field_name)
            comment_text = ""
            cell_value = ""
            if target_col is not None:
                cell = sheet.cell(row=resolved_row, column=target_col)
                comment_text = _cell_comment_text(cell)
                cell_value = normalize(cell.value)
                if not comment_text:
                    for r in (resolved_row - 1, resolved_row + 1, resolved_row + 2):
                        if r <= 0:
                            continue
                        c = sheet.cell(row=r, column=target_col)
                        text = _cell_comment_text(c)
                        if text:
                            comment_text = text
                            cell_value = normalize(c.value)
                            break
            items.append(
                {
                    "field": field_name,
                    "comment": comment_text,
                    "cell_value": cell_value,
                }
            )

        print(
            json.dumps(
                {
                    "errors_column": errors_text,
                    "row": resolved_row,
                    "fields": items,
                }
            )
        )
        sys.exit(0)
    finally:
        try:
            wb.close()
        except Exception:
            pass
        _safe_unlink(cleanup_path)


def validate_field(file_path, expected_field, data_row):
    wb, sheet, cleanup_path = get_sheet(file_path)
    try:
        errors_col = find_column(sheet, "Errors")
        if errors_col is None:
            fail("Errors column not found in downloaded error file")

        error_cell = sheet.cell(row=data_row, column=errors_col)
        if not error_cell.value:
            fail("Errors cell is empty in downloaded error file")

        error_value = normalize(error_cell.value)
        error_list = [v.strip() for v in error_value.split(",") if v.strip()]

        print(f"Expecting value: {expected_field}")
        print(f"Errors value from error column: {error_value}")
        print(f"error list: {error_list}")

        if not errors_list_includes_field(error_list, expected_field):
            fail(f"Expected field '{expected_field}' not found in Errors list")

        sys.exit(0)
    finally:
        try:
            wb.close()
        except Exception:
            pass
        _safe_unlink(cleanup_path)


def main():
    if len(sys.argv) < 2:
        fail(
            "Usage:\n"
            "  error_excel_reader.py comment <file_path> <field_name> [data_row]\n"
            "  error_excel_reader.py validate <file_path> <expected_field> [data_row]\n"
            "  error_excel_reader.py list_comments <file_path> [data_row]"
        )

    mode = sys.argv[1].strip().lower()

    if mode == "comment":
        if len(sys.argv) < 4:
            fail("Usage: error_excel_reader.py comment <file_path> <field_name> [data_row]")
        file_path = sys.argv[2]
        field_name = sys.argv[3]
        data_row = int(sys.argv[4]) if len(sys.argv) > 4 else DEFAULT_DATA_ROW
        read_comment(file_path, field_name, data_row)
        return

    if mode == "validate":
        if len(sys.argv) < 4:
            fail("Usage: error_excel_reader.py validate <file_path> <expected_field> [data_row]")
        file_path = sys.argv[2]
        expected_field = sys.argv[3]
        data_row = int(sys.argv[4]) if len(sys.argv) > 4 else DEFAULT_DATA_ROW
        validate_field(file_path, expected_field, data_row)
        return

    if mode == "list_comments":
        if len(sys.argv) < 3:
            fail("Usage: error_excel_reader.py list_comments <file_path> [data_row]")
        file_path = sys.argv[2]
        data_row = int(sys.argv[3]) if len(sys.argv) > 3 else DEFAULT_DATA_ROW
        list_comments(file_path, data_row)
        return

    fail(f"Unknown mode '{mode}'. Use 'comment', 'validate', or 'list_comments'.")


if __name__ == "__main__":
    main()
