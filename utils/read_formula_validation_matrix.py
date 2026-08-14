#!/usr/bin/env python3
"""Dump Oman formula-validation FullMatrix rows as JSON for the Excel pack generator."""
from __future__ import annotations

import json
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError as exc:  # pragma: no cover
    print(json.dumps({"ok": False, "error": f"openpyxl required: {exc}"}))
    sys.exit(1)


def main() -> int:
    if len(sys.argv) < 2:
        print(
            json.dumps(
                {
                    "ok": False,
                    "error": "Usage: read_formula_validation_matrix.py <xlsx> [out.json]",
                }
            )
        )
        return 1

    path = Path(sys.argv[1])
    out_path = Path(sys.argv[2]) if len(sys.argv) > 2 else None
    if not path.is_file():
        payload = {"ok": False, "error": f"File not found: {path}"}
        text = json.dumps(payload, ensure_ascii=False)
        if out_path:
            out_path.write_text(text, encoding="utf-8")
        else:
            print(text)
        return 1

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = (
            wb["All Testcases"]
            if "All Testcases" in wb.sheetnames
            else wb[wb.sheetnames[0]]
        )
        rows_iter = ws.iter_rows(values_only=True)
        headers = [str(c).strip() if c is not None else "" for c in next(rows_iter)]
        col = {h: i for i, h in enumerate(headers) if h}

        def cell(row, name: str) -> str:
            idx = col.get(name)
            if idx is None or idx >= len(row):
                return ""
            v = row[idx]
            return "" if v is None else str(v).strip()

        cases = []
        for row in rows_iter:
            if not row:
                continue
            tc_id = cell(row, "Test Case ID")
            if not tc_id:
                continue
            cases.append(
                {
                    "id": tc_id,
                    "priority": cell(row, "Priority"),
                    "polarity": cell(row, "Polarity").lower(),
                    "section": cell(row, "Section"),
                    "field": cell(row, "Filed name") or cell(row, "Field name"),
                    "taxCategory": cell(row, "Tax Category"),
                    "currency": cell(row, "Currency"),
                    "lines": cell(row, "Lines"),
                    "ruleId": cell(row, "RuleId"),
                    "title": cell(row, "Testcase Title"),
                    "description": cell(row, "Test Cases Description"),
                    "preconditions": cell(row, "Preconditions"),
                    "steps": cell(row, "Steps of Test case"),
                    "expected": cell(row, "Expected Results"),
                    "automationStatus": cell(row, "Automation status"),
                }
            )
    finally:
        wb.close()

    text = json.dumps({"ok": True, "cases": cases}, ensure_ascii=False)
    if out_path:
        out_path.parent.mkdir(parents=True, exist_ok=True)
        out_path.write_text(text, encoding="utf-8")
        print(json.dumps({"ok": True, "out": str(out_path), "count": len(cases)}))
    else:
        print(text)
    return 0


if __name__ == "__main__":
    # Windows consoles default to cp1252; matrix titles include minus/en-dash chars.
    try:
        sys.stdout.reconfigure(encoding="utf-8")  # type: ignore[attr-defined]
    except Exception:
        pass
    raise SystemExit(main())
