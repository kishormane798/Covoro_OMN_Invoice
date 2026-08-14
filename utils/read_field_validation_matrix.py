#!/usr/bin/env python3
"""Dump Oman field-validation matrix rows as JSON for the Excel pack generator."""
from __future__ import annotations

import json
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError as exc:  # pragma: no cover
    print(json.dumps({"ok": False, "error": f"openpyxl required: {exc}"}))
    sys.exit(1)


def main() -> int:
    if len(sys.argv) < 2:
        print(json.dumps({"ok": False, "error": "Usage: read_field_validation_matrix.py <xlsx>"}))
        return 1

    path = Path(sys.argv[1])
    if not path.is_file():
        print(json.dumps({"ok": False, "error": f"File not found: {path}"}))
        return 1

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb["All Testcases"] if "All Testcases" in wb.sheetnames else wb[wb.sheetnames[0]]
        rows_iter = ws.iter_rows(values_only=True)
        headers = [str(c).strip() if c is not None else "" for c in next(rows_iter)]
        col = {h: i for i, h in enumerate(headers) if h}

        def cell(row, name: str) -> str:
            idx = col.get(name)
            if idx is None or idx >= len(row):
                return ""
            v = row[idx]
            return "" if v is None else str(v).strip()

        cases = []
        for row in rows_iter:
            if not row:
                continue
            tc_id = cell(row, "Test Case ID")
            if not tc_id:
                continue
            cases.append(
                {
                    "id": tc_id,
                    "priority": cell(row, "Priority"),
                    "section": cell(row, "Section"),
                    "field": cell(row, "Filed name") or cell(row, "Field name"),
                    "title": cell(row, "Testcase Title"),
                    "description": cell(row, "Test Cases Description"),
                }
            )
    finally:
        wb.close()

    print(json.dumps({"ok": True, "cases": cases}, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
