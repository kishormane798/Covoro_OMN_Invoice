#!/usr/bin/env python3
"""In-place patch of identity cells on many workbooks (parallel)."""
from __future__ import annotations

import json
import os
import sys
from concurrent.futures import ProcessPoolExecutor, as_completed
from pathlib import Path

from openpyxl import load_workbook


def normalize_header(name: object) -> str:
    return " ".join(str(name if name is not None else "").split()).strip()


def resolve_header_column(ws, header_row: int, field_name: str) -> int:
    wanted = normalize_header(field_name)
    wanted_lower = wanted.lower()
    exact: list[int] = []
    ci: list[int] = []
    for cell in ws[header_row]:
        raw = normalize_header(cell.value)
        if not raw:
            continue
        if raw == wanted:
            exact.append(cell.column)
        elif raw.lower() == wanted_lower:
            ci.append(cell.column)
    cols = exact or ci
    if not cols:
        raise RuntimeError(f"Column not found: {field_name!r}")
    if wanted == wanted.title():
        return cols[-1]
    return cols[0]


def patch_file(args: tuple) -> dict:
    dest, patches, header_row, data_row, sheet_name = args
    try:
        wb = load_workbook(dest)
        ws = wb[sheet_name]
        for field, value in patches:
            col = resolve_header_column(ws, header_row, field)
            ws.cell(row=data_row, column=col).value = value
        wb.save(dest)
        wb.close()
        return {"ok": True, "dest": dest}
    except Exception as exc:  # noqa: BLE001
        return {"ok": False, "dest": dest, "error": str(exc)}


def main() -> int:
    payload = json.loads(Path(sys.argv[1]).read_text(encoding="utf-8"))
    jobs = payload.get("jobs") or []
    # Group patches by destPath
    by_dest: dict[str, list[tuple[str, str]]] = {}
    for job in jobs:
        by_dest.setdefault(job["destPath"], []).append((job["field"], job["value"]))

    workers = max(1, min(6, (os.cpu_count() or 4)))
    items = [
        (dest, patches, 4, 6, "E Invoice")
        for dest, patches in by_dest.items()
    ]
    ok = 0
    errors = []
    with ProcessPoolExecutor(max_workers=workers) as pool:
        futs = [pool.submit(patch_file, item) for item in items]
        for fut in as_completed(futs):
            result = fut.result()
            if result.get("ok"):
                ok += 1
            else:
                errors.append(result)

    print(json.dumps({"ok": len(errors) == 0, "files": ok, "errors": errors[:20]}))
    return 0 if not errors else 1


if __name__ == "__main__":
    raise SystemExit(main())
