#!/usr/bin/env python3
"""
Clone a base Oman invoice workbook and patch one cell per destination.

Uses a thin base (validations stripped) and a process pool. Each destination is
a fresh copy of the thin base, then load → patch → save once (openpyxl cannot
safely wb.save() the same workbook object repeatedly when images are present).
"""
from __future__ import annotations

import json
import os
import shutil
import sys
from concurrent.futures import ProcessPoolExecutor, as_completed
from pathlib import Path

from openpyxl import load_workbook


def normalize_header(name: object) -> str:
    return " ".join(str(name if name is not None else "").split()).strip()


def resolve_header_column(ws, header_row: int, field_name: str) -> int:
    wanted = normalize_header(field_name)
    if not wanted:
        raise RuntimeError("Empty field name for column resolution")
    wanted_lower = wanted.lower()
    exact: list[int] = []
    ci: list[int] = []
    for cell in ws[header_row]:
        raw = normalize_header(cell.value)
        if not raw:
            continue
        if raw == wanted:
            exact.append(cell.column)
        elif raw.lower() == wanted_lower:
            ci.append(cell.column)
    cols = exact or ci
    if not cols:
        raise RuntimeError(f"Column not found: {field_name!r}")
    if wanted == wanted.title():
        return cols[-1]
    return cols[0]


def strip_validations(ws) -> None:
    try:
        ws.data_validations.dataValidation = []
    except Exception:  # noqa: BLE001
        pass


def make_thin_base(base_path: str, thin_path: str) -> None:
    wb = load_workbook(base_path)
    for name in wb.sheetnames:
        strip_validations(wb[name])
    Path(thin_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(thin_path)
    wb.close()


def _xlsx_has_content_types(path: str) -> bool:
    import zipfile

    try:
        with zipfile.ZipFile(path) as zf:
            return "[Content_Types].xml" in zf.namelist()
    except Exception:  # noqa: BLE001
        return False


def worker_chunk(args: tuple) -> dict:
    thin_path, sheet_name, header_row, data_row, jobs = args
    # Resolve columns once; do not reuse this workbook for destination saves.
    wb_meta = load_workbook(thin_path)
    ws_meta = wb_meta[sheet_name]
    col_cache: dict[str, int] = {}
    for job in jobs:
        field = job["field"]
        if field not in col_cache:
            col_cache[field] = resolve_header_column(ws_meta, header_row, field)
    wb_meta.close()

    ok = 0
    errors: list[dict] = []
    for job in jobs:
        dest = job["destPath"]
        field = job["field"]
        value = "" if job.get("value") is None else str(job.get("value"))
        try:
            Path(dest).parent.mkdir(parents=True, exist_ok=True)
            shutil.copy2(thin_path, dest)
            wb = load_workbook(dest)
            try:
                ws = wb[sheet_name]
                ws.cell(row=data_row, column=col_cache[field]).value = value
                wb.save(dest)
            finally:
                wb.close()
            if not _xlsx_has_content_types(dest):
                raise RuntimeError(f"Corrupt workbook after save (missing Content_Types): {dest}")
            ok += 1
        except Exception as exc:  # noqa: BLE001
            try:
                Path(dest).unlink(missing_ok=True)
            except Exception:  # noqa: BLE001
                pass
            errors.append({"destPath": dest, "field": field, "error": str(exc)})
    return {"written": ok, "errors": errors}


def main() -> int:
    if len(sys.argv) < 2:
        print(json.dumps({"ok": False, "error": "Usage: batch_clone_patch_invoice.py <jobs.json>"}))
        return 1

    jobs_path = Path(sys.argv[1])
    payload = json.loads(jobs_path.read_text(encoding="utf-8"))
    base_path = payload["basePath"]
    sheet_name = payload.get("sheetName", "E Invoice")
    header_row = int(payload.get("headerRow", 4))
    data_row = int(payload.get("dataRow", 6))
    jobs = payload.get("jobs") or []
    workers = int(payload.get("workers") or max(1, min(6, (os.cpu_count() or 4))))
    thin_path = payload.get("thinBasePath") or str(
        Path(base_path).with_name(Path(base_path).stem + ".thin.xlsx")
    )

    if not os.path.isfile(base_path):
        print(json.dumps({"ok": False, "error": f"Base not found: {base_path}"}))
        return 1

    if not os.path.isfile(thin_path):
        make_thin_base(base_path, thin_path)

    if not jobs:
        print(json.dumps({"ok": True, "written": 0, "errors": [], "thinBase": thin_path}))
        return 0

    # Split jobs across workers.
    chunks: list[list] = [[] for _ in range(min(workers, len(jobs)))]
    for i, job in enumerate(jobs):
        chunks[i % len(chunks)].append(job)

    written = 0
    errors: list[dict] = []
    with ProcessPoolExecutor(max_workers=len(chunks)) as pool:
        futures = [
            pool.submit(
                worker_chunk,
                (thin_path, sheet_name, header_row, data_row, chunk),
            )
            for chunk in chunks
            if chunk
        ]
        for fut in as_completed(futures):
            result = fut.result()
            written += int(result.get("written") or 0)
            errors.extend(result.get("errors") or [])

    print(
        json.dumps(
            {
                "ok": len(errors) == 0,
                "written": written,
                "errors": errors,
                "thinBase": thin_path,
                "workers": len(chunks),
            }
        )
    )
    return 0 if not errors else 1


if __name__ == "__main__":
    raise SystemExit(main())
