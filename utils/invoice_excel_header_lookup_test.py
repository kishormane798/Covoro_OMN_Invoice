"""Failing-first checks: header lookup must not walk data-validation XFD width."""
from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation

_UTILS = Path(__file__).resolve().parent
if str(_UTILS) not in sys.path:
    sys.path.insert(0, str(_UTILS))

from invoice_excel_writer import (  # noqa: E402
    INVOICE_SUBMIT_CLEAR_LAST_COL,
    _header_lookup,
    _set_by_header_name,
)


def _wide_sheet():
    wb = Workbook()
    ws = wb.active
    ws.title = "E Invoice"
    ws.cell(row=4, column=1, value="Invoice Number")
    ws.cell(row=4, column=119, value="Custom 5")
    ws.cell(row=4, column=16384, value="trap-xfd")
    dv = DataValidation(type="list", formula1='"A"')
    dv.add("A4:XFD4")
    ws.add_data_validation(dv)
    return wb, ws


def test_header_lookup_caps_data_validation_width() -> None:
    _wb, ws = _wide_sheet()
    assert ws.max_column >= 16384
    exact, _lower, last_col = _header_lookup(ws, 4)
    assert last_col <= 256, last_col
    assert last_col >= INVOICE_SUBMIT_CLEAR_LAST_COL
    assert "Invoice Number" in exact
    assert "Custom 5" in exact
    assert "trap-xfd" not in exact


def test_set_by_header_name_uses_capped_lookup() -> None:
    _wb, ws = _wide_sheet()
    _set_by_header_name(ws, 4, 6, "Invoice Number", "INV-1")
    assert ws.cell(row=6, column=1).value == "INV-1"


if __name__ == "__main__":
    test_header_lookup_caps_data_validation_width()
    test_set_by_header_name_uses_capped_lookup()
    print("ok")
