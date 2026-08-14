#!/usr/bin/env python3
"""List unmapped conditional FullMatrix rules for alias suggestions."""
from __future__ import annotations

import re
from collections import Counter, defaultdict
from pathlib import Path

import openpyxl

ALIASES = {
    "Preceding invoice reference": "Preceding Invoice reference",
    "Invoice transaction type": "Invoice Transaction Type Code",
    "Invoice Transaction Type": "Invoice Transaction Type Code",
    "If invoice transaction type": "Invoice Transaction Type Code",
    "If Invoice transaction type": "Invoice Transaction Type Code",
    "When Invoice transaction type": "Invoice Transaction Type Code",
    # Dropdown description → Invoice Transaction Type Code (not Peppol bit-string).
    "Simplified Tax Invoice": "Invoice Transaction Type Code",
    # Peppol BTOM-001 = Invoice Transaction Type.
    "BTOM-001": "Invoice Transaction Type Code",
    "Invoice type code": "Invoice Type Code",
    "If Invoice type code": "Invoice Type Code",
    "if Invoice type code": "Invoice Type Code",
    "If Invoice Type code": "Invoice Type Code",
    "Currency exchange rate": "Currency Exchange Rate",
    "BTOM-003": "Currency Exchange Rate",
    "When Invoice currency code": "Invoice Currency Code",
    "Invoice currency code": "Invoice Currency Code",
    "If Invoice currency code": "Invoice Currency Code",
    "IBT-005": "Invoice Currency Code",
    # IBT-006 Tax Accounting Currency — backend default (intentional skip; do not alias).
    "If the TAX category code for tax category tax amount in accounting currency":
        "Invoice Total Tax Amount In Tax Accounting Currency",
    # IBR-065 / IBT-110 / IBT-111 (IBT-006 itself stays skip)
    "IBT-110": "Invoice Total Tax Amount",
    "IBT-111": "Invoice Total Tax Amount In Tax Accounting Currency",
    "Invoice Total VAT Amount in Tax Accounting Currency":
        "Invoice Total Tax Amount In Tax Accounting Currency",
    "Invoice Total Tax Amount in Accounting Currency":
        "Invoice Total Tax Amount In Tax Accounting Currency",
    "Invoice Total Tax Amount in Tax Accounting Currency":
        "Invoice Total Tax Amount In Tax Accounting Currency",
    # IBR-082
    "BTOM-020": "Total Amount Due (Profit Margin)",
    "Total Amount Due": "Total Amount Due (Profit Margin)",
    "BTOM-017": "Total Amount Including VAT",
    # ALIGNED-IBRP-*-08/09 + IBR-168 — breakdown amounts → closest document totals
    "IBT-116": "Invoice Total Amount Without Tax",
    "VAT Category Taxable Amount": "Invoice Total Amount Without Tax",
    "IBT-117": "Invoice Total Tax Amount",
    "VAT Category Tax Amount": "Invoice Total Tax Amount",
    "IBT-118": "Tax Category",
    "IBT-151": "Tax Category",
    "VAT Category Code": "Tax Category",
    "IBT-131": "Invoice Line Net Amount",
    "IBT-092": "Allowances On Document Level",
    "IBT-099": "Charges On Document Level",
    "IBT-095": "VAT Category - Allowances",
    "IBT-102": "VAT Category - Charges",
    "IBT-119": "Tax Rate",
    "IBT-152": "Tax Rate",
    "IBT-103": "Tax Rate",
    "VAT Category Rate": "Tax Rate",
    "Document Level Charge Tax Rate": "Tax Rate",
    "Document Level Allowance Tax Rate": "Tax Rate",  # only if not field-skipped
    "Exchange Rate": "Currency Exchange Rate",
    "BTOM-016": "Line Item VAT Amount",
    # IBR-035 line allowance/charge amounts
    "IBT-137": "Invoice Line Allowance Amount",
    "IBT-142": "Invoice Line Charge Amount",
    "Invoice Line Allowance/Charge Base Amount (IBT-137 / IBT-142)":
        "Invoice Line Allowance Amount",
    # IBR-046 multi-rate Filed name → Tax Rate (IBT-096 has no separate column)
    "VAT Rate (IBT-096, IBT-103, IBT-119, IBT-152, IBT-193)": "Tax Rate",
    # Compound ALIGNED-*-08 Filed name
    "Invoice Transaction Type (BTOM-001), VAT Category Code (IBT-118), VAT Category Taxable Amount (IBT-116)":
        "Invoice Total Amount Without Tax",
    "Item Type": "Item Type",
    "When Item type": "Item Type",
    "Item classification identifier": "Item classification identifier",
    "Industrial Classification Code must be provided for each ITEM INFORMATION":
        "Industrial Classification Code",
    "Delivery to Country code": "Deliver To Country Code",
    "Deliver to Country code": "Deliver To Country Code",
    "Deliver to address line 1 - Postal code": "Deliver to post code",
    "Customs Declaration number": "Customs Declaration Number",
    "Import date": "Import Date",
    "Invoicing period start date": "Invoicing period start date",
    "Invoicing period Start date": "Invoicing period start date",
    "Invoicing period end date": "Invoicing period end date",
    "Invoicing period End date": "Invoicing period end date",
    # Invoice line period (IBT-134/135) — no Covoro columns; intentional skip in TS helper.
    "Seller identifier": "Seller Identifier",
    "Seller Identifier": "Seller Identifier",
    "IBT-029": "Seller Identifier",
    "Seller Identifier Scheme Identifier": "Seller Identifier - Scheme Identifier",
    "IBT-029-1": "Seller Identifier - Scheme Identifier",
    "Seller Identifier (IBT-029) – Scheme Identifier (IBT-029-1)": "Seller Identifier",
    "Buyer identifier": "Buyer Identifier",
    "Buyer Identifier": "Buyer Identifier",
    "IBT-046": "Buyer Identifier",
    "the value in the Buyer identifier": "Buyer Identifier",
    "Buyer Identifier Scheme Identifier": "Scheme Identifier",
    "IBT-046-1": "Scheme Identifier",
    "Scheme Identifier": "Scheme Identifier",
    "Buyer Identifier (IBT-046), Buyer Identifier Scheme Identifier (IBT-046-1)":
        "Buyer Identifier",
    "Buyer Identifier (IBT-046) / Scheme Identifier (IBT-046-1)": "Buyer Identifier",
    "If Buyer electronic address": "Buyer electronic address",
    "Buyer electronic address": "Buyer electronic address",
    # CL-11 — Profit Margin Item Type (BTOM-025)
    "BTOM-025": "Profit Margin Item Type Code",
    "Profit Margin Item Type Code": "Profit Margin Item Type Code",
    "Profit Margin Item Reason Code": "Profit Margin Item Type Code",
    "Tax Category": "Tax Category",
    "VAT Category": "Tax Category",
    "VAT category code": "Tax Category",
    "Invoiced item VAT category code": "Tax Category",
    "Invoiced item Tax Category": "Tax Category",
    "Tax Rate": "Tax Rate",
    "VAT Rate": "Tax Rate",
    "Invoiced item VAT rate": "Tax Rate",
    "Invoice item VAT rate": "Tax Rate",
    # FullMatrix Filed name for ALIGNED-IBRP-E/S/Z-05-OM (Invoiced item VAT rate / IBT-152).
    "In an Invoice line": "Tax Rate",
    # FullMatrix Filed name for ALIGNED-IBRP-O-05-OM (Invoiced item VAT rate / IBT-152).
    "An Invoice line": "Tax Rate",
    # CL-10 exemption reason codes
    "Tax exemption reason code": "Tax Exemption Reason Code",
    "Tax Exemption Reason Code": "Tax Exemption Reason Code",
    "VAT Exemption Reason Code": "Tax Exemption Reason Code",
    "IBT-121": "Tax Exemption Reason Code",
    "IBT-186": "Tax Exemption Reason Code",
    "IBT-196": "Tax Exemption Reason - Allowances",
    "IBT-198": "Tax Exemption Reason - Charges",
    "Document Level Allowance VAT Exemption Reason Code":
        "Tax Exemption Reason - Allowances",
    "Document Level Allowance TAX Exemption Reason Code":
        "Tax Exemption Reason - Allowances",
    "Document Level Charge VAT Exemption Reason Code":
        "Tax Exemption Reason - Charges",
    "Document Level Charge TAX Exemption Reason Code":
        "Tax Exemption Reason - Charges",
    "Document Level Allowance VAT Exemption Reason Code (IBT-196) / Document Level Allowance VAT Category Code (IBT-095)":
        "Tax Exemption Reason - Allowances",
    "Document Level Allowance TAX Exemption Reason Code (IBT-196) / Document Level Allowance TAX Category Code (IBT-095)":
        "Tax Exemption Reason - Allowances",
    "Document Level Charge VAT Exemption Reason Code (IBT-198) / Document Level Charge VAT Category Code (IBT-102)":
        "Tax Exemption Reason - Charges",
    "VAT Exemption Reason Code (IBT-121) / VAT Exemption Reason Text (IBT-120)":
        "Tax Exemption Reason Code",
    "TAX category tax amount in accounting currency":
        "Invoice Total Tax Amount In Tax Accounting Currency",
    "Tax category tax amount in accounting currency":
        "Invoice Total Tax Amount In Tax Accounting Currency",
    "IBT-190": "Invoice Total Tax Amount In Tax Accounting Currency",
    # IBT-192 Accounting Currency VAT Category Code — no Covoro column (field skip).
    # Amount prose below still maps to tax-in-accounting-currency total.
    "IBT-192(tax category tax amount in accounting currency)":
        "Invoice Total Tax Amount In Tax Accounting Currency",
    # Line-level VAT amount (BTOM-016)
    "In Line VAT information": "Line Item VAT Amount",
    "Each Invoice/CreditNote line must contain Item VAT Amount":
        "Line Item VAT Amount",
    "Line Item VAT Amount": "Line Item VAT Amount",
    "Item VAT Amount": "Line Item VAT Amount",
    "Seller tax identifier": "Seller VAT Identifier (TRN / TIN)",
    "Seller Tax Identifier": "Seller VAT Identifier (TRN / TIN)",
    "Buyer VATIN": "Buyer VAT Identifier",
    "If Document level charge TAX category code": "Vat category - charges",
    "Document level charge TAX category code": "Vat category - charges",
    "If Document level allowance TAX category code": "Vat category - allowances",
    "Document level allowance TAX category code": "Vat category - allowances",
    "Document level allowance": "Allowances on document level",
    "Document level charge": "Charges on document level",
    # FullMatrix Filed name for ALIGNED-IBRP-*-01-OM (IBG-25 + IBG-20/IBG-21 prose).
    # Primary → Allowances; pack helper dual-patches Charges as companion.
    "An Invoice that contains an Invoice line": "Allowances on document level",
    "Charge base amount (IBT-142)": "Charges on document level",
    "Charge base amount": "Charges on document level",
    "Allowance base amount (IBT-137)": "Allowances on document level",
    "Allowance base amount (IBT-093)": "Allowances on document level",
    "Allowance base amount": "Allowances on document level",
    "VAT breakdown(IBG-23)": "Invoice Total Tax Amount",
    "In a VAT breakdown": "Invoice Total Tax Amount",
    # ALIGNED-IBRP-*-01 — breakdown group Filed → Tax Category (drives derived breakdown)
    "VAT Breakdown Group (IBG-23) – VAT Category Code (IBT-118)": "Tax Category",
    "VAT Breakdown Group (IBG-23) - VAT Category Code (IBT-118)": "Tax Category",
    "VAT Breakdown (IBG-23) – VAT Category Code (IBT-118)": "Tax Category",
    "VAT Breakdown (IBG-23) - VAT Category Code (IBT-118)": "Tax Category",
    # Invoicing period (IBR-036/037)
    "IBT-073": "Invoicing Period Start Date",
    "IBT-074": "Invoicing Period End Date",
    "Invoicing Period Start Date": "Invoicing Period Start Date",
    "Invoicing Period End Date": "Invoicing Period End Date",
    "Invoicing Period Start Date (IBT-073) – Invoicing Period End Date (IBT-074)":
        "Invoicing Period Start Date",
    "Invoicing Period Start Date (IBT-073) - Invoicing Period End Date (IBT-074)":
        "Invoicing Period Start Date",
    # Invoice line period (IBR-030 / IBR-CO-20) — template has no line-period columns;
    # map to document Invoicing Period* (closest Covoro headers).
    "IBT-134": "Invoicing Period Start Date",
    "IBT-135": "Invoicing Period End Date",
    "Invoice Line Period Start Date": "Invoicing Period Start Date",
    "Invoice Line Period End Date": "Invoicing Period End Date",
    "Invoice Line Period Start Date (IBT-134)": "Invoicing Period Start Date",
    "Invoice Line Period End Date (IBT-135)": "Invoicing Period End Date",
    "Invoice Line Period Start Date (IBT-134) / Invoice Line Period End Date (IBT-135)":
        "Invoicing Period Start Date",
    "Invoice Line Period Start Date (IBT-134) / Invoice Line Period End Date (IBT-135)":
        "Invoicing Period Start Date",
    # ALIGNED-IBRP-048 / IBT-119
    "VAT Category Rate (IBT-119)": "Tax Rate",
    "IBT-119": "Tax Rate",
    "IBT-118": "Tax Category",
    # IBR-053
    "Invoice Total Tax Amount in Accounting Currency (IBT-111)":
        "Invoice Total Tax Amount In Tax Accounting Currency",
    "IBG-31": "Industrial Classification Code",
    "Item Classification Identifier (HS Code) (IBT-158)":
        "Item Classification Identifier",
    # ALIGNED-IBRP-*-05 line VAT category + rate
    "Invoiced Item VAT Category Code": "Tax Category",
    "Invoiced Item VAT Category Code (IBT-151) – Invoice Item VAT Rate (IBT-152)":
        "Tax Category",
    "Invoiced Item VAT Category Code (IBT-151) - Invoice Item VAT Rate (IBT-152)":
        "Tax Category",
    "Invoiced Item VAT Category Code (IBT-151) – Invoiced Item VAT Rate (IBT-152)":
        "Tax Category",
    "Invoiced Item VAT Category Code (IBT-151) - Invoiced Item VAT Rate (IBT-152)":
        "Tax Category",
    "Invoiced Item VAT Category Code (IBT-151) – Line Item VAT Amount (BTOM-016)":
        "Line Item VAT Amount",
    "Invoiced Item VAT Category Code (IBT-151) - Line Item VAT Amount (BTOM-016)":
        "Line Item VAT Amount",
    # Delivery / export
    "IBT-080": "Deliver To Country Code",
    "Delivery to Country Code": "Deliver To Country Code",
    "Deliver to Country Code": "Deliver To Country Code",
    "Deliver To Country Code": "Deliver To Country Code",
    # Seller / buyer tax ids
    "IBT-031": "Seller VAT Identifier (TRN / TIN)",
    "Seller Tax Identifier": "Seller VAT Identifier (TRN / TIN)",
    "IBT-048": "Buyer VAT Identifier",
    "Buyer VATIN": "Buyer VAT Identifier",
    "Buyer Identifier (IBT-046), Buyer VATIN (IBT-048)": "Buyer Identifier",
    # Country subdivision (IBR-150)
    "BTOM-026": "Buyer Country Subdivision Code",
    "BTOM-024": "Seller Country Subdivision Code",
    "Buyer Country Subdivision Code": "Buyer Country Subdivision Code",
    "Seller Country Subdivision Code": "Seller Country Subdivision Code",
    "Buyer Country Subdivision Code (BTOM-026) / Seller Country Subdivision Code (BTOM-024)":
        "Buyer Country Subdivision Code",
    # Credit / debit reason (IBR-023)
    "IBT-003": "Invoice Type Code",
    "BTOM-032": "Credit Note Or Debit Note Reason Code",
    "Credit Note / Debit Note Reason Code": "Credit Note Or Debit Note Reason Code",
    "Credit Note / Debit Note Reason Code (BTOM-032)":
        "Credit Note Or Debit Note Reason Code",
    "Credit note or Debit Note reason code": "Credit Note Or Debit Note Reason Code",
    # Document charge (IBR-042) — no separate Charge Reason Code column
    "IBT-104": "Charges On Document Level",
    "Document Level Charge Reason Code": "Charges On Document Level",
    "Document Level Charge": "Charges On Document Level",
    "Document Level Charge Reason Code (IBT-104) / Document Level Charge (IBG-21)":
        "Charges On Document Level",
    # Preceding invoice (IBR-032 / IBR-175) — BTOM-031 → Unique Identifier Number
    "IBT-025": "Preceding Invoice Reference",
    "IBT-026": "Preceding Invoice Issue Date",
    "BTOM-031": "Unique Identifier Number",
    "Preceding Invoice UUID": "Unique Identifier Number",
    "Preceding Invoice Reference (IBT-025), Preceding Invoice Issue Date (IBT-026), Preceding Invoice UUID (BTOM-031)":
        "Preceding Invoice Reference",
    "Invoice Transaction Type (BTOM-001) – Preceding Invoice Reference (IBT-025), Preceding Invoice UUID (BTOM-031)":
        "Invoice Transaction Type Code",
    "Invoice Transaction Type (BTOM-001) - Preceding Invoice Reference (IBT-025), Preceding Invoice UUID (BTOM-031)":
        "Invoice Transaction Type Code",
    # Item classification / type (IBR-174 / IBR-091)
    "BTOM-019": "Item Type",
    "IBT-158": "Item Classification Identifier",
    "Item Classification Identifier": "Item Classification Identifier",
    "Item Classification Identifier (HS Code)": "Item Classification Identifier",
    "Item Classification Identifier (HS Code) (IBT-158)": "Item Classification Identifier",
    "Item Classification Identifier (IBT-158) – Invoice Transaction Type (BTOM-001)":
        "Item Classification Identifier",
    "Item Classification Identifier (IBT-158) - Invoice Transaction Type (BTOM-001)":
        "Item Classification Identifier",
    # IBR-104 — rate drives; VAT Accounting Currency remains field-skip
    "VAT Category Rate (IBT-119) – VAT Accounting Currency": "Tax Rate",
    "VAT Category Rate (IBT-119) - VAT Accounting Currency": "Tax Rate",
    # Import details (IBR-085)
    "Import Details (IBG-33-OM)": "Import Date",
    "Import Details": "Import Date",
    "BTOM-021": "Customs Declaration Number",
    "BTOM-022": "Incoterms",
    "Customs Declaration Number": "Customs Declaration Number",
    # Note: matrix labels Import Date as BTOM-020 here (not profit-margin BTOM-020).
    "Import Date (BTOM-020)": "Import Date",
    # Prepayment / paid (IBR-058) + IBR-093 (non-OM): IBT-113 ≠ IBT-180
    "IBT-180": "Paid Amount",
    "Paid Amount": "Paid Amount",
    "IBT-113": "Invoice Total Amount With Tax",
    "Total Paid Amount": "Invoice Total Amount With Tax",
    "Total Paid Amount (IBT-113)": "Invoice Total Amount With Tax",
    "BTOM-027": "Prepayment Invoice Number",
    "BTOM-014": "Prepayment Invoice Uuid",
    "Prepayment Invoice Number": "Prepayment Invoice Number",
    "Prepayment Invoice UUID": "Prepayment Invoice Uuid",
    "Prepayment Invoice Uuid": "Prepayment Invoice Uuid",
    "Prepayment Invoice Number (BTOM-027), Prepayment Invoice UUID (BTOM-014)":
        "Prepayment Invoice Number",
    # Supporting documents (IBR-013)
    "IBT-122": "Supporting Document Reference",
    "BTOM-023": "Supporting Document Uuid",
    "Supporting document reference": "Supporting Document Reference",
    "Supporting Document Reference": "Supporting Document Reference",
    "Supporting document UUID": "Supporting Document Uuid",
    "Supporting Document Uuid": "Supporting Document Uuid",
    "Supporting document reference (IBT-122) and Supporting document UUID (BTOM-023)":
        "Supporting Document Reference",
    # Seller address (IBR-010)
    "Seller Postal Address": "Seller Address Line 1",
    "Seller postal address": "Seller Address Line 1",
    "Seller Address Line 1": "Seller Address Line 1",
    "Seller Address Line 2": "Seller Address Line 2",
    "Seller Address Line 3": "Seller Address Line 3",
    "Seller City": "Seller City",
    "Seller Postal Code": "Seller Post Code",
    "Seller Post Code": "Seller Post Code",
    "IBT-035": "Seller Address Line 1",
    "IBT-036": "Seller Address Line 2",
    "IBT-162": "Seller Address Line 3",
    "IBT-037": "Seller City",
    "IBT-038": "Seller Post Code",
}

# Keep in sync with Helpers/conditionalValidationExcelPackHelper.ts
PSEUDO = re.compile(
    r"^(in a vat breakdown|conditional fields|an invoice that contains|"
    r"in an invoice line|in line vat|the vat category tax amount(?!\s+in accounting)|"
    r"related invoice|vat breakdown)",
    re.I,
)

# Default / formula suite / out-of-scope — do not suggest aliases or generate Excels.
# Keep field skips in sync with Helpers/conditionalValidationExcelPackHelper.ts
INTENTIONAL_SKIP = {
    # IBT-006 — set by backend; no need to enter in Excel.
    "vat accounting currency",
    "tax accounting currency",
    "tax accounting currency code",
    "ibt-006",
    # IBT-134/135: no longer skipped — aliased to document Invoicing Period*
    # (template has no Invoice Line Period columns).
    # IBT-096 — no separate Allowance VAT Rate / percentage column on Covoro template.
    "document level allowance tax rate",
    # IBT-138 / IBT-143 — no separate line allowance/charge percentage columns.
    "invoice line allowance percentage",
    "invoice line charge percentage",
    "for each different value of vat category rate",
    # IBT-192 Accounting Currency VAT Category Code — no Covoro column.
    "accounting currency vat category code",
    "accounting currency vat category",
}
# Keep in sync with Helpers/conditionalValidationExcelPackHelper.ts
INTENTIONAL_SKIP_RULE_IDS = {
    # FORMULA suite — user: keep skip (not needed)
    "IBR-033-OM",
    "IBR-041-OM",
    # BACKEND
    "IBR-066-OM",
    "IBR-096-OM",
    "IBR-097-OM",
    # User batch: not needed / backend / default
    "IBR-073-OM",
    "IBR-074-OM",
    "IBR-173-OM",
    "IBR-059-OM",
}
RULE_RE = re.compile(r"\[([A-Z0-9][A-Z0-9\-]*OM[^\]]*)\]", re.I)
# New FullMatrix Filed names often append Peppol ids, e.g. "Preceding Invoice Reference (IBG-03)".
PEPPOL_SUFFIX_RE = re.compile(
    r"\s*\((?:IBT|IBG|BT|BTOM)-?[0-9A-Za-z\-]+\)\s*$", re.I
)


def norm(s: str) -> str:
    return re.sub(r"\s+", " ", s or "").strip().lower()


def strip_peppol_suffix(field: str) -> str:
    return PEPPOL_SUFFIX_RE.sub("", field or "").strip()


def is_invoice_line_period_field(field: str) -> bool:
    """True when Filed name refers to invoice line period (IBT-134/135 / IBG-26).

    No longer an intentional skip — aliased to document Invoicing Period*
    (template has no separate Invoice Line Period columns).
    Kept for reports / detection only.
    """
    raw = norm(field)
    return (
        "invoice line period" in raw
        or "ibt-134" in raw
        or "ibt-135" in raw
        or "ibg-26" in raw
    )


def is_tax_accounting_currency_field(field: str) -> bool:
    """IBT-006 — backend default; no need to enter in Excel (standalone labels only)."""
    f = (field or "").strip()
    raw = norm(f)
    base = norm(strip_peppol_suffix(f) or f)
    # Standalone labels only — compound multi-field rows stay for other aliases.
    if any(sep in f for sep in (",", "/", "–")):
        return False
    return base in {
        "vat accounting currency",
        "tax accounting currency",
        "tax accounting currency code",
        "ibt-006",
    } or raw == "ibt-006"


def is_accounting_currency_vat_category_field(field: str) -> bool:
    """IBT-192 Accounting Currency VAT Category Code — no Covoro column."""
    raw = norm(field)
    base = norm(strip_peppol_suffix(field) or field)
    if "tax amount" in raw or "tax amount" in base:
        return False  # amount prose maps elsewhere
    return (
        "accounting currency vat category" in raw
        or "accounting currency vat category" in base
        or (
            "ibt-192" in raw
            and "category" in raw
            and "amount" not in raw
        )
    )


def is_intentional_skip_field(field: str) -> bool:
    """True when field is not Excel-mappable by design (not an alias gap)."""
    # Line period (IBT-134/135) is aliased to Invoicing Period* — not a skip.
    if is_tax_accounting_currency_field(field):
        return True
    if is_accounting_currency_vat_category_field(field):
        return True
    f = (field or "").strip()
    base = strip_peppol_suffix(f) or f
    return any(norm(c) in INTENTIONAL_SKIP for c in (f, base))


def cell(row: tuple, col: dict[str, int], *names: str) -> str:
    """Read first matching column; supports GSP-54396 renames."""
    for name in names:
        idx = col.get(name)
        if idx is None or idx >= len(row):
            continue
        v = row[idx]
        if v is None:
            continue
        return str(v).strip()
    return ""


def main() -> None:
    root = Path(__file__).resolve().parents[1]
    template = root / "testData" / "uploads" / "template.xlsx"
    matrix = (
        root
        / "testcase"
        / "conditional_validation"
        / "EINV_OMAN_ConditionalValidation_FullMatrix.xlsx"
    )
    out = (
        root
        / "testcase"
        / "conditional_validation"
        / "UNMAPPED_RULES_FOR_ALIAS.md"
    )

    wb = openpyxl.load_workbook(template, read_only=True, data_only=True)
    ws = wb["E Invoice"]
    row4 = list(next(ws.iter_rows(min_row=4, max_row=4, values_only=True)))
    headers = [str(h).strip() for h in row4 if h]
    header_norm = {norm(h): h for h in headers}
    wb.close()

    def mappable(field: str, title: str = "", description: str = "") -> bool:
        f = (field or "").strip()
        if not f:
            return False
        base = strip_peppol_suffix(f) or f
        candidates = [f, base] if base != f else [f]
        # Matrix often uses en-dash; aliases may use ASCII hyphen.
        expanded: list[str] = []
        for c in candidates:
            expanded.append(c)
            ascii_dash = c.replace("–", "-").replace("—", "-")
            if ascii_dash != c:
                expanded.append(ascii_dash)
        candidates = expanded
        m_rule = RULE_RE.search(title or "")
        if m_rule and m_rule.group(1).strip().upper() in INTENTIONAL_SKIP_RULE_IDS:
            return True  # intentional skip (formula suite / out of scope)
        blob = f"{title}\n{description}".lower()
        if (
            any(norm(c) == "conditional fields" for c in candidates)
            and (
                "invoice line period" in blob
                or "ibt-134" in blob
                or "ibt-135" in blob
                or "ibg-26" in blob
            )
        ):
            return True  # intentional skip — no Covoro column (backend default)
        if is_intentional_skip_field(f):
            return True  # intentional skip — backend default / formula / out of scope
        # ALIGNED-IBRP-O/Z-01 Simplified Tax Invoice exception → Invoice Transaction Type Code.
        if (
            any(norm(c) == "an invoice that contains an invoice line" for c in candidates)
            and "simplified tax invoice exception" in blob
        ):
            return True
        if any(c in ALIASES for c in candidates):
            return True
        if any(PSEUDO.search(c) for c in candidates):
            return False
        for c in candidates:
            aliased = ALIASES.get(c, c)
            if norm(aliased) in header_norm:
                return True
        return False

    wb = openpyxl.load_workbook(matrix, read_only=True, data_only=True)
    ws = wb["All Testcases"]
    rows = ws.iter_rows(values_only=True)
    headers_m = [str(c).strip() if c is not None else "" for c in next(rows)]
    col = {h: i for i, h in enumerate(headers_m) if h}
    if "Test Case ID" not in col:
        raise KeyError(
            f"FullMatrix missing 'Test Case ID'. Found headers: {headers_m}"
        )
    if not any(h in col for h in ("Test Case Title", "Testcase Title", "Test Cases Title")):
        raise KeyError(
            "FullMatrix missing title column "
            "('Test Case Title' / 'Testcase Title'). "
            f"Found headers: {headers_m}"
        )

    by_rule: dict[str, dict[str, dict]] = defaultdict(
        lambda: defaultdict(lambda: {"count": 0, "sections": Counter(), "polarities": Counter()})
    )
    field_totals: Counter[str] = Counter()

    for row in rows:
        if not row:
            continue
        tc_id = cell(row, col, "Test Case ID")
        if not tc_id or tc_id.lower() == "test case id":
            continue  # blank or duplicate header row in new packs
        field = cell(row, col, "Filed name", "Field name")
        title = cell(row, col, "Test Case Title", "Testcase Title", "Test Cases Title")
        desc = cell(row, col, "Test Cases Description")
        if mappable(field, title, desc):
            continue
        sec = cell(row, col, "Section")
        pol = cell(row, col, "Polarity")
        m = RULE_RE.search(title)
        rule = m.group(1).strip() if m else "(no-ruleId)"
        by_rule[rule][field]["count"] += 1
        by_rule[rule][field]["sections"][sec] += 1
        by_rule[rule][field]["polarities"][pol] += 1
        field_totals[field] += 1
    wb.close()

    lines: list[str] = [
        "# Unmapped conditional matrix rules (for alias suggestions)",
        "",
        "Skipped FullMatrix rows: `Filed name` did not map to a Covoro `E Invoice` header "
        "(or matched the pack helper pseudo-field filter).",
        "",
        "Fill **suggested Excel header** for each field (must match template row 4).",
        "",
        f"- Skipped rows: **{sum(field_totals.values())}**",
        f"- Unique field labels: **{len(field_totals)}**",
        f"- Unique ruleIds: **{len(by_rule)}**",
        "",
        "## Closest known template headers (candidates)",
        "",
        "- `Invoice Total Tax Amount In Tax Accounting Currency`",
        "- `Invoice Total Tax Amount`",
        "- `Tax Category` / `Tax Rate`",
        "- `Currency Exchange Rate` / `Invoice Currency Code` / `Source Currency Code`",
        "- `Preceding Invoice reference` / `Invoice Type Code`",
        "",
        "## By ruleId",
        "",
    ]

    for rule in sorted(
        by_rule.keys(),
        key=lambda x: (-sum(v["count"] for v in by_rule[x].values()), x),
    ):
        total = sum(v["count"] for v in by_rule[rule].values())
        lines.append(f"### `{rule}` ({total} cases)")
        lines.append("")
        for field, meta in sorted(
            by_rule[rule].items(), key=lambda kv: -kv[1]["count"]
        ):
            secs = ", ".join(f"{s}:{c}" for s, c in meta["sections"].most_common())
            pols = ", ".join(f"{p}:{c}" for p, c in meta["polarities"].most_common())
            safe = field.replace("|", "/")
            lines.append(f"- **Field label:** `{safe}`")
            lines.append(
                f"  - cases: {meta['count']} | polarity: {pols} | section: {secs}"
            )
            lines.append("  - suggested Excel header: `TBD`")
        lines.append("")

    lines.extend(
        [
            "## Unique field labels (frequency)",
            "",
            "| Count | Matrix Filed name |",
            "|---:|---|",
        ]
    )
    for f, c in field_totals.most_common():
        lines.append(f"| {c} | {f.replace('|', '/')} |")
    lines.append("")

    out.write_text("\n".join(lines), encoding="utf-8")
    print(f"wrote {out}")
    print(f"rules={len(by_rule)} fields={len(field_totals)} rows={sum(field_totals.values())}")
    print("TOP_FIELDS")
    for f, c in field_totals.most_common(20):
        print(f"{c:4d} | {f[:110]}")
    print("TOP_RULES")
    for rule in sorted(
        by_rule.keys(),
        key=lambda x: -sum(v["count"] for v in by_rule[x].values()),
    )[:20]:
        n = sum(v["count"] for v in by_rule[rule].values())
        print(f"{n:4d} | {rule}")


if __name__ == "__main__":
    main()
