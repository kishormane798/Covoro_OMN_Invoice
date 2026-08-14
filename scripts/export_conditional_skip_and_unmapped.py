#!/usr/bin/env python3
"""Export intentional-skip + unmapped lists for user mapping replies."""
from __future__ import annotations

from collections import Counter, defaultdict
from pathlib import Path

import openpyxl

ns: dict = {}
exec(
    Path(__file__)
    .with_name("list_unmapped_conditional_rules.py")
    .read_text(encoding="utf-8")
    .split("def main")[0],
    ns,
)

ALIASES = ns["ALIASES"]
PSEUDO = ns["PSEUDO"]
RULE_RE = ns["RULE_RE"]
INTENTIONAL_SKIP_RULE_IDS = ns["INTENTIONAL_SKIP_RULE_IDS"]
norm = ns["norm"]
strip_peppol_suffix = ns["strip_peppol_suffix"]
cell = ns["cell"]
is_intentional_skip_field = ns["is_intentional_skip_field"]
is_invoice_line_period_field = ns["is_invoice_line_period_field"]

FORMULA_RULES = {
    # User: keep as intentional skip (not needed)
    "IBR-033-OM",
    "IBR-041-OM",
}
BACKEND_RULES = {
    "IBR-066-OM",
    "IBR-096-OM",
    "IBR-097-OM",
    "IBR-073-OM",
    "IBR-074-OM",
    "IBR-173-OM",
    "IBR-059-OM",
}
CL_RULES: set[str] = set()  # CL-06/10/11 mapped — no longer intentional rule skips


def field_reason(field: str) -> str:
    # Line period (IBT-134/135) is aliased to Invoicing Period* — not a skip reason.
    if ns.get("is_tax_accounting_currency_field", lambda _f: False)(field):
        return "backend default / no need to enter (IBT-006 Tax Accounting Currency)"
    if ns.get("is_accounting_currency_vat_category_field", lambda _f: False)(field):
        return "no Covoro column (IBT-192 Accounting Currency VAT Category Code)"
    base = norm(strip_peppol_suffix(field) or field)
    raw = norm(field)
    if base in (
        "vat accounting currency",
        "tax accounting currency",
        "tax accounting currency code",
        "ibt-006",
    ) or raw == "ibt-006":
        return "backend default / no need to enter (IBT-006 Tax Accounting Currency)"
    if base in (
        "accounting currency vat category code",
        "accounting currency vat category",
    ):
        return "no Covoro column (IBT-192 Accounting Currency VAT Category Code)"
    if base in ("vat category tax amount", "the vat category tax amount") or raw in (
        "vat category tax amount",
        "the vat category tax amount",
    ):
        return "no Covoro column / backend default (VAT category tax amount)"
    if base == "btom-016" or raw == "btom-016":
        return "formula / out of scope (BTOM-016)"
    if "for each different value of vat category rate" in raw:
        return "formula / out of scope (VAT category rate aggregation)"
    if "document level allowance tax rate" in base:
        return "no separate column (IBT-096 Allowance VAT Rate)"
    return "no Covoro column / backend default"


def rule_bucket(rule_id: str) -> str:
    if rule_id in FORMULA_RULES:
        return "formula suite"
    if rule_id in BACKEND_RULES:
        return "backend-only"
    if rule_id in CL_RULES:
        return "code-list only"
    return "other intentional rule"


def classify(field: str, title: str, desc: str, header_norm: dict[str, str]):
    f = (field or "").strip()
    if not f:
        return "empty", None, "(no-ruleId)"
    base = strip_peppol_suffix(f) or f
    candidates = [f, base] if base != f else [f]
    expanded: list[str] = []
    for c in candidates:
        expanded.append(c)
        ascii_dash = c.replace("–", "-").replace("—", "-")
        if ascii_dash != c:
            expanded.append(ascii_dash)
    candidates = expanded
    m = RULE_RE.search(title or "")
    rule = m.group(1).strip().upper() if m else "(no-ruleId)"
    if m and m.group(1).strip().upper() in INTENTIONAL_SKIP_RULE_IDS:
        rid = m.group(1).strip().upper()
        return "intentional_rule", rule_bucket(rid), rid
    blob = f"{title}\n{desc}".lower()
    # Line period (IBT-134/135): aliased to Invoicing Period* — fall through to aliases.
    if is_intentional_skip_field(f):
        return "intentional_field", field_reason(f), rule
    if (
        any(norm(c) == "an invoice that contains an invoice line" for c in candidates)
        and "simplified tax invoice exception" in blob
    ):
        return "mapped", None, rule
    if any(c in ALIASES for c in candidates):
        return "mapped", None, rule
    if any(PSEUDO.search(c) for c in candidates):
        return "unmapped", None, rule
    for c in candidates:
        if norm(ALIASES.get(c, c)) in header_norm:
            return "mapped", None, rule
    return "unmapped", None, rule


def main() -> None:
    root = Path(__file__).resolve().parents[1]
    template = root / "testData" / "uploads" / "template.xlsx"
    matrix = (
        root
        / "testcase"
        / "conditional_validation"
        / "EINV_OMAN_ConditionalValidation_FullMatrix.xlsx"
    )
    out = (
        root
        / "testcase"
        / "conditional_validation"
        / "INTENTIONAL_SKIP_AND_UNMAPPED.md"
    )

    wb = openpyxl.load_workbook(template, read_only=True, data_only=True)
    ws = wb["E Invoice"]
    row4 = [
        str(h).strip()
        for h in next(ws.iter_rows(min_row=4, max_row=4, values_only=True))
        if h
    ]
    header_norm = {norm(h): h for h in row4}
    wb.close()

    wb = openpyxl.load_workbook(matrix, read_only=True, data_only=True)
    ws = wb["All Testcases"]
    rows = ws.iter_rows(values_only=True)
    headers = [str(c).strip() if c is not None else "" for c in next(rows)]
    col = {h: i for i, h in enumerate(headers) if h}

    total = mapped = 0
    int_rule_counts: Counter[tuple[str, str]] = Counter()
    int_fields: dict[str, dict] = defaultdict(
        lambda: {"count": 0, "rules": Counter(), "reason": ""}
    )
    unmapped: dict[str, dict] = defaultdict(lambda: {"count": 0, "rules": Counter()})

    for row in rows:
        if not row:
            continue
        tc = cell(row, col, "Test Case ID")
        if not tc or tc.lower() == "test case id":
            continue
        total += 1
        field = cell(row, col, "Filed name", "Field name")
        title = cell(
            row, col, "Test Case Title", "Testcase Title", "Test Cases Title"
        )
        desc = cell(row, col, "Test Cases Description")
        kind, reason, rule = classify(field, title, desc, header_norm)
        if kind == "mapped":
            mapped += 1
        elif kind == "intentional_rule":
            int_rule_counts[(reason or "", rule)] += 1
        elif kind == "intentional_field":
            int_fields[field]["count"] += 1
            int_fields[field]["rules"][rule] += 1
            int_fields[field]["reason"] = reason or ""
        elif kind == "unmapped":
            unmapped[field]["count"] += 1
            unmapped[field]["rules"][rule] += 1
    wb.close()

    int_skip_total = sum(int_rule_counts.values()) + sum(
        v["count"] for v in int_fields.values()
    )
    unmapped_total = sum(v["count"] for v in unmapped.values())

    lines: list[str] = [
        "# Conditional validation — intentional skip vs unmapped",
        "",
        (
            f"Total matrix rows: **{total}** | Mapped: **{mapped}** | "
            f"Intentional skip: **{int_skip_total}** | Unmapped: **{unmapped_total}**"
        ),
        "",
        "How to reply with mappings:",
        "",
        "```",
        '"<matrix Filed name>" → "<Covoro template header>"',
        '"<matrix Filed name>" → skip / backend default',
        "```",
        "",
        "---",
        "",
        "## 1. Intentional skip",
        "",
        "Excluded on purpose — not alias gaps (unless product adds columns).",
        "",
    ]

    for bucket, blurb in (
        (
            "formula suite",
            "Calculated totals / Σ — belong in formula validation, not conditional Excel packs.",
        ),
        (
            "backend-only",
            "Not exposed on Covoro Excel error files.",
        ),
        (
            "code-list only",
            "CL dropdown/code-list — field validation suite unless there is a real if-then.",
        ),
    ):
        items = [(r, c) for (reason, r), c in int_rule_counts.items() if reason == bucket]
        if not items:
            continue
        lines.append(f"### Rules — {bucket} ({sum(c for _, c in items)} cases)")
        lines.append("")
        lines.append(blurb)
        lines.append("")
        lines.append("| Rule ID | Count |")
        lines.append("|---|---:|")
        for r, c in sorted(items, key=lambda x: (-x[1], x[0])):
            lines.append(f"| `{r}` | {c} |")
        lines.append("")

    lines.append("### Fields — no Covoro column / backend default")
    lines.append("")
    lines.append("| Field label | Count | Rule IDs | Reason |")
    lines.append("|---|---:|---|---|")
    for field, meta in sorted(
        int_fields.items(), key=lambda kv: (-kv[1]["count"], kv[0])
    ):
        rids = ", ".join(
            f"`{r}` ({c})" for r, c in meta["rules"].most_common(8)
        )
        reason = meta["reason"]
        for prefix in (
            "no Covoro column / backend default (",
            "backend default / no need to enter (",
            "no separate column (",
        ):
            if reason.startswith(prefix) and reason.endswith(")"):
                reason = reason[len(prefix) : -1]
                break
        safe = field.replace("|", "/")
        lines.append(f"| `{safe}` | {meta['count']} | {rids} | {reason} |")
    lines.append("")

    lines.extend(
        [
            "---",
            "",
            "## 2. Unmapped (needs mapping from user)",
            "",
            f"**{len(unmapped)}** unique field labels · **{unmapped_total}** cases",
            "",
            "Do **not** include intentional skips here.",
            "",
            "| Field label | Count | Example rule IDs |",
            "|---|---:|---|",
        ]
    )
    for field, meta in sorted(
        unmapped.items(), key=lambda kv: (-kv[1]["count"], kv[0])
    ):
        rids = ", ".join(f"`{r}`" for r, _ in meta["rules"].most_common(6))
        safe = field.replace("|", "/")
        lines.append(f"| `{safe}` | {meta['count']} | {rids} |")
    lines.append("")

    out.write_text("\n".join(lines), encoding="utf-8")
    print(f"wrote {out}")
    print(
        f"total={total} mapped={mapped} intentional={int_skip_total} "
        f"unmapped={unmapped_total} fields={len(unmapped)}"
    )


if __name__ == "__main__":
    main()
