#!/usr/bin/env python3
"""
Export FullMatrix with Filed name remapped to Covoro headers.

- Lists cases without proper *-OM rule mention → MISSING_RULE_MENTION.md
- Writes EINV_OMAN_ConditionalValidation_FullMatrix_MappedFiledNames.xlsx
- Yellow-fills intentional-skip rows (backend / default / not needed / field skip)

Reuses ALIASES / skip logic from list_unmapped_conditional_rules.py.
Does not overwrite the source FullMatrix.
"""
from __future__ import annotations

import re
from collections import Counter
from pathlib import Path

import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

ns: dict = {}
exec(
    Path(__file__)
    .with_name("list_unmapped_conditional_rules.py")
    .read_text(encoding="utf-8")
    .split("def main")[0],
    ns,
)

ALIASES = ns["ALIASES"]
PSEUDO = ns["PSEUDO"]
RULE_RE = ns["RULE_RE"]  # requires *OM* in brackets
INTENTIONAL_SKIP_RULE_IDS = ns["INTENTIONAL_SKIP_RULE_IDS"]
norm = ns["norm"]
strip_peppol_suffix = ns["strip_peppol_suffix"]
cell = ns["cell"]
is_intentional_skip_field = ns["is_intentional_skip_field"]
is_tax_accounting_currency_field = ns["is_tax_accounting_currency_field"]
is_accounting_currency_vat_category_field = ns[
    "is_accounting_currency_vat_category_field"
]

ANY_BRACKET_RE = re.compile(r"\[([A-Za-z0-9][A-Za-z0-9\-]*)\]")

YELLOW = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

# Rule-skip buckets for Skip Reason / Automation Triage (sync reference.md + TS helper)
FORMULA_NOT_NEEDED = {"IBR-033-OM", "IBR-041-OM"}
FORMULA_RULE_IDS = FORMULA_NOT_NEEDED | {
    "ALIGNED-IBRP-E-08-OM",
    "ALIGNED-IBRP-E-09-OM",
    "ALIGNED-IBRP-O-08-OM",
    "ALIGNED-IBRP-O-09-OM",
    "ALIGNED-IBRP-S-08-OM",
    "ALIGNED-IBRP-S-09-OM",
    "ALIGNED-IBRP-S-09-OM-WARN",
    "ALIGNED-IBRP-Z-08-OM",
    "ALIGNED-IBRP-Z-09-OM",
    "IBR-035-OM",
    "IBR-046-OM",
    "IBR-065-OM",
    "IBR-071-OM",
    "IBR-075-OM",
    "IBR-082-OM",
    "IBR-158-OM",
    "IBR-168-OM",
    "IBR-168-OM-WARN",
}
BACKEND_ONLY_RULE_IDS = {
    "IBR-066-OM",
    "IBR-095-OM",
    "IBR-096-OM",
    "IBR-097-OM",
}
DEFAULT_RULE_IDS = {"IBR-059-OM"}
# No Covoro column / user batch — keep in testcase, do not automate
IGNORE_RULE_IDS = {"IBR-073-OM", "IBR-074-OM", "IBR-173-OM"}
FIELD_RULE_IDS = {
    "CL-06-OM",
    "CL-10-OM",
    "CL-11-OM",
    "IBR-CL-05-OM",
    "IBR-002-OM",
}
# Legacy union for INTENTIONAL_SKIP_RULE_IDS overlap checks
BACKEND_RULES = BACKEND_ONLY_RULE_IDS | DEFAULT_RULE_IDS | IGNORE_RULE_IDS


def title_of(row, col: dict[str, int]) -> str:
    return cell(row, col, "Test Case Title", "Testcase Title", "Test Cases Title")


def field_candidates(field: str) -> list[str]:
    f = (field or "").strip()
    if not f:
        return []
    base = strip_peppol_suffix(f) or f
    out: list[str] = []
    for c in (f, base) if base != f else (f,):
        out.append(c)
        ascii_dash = c.replace("–", "-").replace("—", "-")
        if ascii_dash != c:
            out.append(ascii_dash)
    seen: set[str] = set()
    uniq: list[str] = []
    for c in out:
        if c not in seen:
            seen.add(c)
            uniq.append(c)
    return uniq


# Split compound matrix Filed names into segments (multi-field cases).
COMPOUND_SPLIT_RE = re.compile(
    r"\s*[–—]\s*|\s*/\s*|\s*,\s*|\s+and\s+",
    re.I,
)


def split_compound_field(field: str) -> list[str]:
    f = (field or "").strip()
    if not f:
        return []
    parts = [p.strip() for p in COMPOUND_SPLIT_RE.split(f) if p and p.strip()]
    return parts if parts else [f]


def resolve_one_header(segment: str, header_norm: dict[str, str]) -> str | None:
    """Resolve a single Filed-name segment to a Covoro template header."""
    for c in field_candidates(segment):
        if c in ALIASES:
            aliased = ALIASES[c]
            return header_norm.get(norm(aliased), aliased)
    for c in field_candidates(segment):
        aliased = ALIASES.get(c, c)
        if norm(aliased) in header_norm:
            return header_norm[norm(aliased)]
        if norm(c) in header_norm:
            return header_norm[norm(c)]
    return None


def resolve_all_headers(field: str, header_norm: dict[str, str]) -> list[str]:
    """
    Resolve every segment of a compound Filed name to Covoro headers.
    Prefer multi-segment results when 2+ segments map; otherwise whole-string alias.
    """
    f = (field or "").strip()
    if not f:
        return []

    segments = split_compound_field(f)
    seg_headers: list[str] = []
    seen: set[str] = set()
    for seg in segments:
        h = resolve_one_header(seg, header_norm)
        if not h:
            continue
        key = norm(h)
        if key in seen:
            continue
        seen.add(key)
        seg_headers.append(h)

    whole = resolve_one_header(f, header_norm)

    # Multi-field: show all distinct mapped headers
    if len(seg_headers) >= 2:
        return seg_headers

    # Single-segment or only one segment resolved — use whole alias if richer/same
    if whole:
        if not seg_headers:
            return [whole]
        # Same or whole is primary collapsing alias — keep segment if it matched
        if norm(whole) == norm(seg_headers[0]):
            return seg_headers
        # Whole-string alias differs (e.g. compound → primary only): if we only
        # got one segment mapped from a multi-part label, prefer whole for status
        # but if segments==1, return that one.
        if len(segments) == 1:
            return [whole]
        # Multi-part but only one segment resolved → include whole if different?
        # Prefer the one resolved segment plus try peppol tokens on unresolved?
        return seg_headers

    return seg_headers


def join_headers(headers: list[str]) -> str:
    return " | ".join(headers)


TITLE_INVOICE_TYPE_381_RE = re.compile(
    r"invoice\s+type\s+code.*(381|383|261)|(381|383|261).*invoice\s+type",
    re.I | re.S,
)

# Peppol Invoice Type Code → Covoro Masters / Master.omnCore dropdown labels
INVOICE_TYPE_CODE_TO_LABEL = {
    "381": "Credit note",
    "383": "Debit note",
    "261": "Self billed credit note",
    "389": "Self-billed invoice",
    "380": "Commercial invoice",
}

CN_DN_261_DESC = "Credit note, Debit note, or Self billed credit note"


def title_implies_invoice_type_codes(title: str, desc: str) -> bool:
    blob = f"{title}\n{desc}"
    if not (
        re.search(r"\b381\b", blob)
        or re.search(r"\b383\b", blob)
        or re.search(r"\b261\b", blob)
        or re.search(r"\b389\b", blob)
    ):
        return False
    return bool(
        re.search(r"invoice\s+type\s+code", blob, re.I)
        or re.search(r"\bIBT-003\b", blob, re.I)
        or re.search(r"self-?billed\s+credit", blob, re.I)
    )


def rewrite_title_codes_to_descriptions(title: str) -> tuple[str, bool]:
    """
    Replace raw Invoice Type Code numbers in titles with Covoro dropdown labels.
    Returns (new_title, changed).
    """
    if not title:
        return title, False
    out = title

    def repl_list(m: re.Match[str]) -> str:
        prefix = m.group("prefix")
        verb = (m.group("verb") or "").strip()
        if verb:
            return f"{prefix} {verb} {CN_DN_261_DESC}"
        return f"{prefix} {CN_DN_261_DESC}"

    # Invoice Type Code [is|is not] 381, 383, or 261
    out = re.sub(
        r"(?P<prefix>Invoice Type Code\s*(?:\(IBT-003\))?)\s*"
        r"(?P<verb>is\s+not|is)?\s*"
        r"['\"]?381['\"]?\s*,\s*['\"]?383['\"]?\s*,?\s*or\s*['\"]?261['\"]?",
        repl_list,
        out,
        flags=re.I,
    )
    # Invoice Type Code [is|is not] '381', '383', or '261'
    out = re.sub(
        r"(?P<prefix>Invoice Type Code\s*(?:\(IBT-003\))?)\s*"
        r"(?P<verb>is\s+not|is)?\s*"
        r"['\"]381['\"]\s*,\s*['\"]383['\"]\s*,?\s*or\s*['\"]261['\"]",
        repl_list,
        out,
        flags=re.I,
    )

    def repl_code(m: re.Match[str]) -> str:
        return INVOICE_TYPE_CODE_TO_LABEL.get(m.group(1), m.group(1))

    out = re.sub(
        r"(?<![A-Za-z0-9-])['\"]?(381|383|261|389|380)['\"]?(?![A-Za-z0-9-])",
        repl_code,
        out,
    )
    out = re.sub(
        r"(Credit note|Debit note|Self billed credit note|Self-billed invoice|Commercial invoice)"
        r"\s*\(\1\)",
        r"\1",
        out,
        flags=re.I,
    )
    out = re.sub(r" {2,}", " ", out)
    return out, out != title


def add_header_to_union(
    token: str,
    header: str,
    header_norm: dict[str, str],
    rule_header_union: dict[str, list[str]],
    rule_header_seen: dict[str, set[str]],
) -> None:
    if not token or not header:
        return
    key = token.upper()
    canon = header_norm.get(norm(header), header)
    seen = rule_header_seen.setdefault(key, set())
    union = rule_header_union.setdefault(key, [])
    nk = norm(canon)
    if nk in seen:
        return
    seen.add(nk)
    if norm(canon) == norm("Invoice Type Code"):
        union.insert(0, canon)
    else:
        union.append(canon)


def rule_skip_reason(rule: str) -> str | None:
    if not rule:
        return None
    r = rule.upper()
    if r in FORMULA_NOT_NEEDED:
        return "skip_not_needed"
    if r in IGNORE_RULE_IDS:
        return "skip_ignore"
    if r in DEFAULT_RULE_IDS:
        return "skip_default"
    if r in BACKEND_ONLY_RULE_IDS:
        return "skip_backend"
    if r in INTENTIONAL_SKIP_RULE_IDS:
        if r in FORMULA_NOT_NEEDED:
            return "skip_not_needed"
        if r in IGNORE_RULE_IDS:
            return "skip_ignore"
        if r in DEFAULT_RULE_IDS:
            return "skip_default"
        return "skip_backend"
    return None


def automation_triage(
    rule: str, status: str, skip_reason: str | None
) -> str:
    """
    Every FullMatrix row gets a destination — including skip/backend/default/ignore.
    Values: conditional | formula | backend | default | field | ignore | unmapped | empty
    """
    r = (rule or "").upper()
    if status == "empty":
        return "empty"
    if r in FORMULA_RULE_IDS:
        return "formula"
    if r in FIELD_RULE_IDS or (r.startswith("CL-") and r.endswith("-OM")):
        return "field"
    if r in BACKEND_ONLY_RULE_IDS:
        return "backend"
    if r in DEFAULT_RULE_IDS:
        return "default"
    if r in IGNORE_RULE_IDS:
        return "ignore"
    if status == "skip":
        if skip_reason == "skip_not_needed":
            return "formula"
        if skip_reason == "skip_default":
            return "default"
        if skip_reason == "skip_backend":
            return "backend"
        if skip_reason == "skip_ignore":
            return "ignore"
        return "ignore"
    if status == "unmapped":
        return "unmapped"
    if status == "mapped":
        return "conditional"
    return "unmapped"


def field_skip_reason(field: str) -> str | None:
    f = (field or "").strip()
    if not f:
        return None
    if is_tax_accounting_currency_field(f):
        return "skip_default"
    if is_accounting_currency_vat_category_field(f):
        return "skip_backend"
    if is_intentional_skip_field(f):
        raw = norm(f)
        if "allowance tax rate" in raw or "percentage" in raw:
            return "skip_backend"
        return "skip_backend"
    return None


def resolve_mapped_header(
    field: str, title: str, desc: str, header_norm: dict[str, str]
) -> tuple[str, str | None, str]:
    """
    Returns (status, resolved_headers_joined|None, skip_reason).
    status: mapped | skip | unmapped | empty
    When mapped, resolved is one or more Covoro headers joined with ` | `.
    """
    f = (field or "").strip()
    if not f:
        return "empty", None, ""

    candidates = field_candidates(f)
    m_om = RULE_RE.search(title or "") or RULE_RE.search(desc or "")
    m_any = ANY_BRACKET_RE.search(title or "") or ANY_BRACKET_RE.search(desc or "")
    rule = ""
    if m_om:
        rule = m_om.group(1).strip().upper()
    elif m_any:
        rule = m_any.group(1).strip().upper()

    rs = rule_skip_reason(rule)
    if rs:
        return "skip", None, rs

    fs = field_skip_reason(f)
    if fs:
        return "skip", None, fs

    blob = f"{title}\n{desc}".lower()
    if (
        any(norm(c) == "an invoice that contains an invoice line" for c in candidates)
        and "simplified tax invoice exception" in blob
    ):
        hdr = header_norm.get(norm("Invoice Transaction Type Code"))
        return "mapped", hdr or "Invoice Transaction Type Code", ""

    headers = resolve_all_headers(f, header_norm)
    if headers:
        return "mapped", join_headers(headers), ""

    if any(PSEUDO.search(c) for c in candidates):
        return "unmapped", None, ""

    return "unmapped", None, ""


def classify_rule_mention(title: str, desc: str) -> tuple[str, str]:
    blob_title = title or ""
    blob_desc = desc or ""
    m_om = RULE_RE.search(blob_title) or RULE_RE.search(blob_desc)
    if m_om:
        return "om", m_om.group(1).strip().upper()
    m_any = ANY_BRACKET_RE.search(blob_title) or ANY_BRACKET_RE.search(blob_desc)
    if m_any:
        return "non_om", m_any.group(1).strip().upper()
    return "missing", ""


def yellow_row(ws, row: int, max_col: int) -> None:
    for c in range(1, max_col + 1):
        ws.cell(row, c).fill = YELLOW


def main() -> None:
    root = Path(__file__).resolve().parents[1]
    template = root / "testData" / "uploads" / "template.xlsx"
    matrix = (
        root
        / "testcase"
        / "conditional_validation"
        / "EINV_OMAN_ConditionalValidation_FullMatrix.xlsx"
    )
    out_xlsx = (
        root
        / "testcase"
        / "conditional_validation"
        / "EINV_OMAN_ConditionalValidation_FullMatrix_MappedFiledNames.xlsx"
    )
    out_missing = (
        root
        / "testcase"
        / "conditional_validation"
        / "MISSING_RULE_MENTION.md"
    )

    if not matrix.is_file():
        raise SystemExit(f"FullMatrix not found: {matrix}")

    twb = openpyxl.load_workbook(template, read_only=True, data_only=True)
    tws = twb["E Invoice"]
    row4 = [
        str(h).strip()
        for h in next(tws.iter_rows(min_row=4, max_row=4, values_only=True))
        if h
    ]
    header_norm = {norm(h): h for h in row4}
    twb.close()

    # --- Task A: scan source ---
    swb = openpyxl.load_workbook(matrix, read_only=True, data_only=True)
    sws = (
        swb["All Testcases"]
        if "All Testcases" in swb.sheetnames
        else swb[swb.sheetnames[0]]
    )
    srows = sws.iter_rows(values_only=True)
    headers = [str(c).strip() if c is not None else "" for c in next(srows)]
    scol = {h: i for i, h in enumerate(headers) if h}

    missing_rows: list[dict] = []
    non_om_rows: list[dict] = []
    total = 0
    om_count = 0
    for row in srows:
        if not row:
            continue
        tc = cell(row, scol, "Test Case ID")
        if not tc or tc.lower() == "test case id":
            continue
        total += 1
        title = title_of(row, scol)
        desc = cell(row, scol, "Test Cases Description")
        field = cell(row, scol, "Filed name", "Field name")
        kind, token = classify_rule_mention(title, desc)
        if kind == "om":
            om_count += 1
        elif kind == "non_om":
            non_om_rows.append(
                {"id": tc, "token": token, "field": field, "title": title}
            )
        else:
            missing_rows.append({"id": tc, "field": field, "title": title})
    swb.close()

    lines = [
        "# Conditional FullMatrix — cases without proper `*-OM` rule mention",
        "",
        f"- Total cases: **{total}**",
        f"- Proper `*-OM` in title/description: **{om_count}**",
        f"- Non-standard bracket id (e.g. `[IBR-030]` without `-OM`): **{len(non_om_rows)}**",
        f"- No bracket rule mention at all: **{len(missing_rows)}**",
        "",
        "These still have a rule token — titles lack `-OM` suffix. Field aliases may still apply.",
        "",
        "## Non-standard bracket ids (not `*-OM`)",
        "",
        "| Test Case ID | Bracket token | Filed name | Test Case Title |",
        "|---|---|---|---|",
    ]
    for r in non_om_rows:
        title_esc = (r["title"] or "").replace("|", "\\|")
        field_esc = (r["field"] or "").replace("|", "\\|")
        lines.append(
            f"| `{r['id']}` | `{r['token']}` | {field_esc} | {title_esc} |"
        )
    if not non_om_rows:
        lines.append("| _(none)_ | | | |")

    lines += [
        "",
        "## Missing rule mention entirely (no `[…]` in title or description)",
        "",
        "| Test Case ID | Filed name | Test Case Title |",
        "|---|---|---|",
    ]
    for r in missing_rows:
        title_esc = (r["title"] or "").replace("|", "\\|")
        field_esc = (r["field"] or "").replace("|", "\\|")
        lines.append(f"| `{r['id']}` | {field_esc} | {title_esc} |")
    if not missing_rows:
        lines.append("| _(none)_ | | |")

    tok_counts = Counter(r["token"] for r in non_om_rows)
    if tok_counts:
        lines += ["", "## Non-OM token frequency", ""]
        for tok, n in tok_counts.most_common():
            lines.append(f"- `{tok}` — {n}")

    out_missing.write_text("\n".join(lines) + "\n", encoding="utf-8")

    # --- Mapped workbook ---
    wb = openpyxl.load_workbook(matrix)
    ws = (
        wb["All Testcases"]
        if "All Testcases" in wb.sheetnames
        else wb[wb.sheetnames[0]]
    )
    header_row = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    headers_live = [str(c).strip() if c is not None else "" for c in header_row]
    col_idx = {h: i + 1 for i, h in enumerate(headers_live) if h}

    filed_col = col_idx.get("Filed name") or col_idx.get("Field name")
    if not filed_col:
        raise SystemExit(f"No Filed name column. Headers: {headers_live}")

    def ensure_col(name: str) -> int:
        if name in col_idx:
            return col_idx[name]
        new_i = max(col_idx.values()) + 1
        ws.cell(1, new_i, name)
        col_idx[name] = new_i
        headers_live.append(name)
        return new_i

    orig_col = ensure_col("Original Filed name")
    status_col = ensure_col("Mapping Status")
    skip_col = ensure_col("Skip Reason")
    triage_col = ensure_col("Automation Triage")
    resolved_col = ensure_col("Resolved Covoro Header")
    orig_title_col = ensure_col("Original Test Case Title")
    title_desc_col = ensure_col("Test Case Title (Descriptions)")
    rule_kind_col = ensure_col("Rule Mention Kind")
    rule_token_col = ensure_col("Rule Token")

    title_col_write = (
        col_idx.get("Test Case Title")
        or col_idx.get("Testcase Title")
        or col_idx.get("Test Cases Title")
    )

    id_col = col_idx.get("Test Case ID")
    title_col = (
        col_idx.get("Test Case Title")
        or col_idx.get("Testcase Title")
        or col_idx.get("Test Cases Title")
    )
    desc_col = col_idx.get("Test Cases Description")

    max_col = max(col_idx.values())
    stats = Counter()
    skip_reasons = Counter()
    triage_counts = Counter()
    updated = 0
    unchanged = 0
    yellowed = 0
    multi_field_rows = 0
    multi_samples: list[dict] = []

    # Pass 1: resolve every row; accumulate per-rule union of Covoro headers.
    # (Later TCs often have a shorter Filed name than the first compound row.)
    row_infos: list[dict] = []
    rule_header_union: dict[str, list[str]] = {}
    rule_header_seen: dict[str, set[str]] = {}

    for r in range(2, ws.max_row + 1):
        tc_val = ws.cell(r, id_col).value if id_col else None
        if tc_val is None or str(tc_val).strip() == "":
            continue
        if str(tc_val).strip().lower() == "test case id":
            continue

        original = ws.cell(r, filed_col).value
        original_s = "" if original is None else str(original).strip()
        title = (
            ""
            if not title_col or ws.cell(r, title_col).value is None
            else str(ws.cell(r, title_col).value).strip()
        )
        desc = (
            ""
            if not desc_col or ws.cell(r, desc_col).value is None
            else str(ws.cell(r, desc_col).value).strip()
        )

        status, resolved, skip_reason = resolve_mapped_header(
            original_s, title, desc, header_norm
        )
        kind, token = classify_rule_mention(title, desc)
        headers = (
            [h.strip() for h in (resolved or "").split(" | ") if h.strip()]
            if status == "mapped" and resolved
            else []
        )

        row_infos.append(
            {
                "r": r,
                "tc_id": str(tc_val).strip(),
                "original_s": original_s,
                "title": title,
                "desc": desc,
                "status": status,
                "headers": headers,
                "skip_reason": skip_reason,
                "kind": kind,
                "token": token,
            }
        )

        if status == "mapped" and token and headers:
            key = token.upper()
            seen = rule_header_seen.setdefault(key, set())
            union = rule_header_union.setdefault(key, [])
            for h in headers:
                nk = norm(h)
                if nk in seen:
                    continue
                seen.add(nk)
                union.append(h)

        # Title/description implies Invoice Type Code 381/383/261 → include that header
        if status == "mapped" and token and title_implies_invoice_type_codes(
            title, desc
        ):
            add_header_to_union(
                token,
                "Invoice Type Code",
                header_norm,
                rule_header_union,
                rule_header_seen,
            )
            # Known companion fields for CN/DN/261 rules
            tok_u = token.upper()
            if tok_u in ("IBR-023-OM",):
                add_header_to_union(
                    token,
                    "Credit Note Or Debit Note Reason Code",
                    header_norm,
                    rule_header_union,
                    rule_header_seen,
                )
            if tok_u in ("ALIGNED-IBRP-028-OM", "IBR-032-OM"):
                add_header_to_union(
                    token,
                    "Preceding Invoice reference",
                    header_norm,
                    rule_header_union,
                    rule_header_seen,
                )
            if tok_u == "IBR-032-OM":
                add_header_to_union(
                    token,
                    "Preceding Invoice issue date",
                    header_norm,
                    rule_header_union,
                    rule_header_seen,
                )
                add_header_to_union(
                    token,
                    "Unique Identifier Number",
                    header_norm,
                    rule_header_union,
                    rule_header_seen,
                )

    # Pass 2: write cells — Filed name = multi-field Covoro union; titles get
    # code→description rewrite; Original Title preserved.
    titles_rewritten = 0
    title_samples: list[dict] = []
    invoice_type_381_cases: list[dict] = []

    for info in row_infos:
        r = info["r"]
        original_s = info["original_s"]
        status = info["status"]
        skip_reason = info["skip_reason"]
        kind = info["kind"]
        token = info["token"]
        title = info["title"]
        desc = info["desc"]
        headers = list(info["headers"])

        if status == "mapped" and token:
            union = rule_header_union.get(token.upper()) or []
            if len(union) >= 2:
                headers = list(union)
            elif len(union) == 1 and not headers:
                headers = list(union)

        resolved = join_headers(headers) if headers else None
        new_title, title_changed = rewrite_title_codes_to_descriptions(title)
        triage = automation_triage(token, status, skip_reason)
        triage_counts[triage] += 1

        ws.cell(r, orig_col, original_s)
        ws.cell(r, status_col, status)
        ws.cell(r, skip_col, skip_reason)
        ws.cell(r, triage_col, triage)
        ws.cell(r, resolved_col, resolved or "")
        ws.cell(r, orig_title_col, title)
        ws.cell(r, title_desc_col, new_title)
        ws.cell(r, rule_kind_col, kind)
        ws.cell(r, rule_token_col, token)

        # Multi-field Filed name for mapped rows (union); skip keeps original
        if status == "mapped" and resolved:
            ws.cell(r, filed_col, resolved)
        else:
            ws.cell(r, filed_col, original_s)

        # Update visible Test Case Title to description form when codes were present
        if title_col_write and title_changed:
            ws.cell(r, title_col_write, new_title)
            titles_rewritten += 1
            if len(title_samples) < 10:
                title_samples.append(
                    {
                        "id": info["tc_id"],
                        "rule": token,
                        "before": title[:120],
                        "after": new_title[:120],
                    }
                )

        if title_implies_invoice_type_codes(title, desc) or title_changed:
            invoice_type_381_cases.append(
                {
                    "id": info["tc_id"],
                    "rule": token,
                    "filed": (resolved or original_s)[:90],
                    "title_before": title[:100],
                    "title_after": new_title[:100],
                }
            )

        if status == "mapped" and resolved:
            n_headers = resolved.count(" | ") + 1
            if n_headers >= 2:
                multi_field_rows += 1
                if len(multi_samples) < 8:
                    multi_samples.append(
                        {
                            "id": info["tc_id"],
                            "rule": token,
                            "filed_name": resolved,
                            "original_filed": original_s[:80],
                        }
                    )
            if original_s != resolved:
                updated += 1
            else:
                unchanged += 1
            stats["mapped"] += 1
        elif status == "skip":
            unchanged += 1
            stats["skip"] += 1
            skip_reasons[skip_reason or "skip"] += 1
            yellow_row(ws, r, max_col)
            yellowed += 1
        elif status == "empty":
            unchanged += 1
            stats["empty"] += 1
        else:
            unchanged += 1
            stats["unmapped"] += 1

    evidence_rule = None
    evidence_rows: list[dict] = []
    for tok, union in sorted(
        rule_header_union.items(), key=lambda kv: -len(kv[1])
    ):
        if len(union) < 2:
            continue
        evidence_rule = tok
        for info in row_infos:
            if info["token"].upper() != tok or info["status"] != "mapped":
                continue
            evidence_rows.append(
                {
                    "id": info["tc_id"],
                    "filed_name": join_headers(union),
                    "original_filed": info["original_s"][:60],
                }
            )
        break

    for c in (
        orig_col,
        status_col,
        skip_col,
        triage_col,
        resolved_col,
        orig_title_col,
        title_desc_col,
        rule_kind_col,
        rule_token_col,
    ):
        ws.column_dimensions[get_column_letter(c)].width = 28
    ws.column_dimensions[get_column_letter(orig_title_col)].width = 55
    ws.column_dimensions[get_column_letter(title_desc_col)].width = 55
    ws.column_dimensions[get_column_letter(resolved_col)].width = 50

    # --- Coverage Summary sheet (all 599 TCs incl. skip/backend/default/ignore) ---
    if "Coverage Summary" in wb.sheetnames:
        del wb["Coverage Summary"]
    cov = wb.create_sheet("Coverage Summary", 0)
    total_rows = sum(triage_counts.values())
    cov.append(["Oman conditional FullMatrix — automation triage"])
    cov.append(["Total test cases", total_rows])
    cov.append([])
    cov.append(["Automation Triage", "Count", "Meaning"])
    triage_blurbs = {
        "conditional": "Excel upload if-then — Playwright conditional suite",
        "formula": "Σ / calculated totals — formula validation suite",
        "backend": "Not exposed on Covoro Excel error file",
        "default": "Backend default — no need to enter in Excel (e.g. IBT-006)",
        "field": "Code list / format — field validation suite",
        "ignore": "Intentional out of scope — tracked, not automated",
        "unmapped": "Needs Filed name → Covoro header mapping",
        "empty": "Missing Filed name",
    }
    for key in (
        "conditional",
        "formula",
        "backend",
        "default",
        "field",
        "ignore",
        "unmapped",
        "empty",
    ):
        n = triage_counts.get(key, 0)
        if n:
            cov.append([key, n, triage_blurbs.get(key, "")])
    cov.append([])
    cov.append(["Mapping Status", "Count"])
    for k, n in sorted(stats.items()):
        cov.append([k, n])
    cov.append([])
    cov.append(["Skip Reason", "Count"])
    for k, n in sorted(skip_reasons.items()):
        cov.append([k or "(blank)", n])

    coverage_md = (
        root
        / "testcase"
        / "conditional_validation"
        / "COVERAGE_SUMMARY.md"
    )
    md_cov: list[str] = [
        "# Conditional validation — full testcase coverage",
        "",
        f"Total matrix rows: **{total_rows}** (includes skip, backend, default, ignore).",
        "",
        "## Automation Triage",
        "",
        "| Triage | Count | Destination |",
        "|---|---:|---|",
    ]
    for key in (
        "conditional",
        "formula",
        "backend",
        "default",
        "field",
        "ignore",
        "unmapped",
        "empty",
    ):
        n = triage_counts.get(key, 0)
        if n:
            md_cov.append(f"| `{key}` | {n} | {triage_blurbs.get(key, '')} |")
    md_cov += [
        "",
        "## Mapping Status",
        "",
        "| Status | Count |",
        "|---|---:|",
    ]
    for k, n in sorted(stats.items()):
        md_cov.append(f"| {k} | {n} |")
    coverage_md.parent.mkdir(parents=True, exist_ok=True)
    coverage_md.write_text("\n".join(md_cov) + "\n", encoding="utf-8")

    out_xlsx.parent.mkdir(parents=True, exist_ok=True)
    try:
        wb.save(out_xlsx)
    except PermissionError:
        alt = out_xlsx.with_name(out_xlsx.stem + "_UPDATED.xlsx")
        wb.save(alt)
        out_xlsx = alt
        print({"warn": f"Primary xlsx locked; wrote {alt}"})
    wb.close()

    try:
        from export_conditional_skip_and_unmapped import main as export_skip_main

        export_skip_main()
    except Exception as exc:  # pragma: no cover
        print({"warn_skip_export": str(exc)})

    # Also write a focused markdown list of title rewrites
    title_md = (
        root
        / "testcase"
        / "conditional_validation"
        / "TITLE_CODE_TO_DESCRIPTION.md"
    )
    md_lines = [
        "# Test Case Title — Invoice Type codes → descriptions",
        "",
        "## Code → Covoro dropdown label",
        "",
        "| Code | Description |",
        "|---|---|",
    ]
    for code, label in INVOICE_TYPE_CODE_TO_LABEL.items():
        md_lines.append(f"| `{code}` | {label} |")
    md_lines += [
        "",
        f"- Titles rewritten: **{titles_rewritten}**",
        f"- Cases mentioning 381/383/261/389 patterns: **{len(invoice_type_381_cases)}**",
        "",
        "## Samples (before → after)",
        "",
    ]
    for s in title_samples:
        md_lines.append(f"### `{s['id']}` ({s['rule']})")
        md_lines.append(f"- Before: {s['before']}")
        md_lines.append(f"- After: {s['after']}")
        md_lines.append("")
    md_lines.append("## All matched cases")
    md_lines.append("")
    md_lines.append("| TC | Rule | Title (after) |")
    md_lines.append("|---|---|---|")
    for c in invoice_type_381_cases:
        md_lines.append(
            f"| `{c['id']}` | `{c['rule']}` | {c['title_after'].replace('|', '/')} |"
        )
    title_md.write_text("\n".join(md_lines) + "\n", encoding="utf-8")

    print(
        {
            "mapped_xlsx": str(out_xlsx),
            "title_rewrite_report": str(title_md),
            "code_to_description": INVOICE_TYPE_CODE_TO_LABEL,
            "titles_rewritten": titles_rewritten,
            "title_samples": title_samples,
            "invoice_type_code_cases": len(invoice_type_381_cases),
            "status_counts": dict(stats),
            "triage_counts": dict(triage_counts),
            "coverage_summary_md": str(coverage_md),
            "yellow_rows": yellowed,
            "multi_field_filed_name_rows": multi_field_rows,
            "filed_name_policy": (
                "Filed name = per-rule multi Covoro headers ( | ); "
                "Test Case Title rewritten to dropdown descriptions; "
                "Original Test Case Title preserved"
            ),
            "evidence_rule": evidence_rule,
            "evidence_rows": evidence_rows[:6],
        }
    )


if __name__ == "__main__":
    main()
