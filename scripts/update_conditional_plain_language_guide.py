#!/usr/bin/env python3
"""Regenerate Oman conditional plain-language Guide (xlsx + md + pdf).

Fixes OE (PINT-OM) semantics, maps field names → Covoro Excel headers via ALIASES,
and writes plain business if-then wording per rule (not QA PASS/FAIL jargon).
"""
from __future__ import annotations

import re
from pathlib import Path

import openpyxl
from fpdf import FPDF
from openpyxl.styles import Alignment, Font

ROOT = Path(__file__).resolve().parents[1]
DIR = ROOT / "testcase" / "conditional_validation"
XLSX = DIR / "EINV_OMAN_Conditional_Rules_PlainLanguage_Guide.xlsx"
MD = DIR / "EINV_OMAN_Conditional_Rules_PlainLanguage_Guide.md"
PDF = DIR / "EINV_OMAN_Conditional_Rules_PlainLanguage.pdf"
PEPPOL_BASE = (
    "https://test-docs.peppol.eu/pint/pint-om/2026-Q2-v1.0.1/pint-om/trn-invoice/rule"
)

# Synced with scripts/list_unmapped_conditional_rules.py /
# Helpers/conditionalValidationExcelPackHelper.ts
ALIASES: dict[str, str] = {
    "Preceding invoice reference": "Preceding Invoice reference",
    "Invoice transaction type": "Invoice Transaction Type Code",
    "Invoice Transaction Type": "Invoice Transaction Type Code",
    "If invoice transaction type": "Invoice Transaction Type Code",
    "If Invoice transaction type": "Invoice Transaction Type Code",
    "When Invoice transaction type": "Invoice Transaction Type Code",
    "Simplified Tax Invoice": "Invoice Transaction Type Code",
    "BTOM-001": "Invoice Transaction Type Code",
    "Invoice type code": "Invoice Type Code",
    "If Invoice type code": "Invoice Type Code",
    "if Invoice type code": "Invoice Type Code",
    "If Invoice Type code": "Invoice Type Code",
    "Currency exchange rate": "Currency Exchange Rate",
    "When Invoice currency code": "Invoice Currency Code",
    "Invoice currency code": "Invoice Currency Code",
    "If Invoice currency code": "Invoice Currency Code",
    "If the TAX category code for tax category tax amount in accounting currency":
        "Invoice total tax amount in tax accounting currency",
    "Item Type": "Item Type",
    "When Item type": "Item Type",
    "Item classification identifier": "Item classification identifier",
    "Industrial Classification Code must be provided for each ITEM INFORMATION":
        "Industrial Classification Code",
    "Delivery to Country code": "Deliver to country code",
    "Deliver to Country code": "Deliver to country code",
    "Deliver to address line 1 - Postal code": "Deliver to post code",
    "Customs Declaration number": "Customs Declaration number",
    "Import date": "Import date",
    "Invoicing period start date": "Invoicing period start date",
    "Invoicing period Start date": "Invoicing period start date",
    "Invoicing period end date": "Invoicing period end date",
    "Invoicing period End date": "Invoicing period end date",
    "Seller identifier": "Seller identifier",
    "Buyer identifier": "Buyer identifier",
    "the value in the Buyer identifier": "Buyer identifier",
    "If Buyer electronic address": "Buyer electronic address",
    "Buyer electronic address": "Buyer electronic address",
    "Tax Category": "Tax Category",
    "VAT Category": "Tax Category",
    "VAT category code": "Tax Category",
    "Invoiced item VAT category code": "Tax Category",
    "Invoiced item Tax Category": "Tax Category",
    "Tax Rate": "Tax Rate",
    "VAT Rate": "Tax Rate",
    "Invoiced item VAT rate": "Tax Rate",
    "Invoice item VAT rate": "Tax Rate",
    "In an Invoice line": "Tax Rate",
    "An Invoice line": "Tax Rate",
    "Tax exemption reason code": "Tax exemption reason code",
    "TAX category tax amount in accounting currency":
        "Invoice total tax amount in tax accounting currency",
    "Tax category tax amount in accounting currency":
        "Invoice total tax amount in tax accounting currency",
    "IBT-190": "Invoice total tax amount in tax accounting currency",
    "IBT-192(tax category tax amount in accounting currency)":
        "Invoice total tax amount in tax accounting currency",
    "In Line VAT information": "Line item VAT amount",
    "Each Invoice/CreditNote line must contain Item VAT Amount":
        "Line item VAT amount",
    "Line Item VAT Amount": "Line item VAT amount",
    "Item VAT Amount": "Line item VAT amount",
    "Seller tax identifier": "Seller VAT Identifier (TRN / TIN)",
    "Seller postal address": "Seller post code",
    "Buyer VATIN": "Buyer VAT identifier",
    "If Document level charge TAX category code": "Vat category - charges",
    "Document level charge TAX category code": "Vat category - charges",
    "If Document level allowance TAX category code": "Vat category - allowances",
    "Document level allowance TAX category code": "Vat category - allowances",
    "Document level allowance": "Allowances on document level",
    "Document level charge": "Charges on document level",
    "An Invoice that contains an Invoice line": "Allowances on document level",
    "Charge base amount (IBT-142)": "Charges on document level",
    "Charge base amount": "Charges on document level",
    "Allowance base amount (IBT-137)": "Allowances on document level",
    "Allowance base amount (IBT-093)": "Allowances on document level",
    "Allowance base amount": "Allowances on document level",
    "VAT breakdown(IBG-23)": "Invoice total tax amount",
    "In a VAT breakdown": "Invoice total tax amount",
    "Credit Note or Debit Note reason code": "Credit note or Debit Note reason code",
    "Credit note or Debit Note reason code": "Credit note or Debit Note reason code",
    "Rounding amount": "Rounding amount",
    "Seller country code": "Seller country code",
    "Deliver to country code": "Deliver to country code",
    "Related invoice fields": "Invoice total tax amount",
    "VAT category tax amount": "Invoice total tax amount",
}

# Explicit OE-correct overrides (main / when / simple / trigger / dependent / effect).
# effect: mandatory | forbidden | value | formula | codelist | exclusive
OVERRIDES: dict[str, dict[str, str]] = {
    "ALIGNED-IBRP-028-OM": {
        "main": "Preceding Invoice reference",
        "when": "Applies when Invoice Type Code is Credit note (381), Debit note (383), or Self billed credit note (261)",
        "simple": "If Invoice Type Code is Credit note / Debit note / Self billed credit note, then Preceding Invoice reference must be filled.",
        "trigger": "Invoice Type Code = Credit note / Debit note / Self billed credit note",
        "dependent": "Preceding Invoice reference",
        "effect": "mandatory",
    },
    "ALIGNED-IBRP-E-01-OM": {
        "main": "Tax Category (E) -> VAT breakdown (Invoice total tax amount)",
        "when": "Applies when any line Tax Category, Vat category - allowances, or Vat category - charges is Exempt (E)",
        "simple": "If Tax Category (or document allowance/charge VAT category) is Exempt (E), the invoice must include exactly one E VAT breakdown — unless Invoice Transaction Type Code is Simplified Tax Invoice.",
        "trigger": "Tax Category / Vat category - allowances / Vat category - charges = Exempt (E)",
        "dependent": "Invoice total tax amount (VAT breakdown for E)",
        "effect": "mandatory",
    },
    "ALIGNED-IBRP-E-05-OM": {
        "main": "Tax Rate",
        "when": "Applies when Tax Category is Exempt (E)",
        "simple": "If Tax Category = Exempt (E), then Tax Rate must be left empty.",
        "trigger": "Tax Category = Exempt (E)",
        "dependent": "Tax Rate",
        "effect": "forbidden",
    },
    "ALIGNED-IBRP-O-01-OM": {
        "main": "Tax Category (O) -> VAT breakdown (Invoice total tax amount)",
        "when": "Applies when any line/allowance/charge Tax Category is Not subject to VAT (O)",
        "simple": "If Tax Category (or document allowance/charge VAT category) is Not subject to VAT (O), include exactly one O VAT breakdown — unless Simplified Tax Invoice.",
        "trigger": "Tax Category / Vat category - allowances / Vat category - charges = Not subject to VAT (O)",
        "dependent": "Invoice total tax amount (VAT breakdown for O)",
        "effect": "mandatory",
    },
    "ALIGNED-IBRP-O-05-OM": {
        "main": "Tax Rate",
        "when": "Applies when Tax Category is Not subject to VAT (O)",
        "simple": "If Tax Category = Not subject to VAT (O), then Tax Rate must be left empty.",
        "trigger": "Tax Category = Not subject to VAT (O)",
        "dependent": "Tax Rate",
        "effect": "forbidden",
    },
    "ALIGNED-IBRP-S-01-OM": {
        "main": "Tax Category (S) -> VAT breakdown (Invoice total tax amount)",
        "when": "Applies when any line/allowance/charge Tax Category is Standard (S)",
        "simple": "If Tax Category (or document allowance/charge VAT category) is Standard (S), the VAT breakdown must include at least one S category.",
        "trigger": "Tax Category / Vat category - allowances / Vat category - charges = Standard (S)",
        "dependent": "Invoice total tax amount (VAT breakdown for S)",
        "effect": "mandatory",
    },
    "ALIGNED-IBRP-S-05-OM": {
        "main": "Tax Rate",
        "when": "Applies when Tax Category is Standard (S)",
        "simple": "If Tax Category = Standard (S), then Tax Rate must be exactly 5.",
        "trigger": "Tax Category = Standard (S)",
        "dependent": "Tax Rate",
        "effect": "value",
        "value": "5",
    },
    "ALIGNED-IBRP-S-10-OM": {
        "main": "Tax exemption reason code",
        "when": "Applies when Tax Category (VAT breakdown) is Standard (S)",
        "simple": "If Tax Category = Standard (S), then Tax exemption reason code must be left empty.",
        "trigger": "Tax Category = Standard (S)",
        "dependent": "Tax exemption reason code",
        "effect": "forbidden",
    },
    "ALIGNED-IBRP-Z-01-OM": {
        "main": "Tax Category (Z) -> VAT breakdown (Invoice total tax amount)",
        "when": "Applies when any line/allowance/charge Tax Category is Zero rated (Z)",
        "simple": "If Tax Category (or document allowance/charge VAT category) is Zero rated (Z), include exactly one Z VAT breakdown — unless Simplified Tax Invoice.",
        "trigger": "Tax Category / Vat category - allowances / Vat category - charges = Zero rated (Z)",
        "dependent": "Invoice total tax amount (VAT breakdown for Z)",
        "effect": "mandatory",
    },
    "ALIGNED-IBRP-Z-05-OM": {
        "main": "Tax Rate",
        "when": "Applies when Tax Category is Zero rated (Z)",
        "simple": "If Tax Category = Zero rated (Z), then Tax Rate must be 0.",
        "trigger": "Tax Category = Zero rated (Z)",
        "dependent": "Tax Rate",
        "effect": "value",
        "value": "0",
    },
    "ALIGNED-IBRP-048": {
        "main": "Tax Rate (VAT breakdown rate)",
        "when": "Applies for each VAT breakdown unless the invoice is not subject to VAT (O)",
        "simple": "Each VAT breakdown must have a Tax Rate, except when the invoice is Not subject to VAT.",
        "trigger": "VAT breakdown present and not Not-subject-to-VAT",
        "dependent": "Tax Rate",
        "effect": "mandatory",
    },
    "CL-10-OM": {
        "main": "Tax exemption reason code",
        "when": "Applies when Tax Category is Zero rated (Z)",
        "simple": "If Tax Category = Zero rated (Z), Tax exemption reason code must be from the Zero rating codelist (VATZR-OM-01 .. VATZR-OM-16). This is not the same as IBR-069 (presence of a reason for E or Z).",
        "trigger": "Tax Category = Zero rated (Z)",
        "dependent": "Tax exemption reason code",
        "effect": "codelist",
        "value": "VATZR-OM-01..VATZR-OM-16",
    },
    "CL-06-OM": {
        "main": "Buyer identifier / Seller identifier (scheme)",
        "when": "Applies when Buyer identifier or Seller identifier is provided",
        "simple": "If Buyer identifier or Seller identifier is filled, its scheme identifier must use the Buyer/Seller Identifier codelist.",
        "trigger": "Buyer identifier or Seller identifier provided",
        "dependent": "Buyer identifier / Seller identifier scheme",
        "effect": "codelist",
    },
    "CL-11-OM": {
        "main": "Profit margin item reason code",
        "when": "Applies when Invoice Transaction Type Code is Profit margin invoice or Profit Margin Self-Invoice",
        "simple": "If Invoice Transaction Type Code is Profit margin invoice / Profit Margin Self-Invoice, Profit margin item reason code must be present and from CL-11 (VATPM-OM-01..05).",
        "trigger": "Invoice Transaction Type Code = Profit margin invoice / Profit Margin Self-Invoice",
        "dependent": "Profit margin item reason code",
        "effect": "codelist",
    },
    "IBR-004-OM": {
        "main": "Currency Exchange Rate",
        "when": "Applies when Invoice Currency Code is not OMR",
        "simple": "If Invoice Currency Code is not OMR, then Currency Exchange Rate must be filled.",
        "trigger": "Invoice Currency Code ≠ OMR",
        "dependent": "Currency Exchange Rate",
        "effect": "mandatory",
    },
    "IBR-006-OM": {
        "main": "Seller VAT Identifier (TRN / TIN)",
        "when": "Applies in ALL cases EXCEPT when Invoice Transaction Type Code is Import of Goods, Import of services for RCM, or Profit Margin Self-Invoice",
        "simple": "Seller VAT Identifier (TRN / TIN) must be filled on almost every invoice. It is NOT required only when Invoice Transaction Type Code is Import of Goods, Import of services for RCM, or Profit Margin Self-Invoice.",
        "trigger": "Invoice Transaction Type Code NOT IN (Import of Goods, Import of services for RCM, Profit Margin Self-Invoice)",
        "dependent": "Seller VAT Identifier (TRN / TIN)",
        "effect": "mandatory",
    },
    "IBR-012-OM": {
        "main": "Deliver to country code",
        "when": "Applies when Invoice Transaction Type Code is Export Invoice and at least one Tax exemption reason code is Export of service (VATZR-OM-09)",
        "simple": "If Invoice Transaction Type Code is Export Invoice and Tax exemption reason code is Export of service (VATZR-OM-09), Deliver to country code must NOT be OM.",
        "trigger": "Invoice Transaction Type Code = Export Invoice AND Tax exemption reason code = VATZR-OM-09",
        "dependent": "Deliver to country code",
        "effect": "forbidden",
        "value": "OM",
    },
    "IBR-023-OM": {
        "main": "Credit note or Debit Note reason code",
        "when": "Applies when Invoice Type Code is 381, 383, or 261",
        "simple": "If Invoice Type Code is Credit note / Debit note / Self billed credit note, then Credit note or Debit Note reason code must be filled.",
        "trigger": "Invoice Type Code = Credit note / Debit note / Self billed credit note",
        "dependent": "Credit note or Debit Note reason code",
        "effect": "mandatory",
    },
    "IBR-032-OM": {
        "main": "Preceding Invoice reference (+ issue date + UUID)",
        "when": "Applies when Invoice Type Code is 381, 383, or 261",
        "simple": "If Invoice Type Code is Credit note / Debit note / Self billed credit note, then Preceding Invoice reference, Preceding Invoice issue date, and Preceding invoice UUID must all be filled.",
        "trigger": "Invoice Type Code = Credit note / Debit note / Self billed credit note",
        "dependent": "Preceding Invoice reference / Preceding Invoice issue date / Preceding invoice UUID",
        "effect": "mandatory",
    },
    "IBR-034-OM": {
        "main": "VAT accounting currency (IBT-006 — no Covoro Excel column yet)",
        "when": "Applies when Invoice Currency Code is not OMR",
        "simple": "If Invoice Currency Code is not OMR, then VAT accounting currency must be provided (backend/XML; not a Covoro Excel error-file column yet).",
        "trigger": "Invoice Currency Code ≠ OMR",
        "dependent": "VAT accounting currency (IBT-006)",
        "effect": "mandatory",
    },
    "IBR-038-OM": {
        "main": "Line item VAT amount",
        "when": "Applies for each line EXCEPT when Invoice Transaction Type Code is Simplified Tax Invoice",
        "simple": "Each line must have Line item VAT amount, except on Simplified Tax Invoice.",
        "trigger": "Invoice Transaction Type Code ≠ Simplified Tax Invoice",
        "dependent": "Line item VAT amount",
        "effect": "mandatory",
    },
    "IBR-047-OM": {
        "main": "Allowance TAX rate (Vat category - allowances / rate)",
        "when": "Applies when Vat category - allowances is Standard (S)",
        "simple": "If Vat category - allowances = Standard (S), document-level allowance TAX rate must be 5.",
        "trigger": "Vat category - allowances = Standard (S)",
        "dependent": "Document level allowance TAX rate (IBT-096)",
        "effect": "value",
        "value": "5",
    },
    "IBR-069-OM": {
        "main": "Tax exemption reason code",
        "when": "Applies when Tax Category (VAT breakdown IBT-118) is Exempt (E) or Zero rated (Z)",
        "simple": "If Tax Category (VAT breakdown) is Exempt (E) or Zero rated (Z), then Tax exemption reason code must be filled.",
        "trigger": "Tax Category = Exempt (E) or Zero rated (Z)",
        "dependent": "Tax exemption reason code",
        "effect": "mandatory",
    },
    "IBR-070-OM": {
        "main": "Tax exemption reason code",
        "when": "Applies when Tax Category (VAT breakdown) is Not subject to VAT (O)",
        "simple": "If Tax Category = Not subject to VAT (O), then Tax exemption reason code must be left empty.",
        "trigger": "Tax Category = Not subject to VAT (O)",
        "dependent": "Tax exemption reason code",
        "effect": "forbidden",
    },
    "IBR-078-OM": {
        "main": "Item Type",
        "when": "Applies for each item EXCEPT when Invoice Transaction Type Code is Simplified Tax Invoice",
        "simple": "Item Type is required on each line, except for Simplified Tax Invoice.",
        "trigger": "Invoice Transaction Type Code ≠ Simplified Tax Invoice",
        "dependent": "Item Type",
        "effect": "mandatory",
    },
    "IBR-081-OM": {
        "main": "Industrial Classification Code",
        "when": "Applies for each item EXCEPT when Invoice Transaction Type Code is Simplified Tax Invoice, Import of Goods, Import of services for RCM, or Profit Margin Self-Invoice",
        "simple": "Each item must have Industrial Classification Code, except on Simplified Tax Invoice / Import of Goods / Import RCM / Profit Margin Self-Invoice.",
        "trigger": "Invoice Transaction Type Code NOT IN (Simplified Tax Invoice, Import of Goods, Import of services for RCM, Profit Margin Self-Invoice)",
        "dependent": "Industrial Classification Code",
        "effect": "mandatory",
    },
    "IBR-092-OM": {
        "main": "Document level allowance TAX rate",
        "when": "Applies when Vat category - allowances is Exempt (E)",
        "simple": "If Vat category - allowances = Exempt (E), then document-level allowance TAX rate must be left empty.",
        "trigger": "Vat category - allowances = Exempt (E)",
        "dependent": "Document level allowance TAX rate",
        "effect": "forbidden",
    },
    "IBR-093-OM": {
        "main": "Document level allowance TAX rate",
        "when": "Applies when Vat category - allowances is Not subject to VAT (O)",
        "simple": "If Vat category - allowances = Not subject to VAT (O), then document-level allowance TAX rate must be left empty.",
        "trigger": "Vat category - allowances = Not subject to VAT (O)",
        "dependent": "Document level allowance TAX rate",
        "effect": "forbidden",
    },
    "IBR-094-OM": {
        "main": "Document level allowance TAX rate",
        "when": "Applies when Vat category - allowances is Zero rated (Z)",
        "simple": "If Vat category - allowances = Zero rated (Z), document-level allowance TAX rate must be 0.",
        "trigger": "Vat category - allowances = Zero rated (Z)",
        "dependent": "Document level allowance TAX rate",
        "effect": "value",
        "value": "0",
    },
    "IBR-098-OM": {
        "main": "Document level charge TAX rate",
        "when": "Applies when Vat category - charges is Exempt (E)",
        "simple": "If Vat category - charges = Exempt (E), then document-level charge TAX rate must be left empty.",
        "trigger": "Vat category - charges = Exempt (E)",
        "dependent": "Document level charge TAX rate",
        "effect": "forbidden",
    },
    "IBR-099-OM": {
        "main": "Document level charge TAX rate",
        "when": "Applies when Vat category - charges is Not subject to VAT (O)",
        "simple": "If Vat category - charges = Not subject to VAT (O), then document-level charge TAX rate must be left empty.",
        "trigger": "Vat category - charges = Not subject to VAT (O)",
        "dependent": "Document level charge TAX rate",
        "effect": "forbidden",
    },
    "IBR-100-OM": {
        "main": "Document level charge TAX rate",
        "when": "Applies when Vat category - charges is Zero rated (Z)",
        "simple": "If Vat category - charges = Zero rated (Z), document-level charge TAX rate must be 0.",
        "trigger": "Vat category - charges = Zero rated (Z)",
        "dependent": "Document level charge TAX rate",
        "effect": "value",
        "value": "0",
    },
    "IBR-172-OM": {
        "main": "Currency Exchange Rate",
        "when": "Applies when Invoice Currency Code is OMR",
        "simple": "If Invoice Currency Code is OMR, then Currency Exchange Rate must be left empty.",
        "trigger": "Invoice Currency Code = OMR",
        "dependent": "Currency Exchange Rate",
        "effect": "forbidden",
    },
    "IBR-175-OM": {
        "main": "Preceding Invoice reference (+ Preceding invoice UUID)",
        "when": "Applies when Invoice Transaction Type Code (BTOM-001) is Profit margin invoice",
        "simple": "If Invoice Transaction Type Code (BTOM-001) is Profit margin invoice, then Preceding Invoice reference and Preceding invoice UUID must be filled.",
        "trigger": "Invoice Transaction Type Code = Profit margin invoice",
        "dependent": "Preceding Invoice reference / Preceding invoice UUID",
        "effect": "mandatory",
        "official_fix": True,
    },
    "IBR-066-OM": {
        "main": "Invoice total tax amount in tax accounting currency (+ IBT-192/193)",
        "when": "Applies when Invoice Currency Code ≠ OMR and at least one Tax Category is Standard (S)",
        "simple": "If Invoice Currency Code is not OMR and Tax Category includes Standard (S), accounting-currency tax amount fields (IBT-190/192/193) must be provided (BACKEND — not Excel error columns yet).",
        "trigger": "Invoice Currency Code ≠ OMR AND Tax Category includes S",
        "dependent": "Invoice total tax amount in tax accounting currency (proxy for IBT-190)",
        "effect": "mandatory",
    },
    "IBR-095-OM": {
        "main": "TAX category rate in accounting currency (IBT-193)",
        "when": "Applies when accounting-currency Tax Category code (IBT-192) is Exempt (E)",
        "simple": "If IBT-192 = E, IBT-193 must NOT be present (BACKEND).",
        "trigger": "IBT-192 = E",
        "dependent": "IBT-193",
        "effect": "forbidden",
    },
    "IBR-096-OM": {
        "main": "TAX category rate in accounting currency (IBT-193)",
        "when": "Applies when accounting-currency Tax Category code (IBT-192) is Not subject to VAT (O)",
        "simple": "If IBT-192 = O, IBT-193 must NOT be present (BACKEND).",
        "trigger": "IBT-192 = O",
        "dependent": "IBT-193",
        "effect": "forbidden",
    },
    "IBR-097-OM": {
        "main": "TAX category rate in accounting currency (IBT-193)",
        "when": "Applies when accounting-currency Tax Category code (IBT-192) is Zero rated (Z)",
        "simple": "If IBT-192 = Z, IBT-193 must be 0 (BACKEND).",
        "trigger": "IBT-192 = Z",
        "dependent": "IBT-193",
        "effect": "value",
        "value": "0",
    },
    "IBR-CL-05-OM": {
        "main": "Tax exemption reason code (allowance IBT-196)",
        "when": "Applies when Vat category - allowances is Zero rated (Z) or Exempt (E)",
        "simple": "If Vat category - allowances is Z or E, document-level allowance Tax exemption reason code must use the Zero rating / Exemption reason codelist.",
        "trigger": "Vat category - allowances = Z or E",
        "dependent": "Document level allowance TAX exemption reason code",
        "effect": "codelist",
    },
    "IBR-CL-10-OM": {
        "main": "Tax exemption reason code (allowance IBT-196)",
        "when": "Applies when Vat category - allowances is Zero rated (Z) or Exempt (E)",
        "simple": "If Vat category - allowances is Z or E, document-level allowance Tax exemption reason code must use the Zero rating / Exemption reason codelist.",
        "trigger": "Vat category - allowances = Z or E",
        "dependent": "Document level allowance TAX exemption reason code",
        "effect": "codelist",
    },
}


def map_field(label: str) -> str:
    lab = (label or "").strip()
    if not lab:
        return lab
    if lab in ALIASES:
        return ALIASES[lab]
    # partial keys
    for k, v in ALIASES.items():
        if lab.lower() == k.lower():
            return v
    return lab


def effect_requirement(o: dict[str, str]) -> str:
    """Plain requirement phrase for the dependent column."""
    effect = o["effect"]
    val = o.get("value", "")
    if effect == "mandatory":
        return "must be filled"
    if effect == "forbidden":
        if val:
            return f"must not be set to {val}"
        return "must be left empty"
    if effect == "value":
        return f"must be {val}"
    if effect == "codelist":
        if val:
            return f"must use a valid code from the codelist ({val})"
        return "must use a valid code from the required codelist"
    if effect == "formula":
        return "must match the calculated total / formula"
    if effect == "exclusive":
        return "must follow the exclusive / either-or requirement"
    return "must follow the rule"


def build_simple_rule(o: dict[str, str]) -> str:
    """Primary business if-then line (no PASS/FAIL / Trigger jargon)."""
    # Prefer curated simple English when present; else synthesize from fields.
    simple = (o.get("simple") or "").strip()
    if simple and not re.search(r"\b(PASS|FAIL|OFF|Trigger:|Dependent)\b", simple):
        # Ensure it starts with If / When when possible
        if re.match(r"^(If|When|Each|Seller|Item|Document)\b", simple, re.I):
            return simple
        return simple
    trig = o["trigger"]
    dep = o["dependent"]
    return f"If {trig}, then {dep} {effect_requirement(o)}."


def build_related_note(o: dict[str, str]) -> str:
    """Optional second line when related columns must move together."""
    dep = o["dependent"]
    # Multi-column dependents → "If you fill A, also fill B" style hint
    parts = [p.strip() for p in re.split(r"\s*/\s*", dep) if p.strip()]
    if len(parts) >= 2:
        first, rest = parts[0], parts[1:]
        also = " and ".join(rest)
        return f"If you enter a value in {first}, you must also enter a value in {also}."
    return ""


def build_what_we_test(o: dict[str, str]) -> str:
    """Plain-language meaning of the rule (Excel column view) — no PASS/FAIL."""
    trig = o["trigger"]
    dep = o["dependent"]
    req = effect_requirement(o)
    lines: list[str] = [
        build_simple_rule(o),
        f"In the Covoro Excel: when ({trig}), the column(s) {dep} {req}.",
    ]
    related = build_related_note(o)
    if related:
        lines.append(related)
    return "\n".join(lines)


def build_how_we_test(o: dict[str, str]) -> str:
    """How to check the rule in Excel — plain steps, no PASS/FAIL/OFF labels."""
    trig = o["trigger"]
    dep = o["dependent"]
    effect = o["effect"]
    val = o.get("value", "")
    lines = [
        "1) Open Upload e-invoice on the COVORO Oman portal.",
        "2) Start from a valid Covoro/OMN Excel invoice row.",
        f"3) Set the condition: {trig}.",
    ]
    if effect == "mandatory":
        lines.append(f"4) Correct sheet: fill {dep} with a valid value.")
        lines.append(f"5) Incorrect sheet: leave {dep} blank — the upload should report an error on {dep}.")
        lines.append(
            f"6) When the condition does not apply, leaving {dep} blank is allowed for this rule."
        )
    elif effect == "forbidden":
        if val:
            lines.append(f"4) Correct sheet: do not set {dep} to {val}.")
            lines.append(f"5) Incorrect sheet: set {dep} = {val} — the upload should report an error.")
        else:
            lines.append(f"4) Correct sheet: leave {dep} empty.")
            lines.append(f"5) Incorrect sheet: fill {dep} — the upload should report an error.")
        lines.append(
            "6) When the condition does not apply, this rule does not require the field to be empty."
        )
    elif effect == "value":
        lines.append(f"4) Correct sheet: set {dep} = {val}.")
        lines.append(
            f"5) Incorrect sheet: leave {dep} blank or use a different value — the upload should report an error."
        )
        lines.append(
            "6) When the condition does not apply, other values for this field are out of scope for this rule."
        )
    elif effect == "codelist":
        code_hint = f" from {val}" if val else ""
        lines.append(f"4) Correct sheet: set {dep} to a valid code{code_hint}.")
        lines.append(
            f"5) Incorrect sheet: leave {dep} blank or use an invalid code — the upload should report an error."
        )
        lines.append(
            "6) When the condition does not apply, this rule does not enforce the codelist."
        )
    else:
        lines.append(f"4) Correct sheet: satisfy the requirement for {dep}.")
        lines.append(f"5) Incorrect sheet: break the requirement for {dep} — expect an error.")
    lines.append(f"7) Rule in simple words: {build_simple_rule(o)}")
    return "\n".join(lines)


def infer_from_official(rule_id: str, main: str, official: str, when: str) -> dict[str, str]:
    """Heuristic fallback when no explicit override exists."""
    mapped = map_field(main)
    off = official or ""
    # Fix known BTOM-003 mislabel in official text for transaction type
    off_fixed = re.sub(
        r"Invoice transaction type\s*\(BTOM-003\)",
        "Invoice transaction type (BTOM-001)",
        off,
        flags=re.I,
    )
    is_except = bool(re.search(r"\bexcept\b", off_fixed, re.I))
    is_forbid = bool(
        re.search(r"MUST NOT|must not|MUST not|shall not|leave .* empty|not be present", off_fixed, re.I)
    )
    must_value = re.search(
        r"MUST be\s+([0-9]+(?:\.[0-9]+)?|0 \(zero\)|exactly 5|5\b)",
        off_fixed,
        re.I,
    )
    value = ""
    if must_value:
        raw = must_value.group(1)
        value = "0" if "0" in raw and "zero" in raw.lower() or raw.strip().startswith("0") else (
            "5" if "5" in raw else raw
        )

    # Extract trigger-ish phrase from when/official
    when_s = when or ""
    # Clean inverted except "Applies when: <exception cases>"
    if is_except and re.search(r"Applies when:.*(Simplified|import|profit margin self)", when_s, re.I):
        when_s = (
            "Applies in all cases EXCEPT the transaction types listed in the official rule text"
        )

    # Prefer mapped Covoro header as dependent
    dependent = mapped
    # Prefer simple English from official without "becomes mandatory" wrapper
    simple = off_fixed
    if len(simple) > 280:
        simple = simple[:277] + "..."
    # Prefer if-then from mapped fields
    trig = when_s.replace("Applies when: ", "").replace("Applies if: ", "").replace(
        "Applies where: ", ""
    ).strip()
    if not trig:
        trig = "as described in the official rule text"

    if is_except:
        effect = "mandatory"
        simple = (
            f"{dependent} must be filled on almost every invoice, except for the "
            f"Invoice Transaction Type Code / cases listed in the official Peppol text."
        )
        trig = f"NOT the exception cases in {rule_id}"
    elif is_forbid:
        effect = "forbidden"
        if value:
            simple = f"If {trig}, then {dependent} must not be set to {value}."
        else:
            simple = f"If {trig}, then {dependent} must be left empty."
    elif value:
        effect = "value"
        simple = f"If {trig}, then {dependent} must be {value}."
    else:
        effect = "mandatory"
        # Avoid saying trigger field is mandatory
        if re.search(r"becomes mandatory", (when or "") + main, re.I):
            simple = off_fixed[:240]
        else:
            simple = f"If {trig}, then {dependent} must be filled / must satisfy the official rule."

    return {
        "main": mapped if mapped else main,
        "when": when_s if when_s.startswith("Applies") else f"Applies when: {trig}",
        "simple": simple,
        "trigger": trig,
        "dependent": dependent,
        "effect": effect,
        "value": value,
        "official": off_fixed,
    }


def load_rows(ws) -> list[dict]:
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = []
    for excel_row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        vals = [c.value for c in row]
        if not vals or not vals[0]:
            continue
        d = {headers[i]: vals[i] for i in range(len(headers))}
        d["_excel_row"] = excel_row_idx
        rows.append(d)
    return rows


def enrich(row: dict) -> dict:
    rid = str(row["Rule ID"]).strip()
    official = str(row.get("Official rule text") or "")
    # Always fix BTOM-003 mislabel for transaction type
    official = re.sub(
        r"Invoice transaction type\s*\(BTOM-003\)",
        "Invoice transaction type (BTOM-001)",
        official,
        flags=re.I,
    )
    main_raw = str(row.get("Main field") or "")
    when_raw = str(row.get("When it applies (simple)") or "")

    if rid == "IBT-005":
        # Duplicate of IBR-066-OM — keep but mark clearly
        o = OVERRIDES["IBR-066-OM"].copy()
        o["main"] = o["main"] + " [duplicate matrix row — use IBR-066-OM]"
        o["simple"] = (
            "Duplicate of IBR-066-OM (matrix sometimes labeled IBT-005). "
            + o["simple"]
        )
    elif rid in OVERRIDES:
        o = OVERRIDES[rid].copy()
    else:
        o = infer_from_official(rid, main_raw, official, when_raw)

    if "official" not in o:
        o["official"] = official
    else:
        o["official"] = o.get("official") or official

    what = build_what_we_test(o)
    how = build_how_we_test(o)
    return {
        "Rule ID": rid,
        "Rule type": row.get("Rule type"),
        "Main field": o["main"],
        "When it applies (simple)": o["when"],
        "What the rule says (simple English)": o["simple"],
        "What we test": what,
        "How we test (step by step)": how,
        "Official rule text": o["official"],
        "Peppol link": row.get("Peppol link")
        or f"{PEPPOL_BASE}/{rid}/",
        "Analysis sheet row": row.get("Analysis sheet row"),
        "_trigger": o["trigger"],
        "_dependent": o["dependent"],
        "_effect": o["effect"],
        "_value": o.get("value", ""),
        "_simple_rule": build_simple_rule(o),
        "_related": build_related_note(o),
        "_requirement": effect_requirement(o),
    }


def write_sheet(ws, rows: list[dict]) -> None:
    headers = [
        "Rule ID",
        "Rule type",
        "Main field",
        "When it applies (simple)",
        "What the rule says (simple English)",
        "What we test",
        "How we test (step by step)",
        "Official rule text",
        "Peppol link",
        "Analysis sheet row",
    ]
    # Clear data rows
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)
    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        cell.font = Font(bold=True)
    for r in rows:
        ws.append([r.get(h) for h in headers])
    for col in ws.columns:
        for cell in col:
            cell.alignment = Alignment(wrap_text=True, vertical="top")


def write_md(all_rows: list[dict]) -> None:
    by_type: dict[str, list[dict]] = {}
    for r in all_rows:
        t = str(r.get("Rule type") or "Other")
        by_type.setdefault(t, []).append(r)

    order = [
        "Backend",
        "Code list",
        "Conditional (if-then)",
        "Formula / totals",
        "Other",
    ]
    # include any extras
    for t in by_type:
        if t not in order:
            order.append(t)

    lines = [
        "# Oman Conditional Validation — Plain Language Rule Guide",
        "**Purpose:** One easy document so anyone (business, BA, QA, Dev) can understand each rule.",
        "For every rule we answer:",
        "1. **When does it apply?** (plain English)",
        "2. **Rule in simple words** — If *this column* has *this value*, then *that column* must be filled / empty / this exact value.",
        "3. **Optional Excel check steps** (how to verify in an upload) — secondary; full pass/fail packs live in the matrix workbook.",
        "",
        "**Sources:** Analysis sheet *Conditional Validations* + [PINT OM](https://test-docs.peppol.eu/pint/pint-om/2026-Q2-v1.0.1/pint-om/)",
        "**Excel mapping:** Covoro headers via `CONDITIONAL_FIELD_TO_ROW_KEY` / `ALIASES`",
        "",
        f"**Total rules:** {len(all_rows)}",
        "",
        "**Detailed Excel case packs:** `EINV_OMAN_ConditionalValidation_FullMatrix.xlsx`",
        "",
        "---",
        "",
    ]
    for t in order:
        rows = by_type.get(t)
        if not rows:
            continue
        lines.append(f"## {t}")
        lines.append("")
        for r in rows:
            rid = r["Rule ID"]
            lines.append(f"### {rid}")
            lines.append(f"- **Main Excel field(s):** {r['Main field']}")
            lines.append(f"- **When it applies:** {r['When it applies (simple)']}")
            lines.append("")
            lines.append("**Rule in simple words:**  ")
            lines.append(str(r.get("_simple_rule") or r["What the rule says (simple English)"]))
            related = r.get("_related") or ""
            if related:
                lines.append("")
                lines.append(f"**Related columns:** {related}")
            lines.append("")
            lines.append("**In the Excel (plain):**")
            lines.append("")
            lines.append(str(r["What we test"]))
            lines.append("")
            lines.append("<details><summary>Optional: how to check in an upload</summary>")
            lines.append("")
            lines.append(str(r["How we test (step by step)"]))
            lines.append("")
            lines.append("</details>")
            lines.append("")
            lines.append("<details><summary>Official Peppol text (short reference)</summary>")
            lines.append("")
            lines.append(str(r["Official rule text"]))
            lines.append("")
            lines.append(str(r["Peppol link"]))
            lines.append("")
            lines.append("</details>")
            lines.append("")
            lines.append("---")
            lines.append("")
    MD.write_text("\n".join(lines), encoding="utf-8")


class GuidePDF(FPDF):
    def footer(self) -> None:
        self.set_y(-12)
        self.set_font("Helvetica", size=8)
        self.cell(0, 8, f"Page {self.page_no()}/{{nb}}", align="C")


def pdf_safe(text: str) -> str:
    return (
        (text or "")
        .replace("→", "->")
        .replace("≠", "!=")
        .replace("…", "...")
        .replace("—", "-")
        .replace("–", "-")
        .replace("≤", "<=")
        .replace("≥", ">=")
        .encode("latin-1", "replace")
        .decode("latin-1")
    )


def soft_wrap(text: str, width: int = 72) -> str:
    """Insert break opportunities into very long tokens so Helvetica can wrap.

    Character budget is intentionally below A4 usable width at 9–10pt so
    unbroken URLs / codes cannot push past the right margin.
    """
    lines_out: list[str] = []
    for line in (text or "").splitlines():
        out: list[str] = []
        for word in line.split(" "):
            while len(word) > width:
                out.append(word[:width])
                word = word[width:]
            out.append(word)
        lines_out.append(" ".join(out))
    return "\n".join(lines_out)


def _pdf_block(pdf: FPDF, h: float, text: str, *, bold: bool = False, size: float = 9) -> None:
    """Word-wrap a block within page margins, then return to left margin.

    fpdf2 multi_cell defaults to new_x=RIGHT, which leaves the cursor at the
    right edge so the next block draws off-page (clipped). Always reset X.
    """
    pdf.set_font("Helvetica", "B" if bold else "", size)
    pdf.multi_cell(
        0,
        h,
        text,
        align="L",
        new_x="LMARGIN",
        new_y="NEXT",
        wrapmode="WORD",
    )


def write_pdf(all_rows: list[dict]) -> None:
    # >= 0.6" margins (~15.2mm); use 16mm for readable gutters
    margin_mm = 16
    pdf = GuidePDF(orientation="P", unit="mm", format="A4")
    pdf.alias_nb_pages()
    pdf.set_margins(margin_mm, margin_mm, margin_mm)
    pdf.set_auto_page_break(auto=True, margin=18)
    pdf.add_page()

    _pdf_block(pdf, 7, "Oman E-Invoice Conditional Validation Rules", bold=True, size=14)
    _pdf_block(
        pdf,
        5,
        pdf_safe(
            "Plain-language guide aligned with Peppol PINT OM + Covoro Excel column mapping.\n"
            "Each rule is written as: If this field has this value, then that field must be "
            "filled / empty / this exact value.\n"
            "Source: Conditional Validations sheet | Peppol PINT OM 2026-Q2\n"
            f"Total rules: {len(all_rows)}"
        ),
        size=10,
    )
    pdf.ln(2)

    for i, r in enumerate(all_rows, 1):
        rid = r["Rule ID"]
        # Ensure room for heading + a few body lines
        if pdf.get_y() > 255:
            pdf.add_page()
        _pdf_block(pdf, 6, pdf_safe(f"{i}. {rid}"), bold=True, size=11)
        related = r.get("_related") or ""
        related_block = f"\nRelated: {related}" if related else ""
        # Short Peppol note — truncated so it does not dominate
        official = str(r.get("Official rule text") or "").strip()
        if len(official) > 180:
            official = official[:177] + "..."
        body = soft_wrap(
            pdf_safe(
                f"Main Excel field(s): {r['Main field']}\n"
                f"When it applies: {r['When it applies (simple)']}\n"
                f"Rule in simple words: {r.get('_simple_rule') or r['What the rule says (simple English)']}"
                + related_block
                + "\n"
                + (f"Note (Peppol): {official}\n" if official else "")
            )
        )
        _pdf_block(pdf, 4.5, body, size=9)
        pdf.ln(1.5)

    pdf.output(str(PDF))


def main() -> None:
    wb = openpyxl.load_workbook(XLSX)
    all_ws = wb["All Rules Guide"]
    raw = load_rows(all_ws)
    enriched = [enrich(r) for r in raw]

    # Write All Rules Guide
    write_sheet(all_ws, enriched)

    # Rewrite type sheets from enriched
    type_to_sheet = {
        "Conditional (if-then)": "Conditional (if-then)",
        "Formula / totals": "Formula - totals",
        "Code list": "Code list",
        "Backend": "Backend",
    }
    for rtype, sheet_name in type_to_sheet.items():
        if sheet_name not in wb.sheetnames:
            continue
        subset = [r for r in enriched if str(r.get("Rule type")) == rtype]
        write_sheet(wb[sheet_name], subset)

    wb.save(XLSX)
    write_md(enriched)
    write_pdf(enriched)
    print(f"Updated {XLSX.name}: {len(enriched)} rules")
    print(f"Updated {MD.name}")
    print(f"Updated {PDF.name}")
    # Sample if-then lines
    for rid in [
        "ALIGNED-IBRP-028-OM",
        "ALIGNED-IBRP-E-05-OM",
        "IBR-023-OM",
        "IBR-006-OM",
        "CL-10-OM",
        "IBR-175-OM",
    ]:
        r = next(x for x in enriched if x["Rule ID"] == rid)
        print(f"  {rid}: {r['_simple_rule']}")


if __name__ == "__main__":
    main()
