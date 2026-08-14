from pathlib import Path
import re
import openpyxl
from openpyxl import Workbook

input_path = Path(r'C:/Users/Kishor Mane/Downloads/22- GSP-54396 - E Invoice __ OMAN __ As a system, it should validate the fields of E Invoices considering additional Dependent validations whenever the user uploads the excel.xlsx')
output_path = Path(r'C:/Users/Kishor Mane/Downloads/tc condition.xlsx')

rule_codes = [
    'ALIGNED-IBRP-028-OM','ALIGNED-IBRP-E-01-OM','ALIGNED-IBRP-E-05-OM','ALIGNED-IBRP-E-08-OM','ALIGNED-IBRP-E-09-OM',
    'ALIGNED-IBRP-O-01-OM','ALIGNED-IBRP-O-05-OM','ALIGNED-IBRP-O-08-OM','ALIGNED-IBRP-O-09-OM',
    'ALIGNED-IBRP-S-01-OM','ALIGNED-IBRP-S-05-OM','ALIGNED-IBRP-S-08-OM','ALIGNED-IBRP-S-09-OM','ALIGNED-IBRP-S-09-OM-WARN','ALIGNED-IBRP-S-10-OM',
    'ALIGNED-IBRP-Z-01-OM','ALIGNED-IBRP-Z-05-OM','ALIGNED-IBRP-Z-08-OM','ALIGNED-IBRP-Z-09-OM',
    'IBR-CL-05-OM','CL-06-OM','CL-10-OM','CL-11-OM','IBR-004-OM','IBR-005-OM','IBR-006-OM','IBR-007-OM','IBR-010-OM','IBR-012-OM','IBR-013-OM','IBR-014-OM','IBR-015-OM','IBR-016-OM','IBR-017-OM','IBR-019-OM','IBR-020-OM','IBR-023-OM','IBR-032-OM','IBR-033-OM','IBR-034-OM','IBR-035-OM','IBR-036-OM','IBR-037-OM','IBR-038-OM','IBR-039-OM','IBR-040-OM','IBR-041-OM','IBR-042-OM','IBR-045-OM','IBR-046-OM','IBR-047-OM','IBR-053-OM','IBR-054-OM','IBR-056-OM','IBR-057-OM','IBR-058-OM','IBR-059-OM','IBR-061-OM','IBR-062-OM','IBR-064-OM','IBR-065-OM','IBR-066-OM','IBR-067-OM','IBR-069-OM','IBR-070-OM','IBR-072-OM','IBR-073-OM','IBR-074-OM','IBR-077-OM','IBR-078-OM','IBR-079-OM','IBR-080-OM','IBR-081-OM','IBR-082-OM','IBR-084-OM','IBR-085-OM','IBR-086-OM','IBR-087-OM','IBR-091-OM','IBR-092-OM','IBR-093-OM','IBR-094-OM','IBR-095-OM','IBR-096-OM','IBR-097-OM','IBR-098-OM','IBR-099-OM','IBR-100-OM','IBR-104-OM','IBR-137-OM','IBR-138-OM','IBR-139-OM','IBR-140-OM','IBR-141-OM','IBR-142-OM','IBR-143-OM','IBR-144-OM','IBR-145-OM','IBR-146-OM','IBR-147-OM','IBR-148-OM','IBR-149-OM','IBR-150-OM','IBR-151-OM','IBR-152-OM','IBR-153-OM','IBR-155-OM','IBR-158-OM','IBR-160-OM','IBR-172-OM','IBR-173-OM','IBR-174-OM','IBR-175-OM','IBR-176-OM','IBR-177-OM','IBR-029','IBR-030','IBR-053','IBR-077','IBR-093','IBR-CO-19','IBR-CO-20','ALIGNED-IBRP-048','IBR-168-OM','IBR-168-OM-WARN'
]

wb = openpyxl.load_workbook(input_path, data_only=True)
ws = wb.active

code_map = {code: [] for code in rule_codes}
extra_codes = {}
for row in ws.iter_rows(min_row=2, values_only=True):
    if not row or not row[0]:
        continue
    text = ' '.join(str(cell) for cell in row if cell is not None)
    for code in rule_codes:
        if code in text:
            code_map[code].append(row[0])
    for match in re.findall(r'\[([A-Za-z0-9\-_.]+)\]', text):
        if match not in code_map:
            extra_codes[match] = extra_codes.get(match, 0) + 1

out_wb = Workbook()
out_ws = out_wb.active
out_ws.title = 'Rule Coverage'
out_ws.append(['Rule Code', 'Present', 'TC Count', 'Sample TC IDs'])
for code in rule_codes:
    ids = code_map[code]
    out_ws.append([code, 'Yes' if ids else 'No', len(ids), ', '.join(ids[:5])])

out_wb.save(output_path)
print(f'Saved rule coverage to {output_path}')
print('Missing codes:', [code for code, ids in code_map.items() if not ids])
print('Extra bracket codes sample:', sorted(extra_codes.items())[:20])
