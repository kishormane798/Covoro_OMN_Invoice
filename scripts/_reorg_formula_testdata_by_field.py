#!/usr/bin/env python3
"""Reorganize formula TestData packs from SECTION folders to FIELD_NAME folders."""
from __future__ import annotations

import re
import shutil
from pathlib import Path

import openpyxl

ROOT = Path("testcase/formula_validation/TestData")
MATRIX = Path("testcase/formula_validation/EINV_OMAN_FormulaValidation_FullMatrix.xlsx")


def folder_name(field: str) -> str:
    cleaned = re.sub(r"[()[\]{}]", " ", field or "UNKNOWN_FIELD")
    cleaned = re.sub(r"[^\w\s.-]+", " ", cleaned).strip() or "UNKNOWN_FIELD"
    return re.sub(r"\s+", "_", cleaned).upper()


def main() -> int:
    wb = openpyxl.load_workbook(MATRIX, data_only=True)
    ws = wb["All Testcases"]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = [
        dict(zip(headers, r))
        for r in ws.iter_rows(min_row=2, values_only=True)
        if r[0]
    ]

    existing: dict[str, Path] = {}
    for p in ROOT.rglob("TC-*.xlsx"):
        if "_tmp" in p.parts:
            continue
        existing[p.stem.upper()] = p

    moved = 0
    missing: list[str] = []
    for r in rows:
        pol = str(r.get("Polarity") or "").lower()
        if pol == "gap":
            continue
        tc_id = str(r["Test Case ID"]).strip()
        tc_key = tc_id.upper()
        field = str(r.get("Field name") or "").strip()
        bucket = "positive" if pol == "positive" else "negative"
        dest_dir = ROOT / folder_name(field) / bucket
        dest_dir.mkdir(parents=True, exist_ok=True)
        dest = dest_dir / f"{tc_id}.xlsx"
        src = existing.get(tc_key)
        if not src:
            missing.append(tc_id)
            continue
        if src.resolve() == dest.resolve():
            continue
        if dest.exists() and dest.resolve() != src.resolve():
            dest.unlink()
        shutil.move(str(src), str(dest))
        moved += 1

    removed = 0
    for d in sorted(ROOT.rglob("*"), reverse=True):
        if not d.is_dir() or d.name == "_tmp" or d == ROOT:
            continue
        try:
            next(d.iterdir())
        except StopIteration:
            d.rmdir()
            removed += 1

    folders = sorted(p.name for p in ROOT.iterdir() if p.is_dir() and p.name != "_tmp")
    xlsx = len(list(ROOT.rglob("TC-*.xlsx")))
    print(
        {
            "moved": moved,
            "missing": missing,
            "empty_removed": removed,
            "folders": folders,
            "xlsx": xlsx,
        }
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
