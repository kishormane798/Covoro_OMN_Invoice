#!/usr/bin/env python3
"""Export remaining unmapped rules with exact titles for alias suggestions."""
from __future__ import annotations

import re
from collections import defaultdict
from pathlib import Path

import openpyxl

# Keep aliases/PSEUDO in sync with scripts/list_unmapped_conditional_rules.py
ns: dict = {}
src = Path(__file__).with_name("list_unmapped_conditional_rules.py").read_text(
    encoding="utf-8"
)
# Import helpers + skip sets (everything before main).
exec(src.split("def main")[0], ns)
ALIASES = ns["ALIASES"]
PSEUDO = ns["PSEUDO"]
RULE_RE = ns["RULE_RE"]
INTENTIONAL_SKIP_RULE_IDS = ns.get("INTENTIONAL_SKIP_RULE_IDS", set())
norm = ns["norm"]
strip_peppol_suffix = ns["strip_peppol_suffix"]
is_intentional_skip_field = ns["is_intentional_skip_field"]
cell = ns["cell"]


def main() -> None:
    root = Path(__file__).resolve().parents[1]
    template = root / "testData" / "uploads" / "template.xlsx"
    matrix = (
        root
        / "testcase"
        / "conditional_validation"
        / "EINV_OMAN_ConditionalValidation_FullMatrix.xlsx"
    )
    out = (
        root
        / "testcase"
        / "conditional_validation"
        / "UNMAPPED_RULES_EXACT_TITLES.md"
    )

    wb = openpyxl.load_workbook(template, read_only=True, data_only=True)
    ws = wb["E Invoice"]
    row4 = list(next(ws.iter_rows(min_row=4, max_row=4, values_only=True)))
    header_norm = {norm(str(h)): str(h).strip() for h in row4 if h}
    wb.close()

    def mappable(field: str, title: str = "", description: str = "") -> bool:
        f = (field or "").strip()
        if not f:
            return False
        base = strip_peppol_suffix(f) or f
        candidates = [f, base] if base != f else [f]
        m_rule = RULE_RE.search(title or "")
        if m_rule and m_rule.group(1).strip().upper() in INTENTIONAL_SKIP_RULE_IDS:
            return True  # intentional skip (formula suite / out of scope)
        blob = f"{title}\n{description}".lower()
        if (
            any(norm(c) == "conditional fields" for c in candidates)
            and (
                "invoice line period" in blob
                or "ibt-134" in blob
                or "ibt-135" in blob
                or "ibg-26" in blob
            )
        ):
            return True  # intentional skip — no Covoro column (backend default)
        if is_intentional_skip_field(f):
            return True  # intentional skip — backend default / formula / out of scope
        # ALIGNED-IBRP-O/Z-01 Simplified Tax Invoice exception → Invoice Transaction Type Code.
        if (
            any(norm(c) == "an invoice that contains an invoice line" for c in candidates)
            and "simplified tax invoice exception" in blob
        ):
            return True
        if any(c in ALIASES for c in candidates):
            return True
        if any(PSEUDO.search(c) for c in candidates):
            return False
        for c in candidates:
            if norm(ALIASES.get(c, c)) in header_norm:
                return True
        return False

    wb = openpyxl.load_workbook(matrix, read_only=True, data_only=True)
    ws = wb["All Testcases"]
    rows = ws.iter_rows(values_only=True)
    headers = [str(c).strip() if c is not None else "" for c in next(rows)]
    col = {h: i for i, h in enumerate(headers) if h}
    if "Test Case ID" not in col:
        raise KeyError(
            f"FullMatrix missing 'Test Case ID'. Found headers: {headers}"
        )
    if not any(h in col for h in ("Test Case Title", "Testcase Title", "Test Cases Title")):
        raise KeyError(
            "FullMatrix missing title column "
            "('Test Case Title' / 'Testcase Title'). "
            f"Found headers: {headers}"
        )

    # field -> ruleId -> samples
    by_field: dict[str, dict[str, list[dict]]] = defaultdict(
        lambda: defaultdict(list)
    )

    for row in rows:
        if not row:
            continue
        tc = cell(row, col, "Test Case ID")
        if not tc or tc.lower() == "test case id":
            continue
        field = cell(row, col, "Filed name", "Field name")
        title = cell(row, col, "Test Case Title", "Testcase Title", "Test Cases Title")
        desc = cell(row, col, "Test Cases Description")
        if mappable(field, title, desc):
            continue
        sec = cell(row, col, "Section")
        pol = cell(row, col, "Polarity")
        m = RULE_RE.search(title)
        rule = m.group(1).strip() if m else "(no-ruleId-in-title)"
        samples = by_field[field][rule]
        # one sample per polarity for this rule
        if any(s["pol"] == pol for s in samples):
            continue
        samples.append(
            {
                "tc": tc,
                "pol": pol,
                "sec": sec,
                "title": title,
                "desc": desc[:320],
            }
        )
    wb.close()

    lines: list[str] = [
        "# Remaining unmapped rules — exact titles for alias suggestions",
        "",
        "For each matrix **Filed name**, the Peppol/rule IDs and sample testcase titles.",
        "Fill **Suggested Excel header(s)** (one column, or list multiple if the rule needs several).",
        "",
        f"Remaining field labels: **{len(by_field)}**",
        "",
    ]

    for field in sorted(
        by_field.keys(),
        key=lambda f: (-len(by_field[f]), f),
    ):
        rules = by_field[field]
        safe = field.replace("|", "/")
        lines.append(f"## Field label: `{safe}`")
        lines.append("")
        lines.append(f"- Distinct ruleIds: **{len(rules)}**")
        lines.append("- **Suggested Excel header(s):** `TBD`")
        lines.append("")
        for rule in sorted(rules.keys()):
            lines.append(f"### `{rule}`")
            for s in rules[rule]:
                lines.append(f"- `{s['tc']}` ({s['pol']}, {s['sec']})")
                lines.append(f"  - Title: {s['title']}")
                if s["desc"]:
                    lines.append(f"  - Description: {s['desc']}")
            lines.append("")

    out.write_text("\n".join(lines), encoding="utf-8")
    print(f"wrote {out}")
    print(f"fields={len(by_field)}")
    for f in sorted(by_field.keys(), key=lambda x: -len(by_field[x])):
        print(f"{len(by_field[f]):2d} rules | {f}")


if __name__ == "__main__":
    main()
