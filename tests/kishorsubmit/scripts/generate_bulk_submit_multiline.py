"""
Multiline Covoro submit Excel generator (local one-off).

Uses the first case in SubmitInvoiceMultiItem.compact.json (5 lines / invoice).
Invoice #: INV-{epochMs}{worker}{seq}

Excel .xlsx hard limit: 1,048,576 rows/sheet. Template data starts at row 6,
so max data rows = 1,048,571. With 5 lines/invoice -> max invoices = 209,714.

Default batches (no args): 10k, 20k, 30k, ... (step 10k) up to max, last = excel_max
Or pass counts: 10000 20000 30000  (also accepts 10k / 20k / max)

Writes to Downloads:
  submit_multiline_invoices_10k.xlsx / 20k / ... / submit_multiline_invoices_excel_max.xlsx

Run from repo root:
  python tests/kishorsubmit/scripts/generate_bulk_submit_multiline.py
  python tests/kishorsubmit/scripts/generate_bulk_submit_multiline.py 10000
  python tests/kishorsubmit/scripts/generate_bulk_submit_multiline.py 10k 20k 30k max
"""

from __future__ import annotations

import json
import math
import os
import re
import sys
import time
from datetime import datetime
from typing import Any

from openpyxl import load_workbook

REPO_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), "../../.."))
UTILS_DIR = os.path.join(REPO_ROOT, "utils")
if UTILS_DIR not in sys.path:
    sys.path.insert(0, UTILS_DIR)

from invoice_excel_writer import (  # noqa: E402
    DOCUMENT_LEVEL_CLEARED_FOR_VAT_REVERSE_CHARGE,
    EXEMPT_BLANK_TAX_FIELD_HEADERS,
    INVOICE_SUBMIT_CLEAR_LAST_COL,
    INVOICE_TEMPLATE_DATA_ROW,
    VAT_REVERSE_CHARGE_TAX_CATEGORY,
    _set_by_header_name,
    clear_cell_optional,
    effective_tax_rate,
    get_col,
    get_header_map,
    is_exempt_from_tax_tax_category,
    normalize_category,
)

SHEET_NAME = "E Invoice"
HEADER_ROW = 4
DATA_ROW = INVOICE_TEMPLATE_DATA_ROW
TEMPLATE_PATH = os.path.join(REPO_ROOT, "testData", "uploads", "template.xlsx")
MULTI_ITEM_JSON = os.path.join(
    os.path.dirname(__file__), "..", "testData", "SubmitInvoiceMultiItem.compact.json"
)
OUTPUT_DIR = r"C:\Users\Kishor Mane\Downloads"
ISSUE_DATE_FORMAT = "yyyy-mm-dd"
WORKER = 0

# Excel .xlsx worksheet row ceiling (inclusive).
EXCEL_MAX_ROW = 1_048_576
# Usable data rows when first invoice starts at DATA_ROW (row 6).
MAX_DATA_ROWS = EXCEL_MAX_ROW - DATA_ROW + 1  # 1_048_571
BATCH_STEP = 10_000


def default_counts(max_invoices: int) -> list[int]:
    """10k, 20k, ... up to largest multiple of 10k <= max, then excel_max last."""
    counts = list(range(BATCH_STEP, max_invoices + 1, BATCH_STEP))
    if not counts or counts[-1] != max_invoices:
        counts.append(max_invoices)
    return counts


CALCULATED_HEADERS = (
    "Item net price",
    "Invoice line net amount",
    "Invoice line amount in AED",
    "VAT line amount in AED",
    "Sum of Invoice line net amount",
    "Invoice total amount without tax",
    "Invoice total tax amount",
    "Invoice total amount with tax",
    "Amount due for payment",
    "Invoice total tax amount in tax accounting currency",
)


def build_unique_submit_invoice_number(seq: int, base_ms: int, worker: int = WORKER) -> str:
    suffix = f"{max(0, worker)}{str(seq).zfill(5)}"
    raw = f"INV-{base_ms}{suffix}"
    return raw if len(raw) <= 64 else raw[:64]


def _to_number(value: Any, fallback: float = 0.0) -> float:
    if value is None or str(value).strip() == "":
        return fallback
    try:
        parsed = float(value)
    except (TypeError, ValueError):
        return fallback
    return parsed if parsed == parsed else fallback


def _fix6(n: float) -> float:
    return float(f"{n:.6f}")


def _ceil2(n: float) -> float:
    return math.ceil(n * 100 - 1e-12) / 100


def load_first_multiline_case() -> list[dict[str, Any]]:
    with open(MULTI_ITEM_JSON, encoding="utf-8-sig") as f:
        cases = json.load(f)
    if not cases:
        raise SystemExit(f"No multi-item cases in {MULTI_ITEM_JSON}")
    tc = cases[0]
    common = dict(tc.get("common") or {})
    for h in CALCULATED_HEADERS:
        common.pop(h, None)
    lines = tc.get("lines") or []
    if not lines:
        raise SystemExit(f"First multi-item case has no lines: {tc.get('name')}")
    rows: list[dict[str, Any]] = []
    for line in lines:
        row = {**common, **line}
        for h in CALCULATED_HEADERS:
            row.pop(h, None)
        rows.append(row)
    print(f"Using multi-item case: {tc.get('name')} ({len(rows)} lines / invoice)")
    return rows


def apply_multiline_totals(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Mirror utils/invoiceExcel.ts generateInvoiceFromSubmitRows totals."""
    currency_code = str(rows[0].get("Invoice Currency Code") or "AED").strip() or "AED"
    currency_rate = _to_number(rows[0].get("Currency Exchange Rate"), 1.0) or 1.0
    doc_charges = _to_number(rows[0].get("Charges on document level"), 0.0)
    doc_allowances = _to_number(rows[0].get("Allowances on document level"), 0.0)
    paid_amount = _to_number(rows[0].get("Paid amount"), 0.0)
    rounding_amount = _to_number(rows[0].get("Rounding amount"), 0.0)

    first_rate = effective_tax_rate(
        rows[0].get("Tax Category"),
        _to_number(rows[0].get("Tax Rate"), 0.0),
    )
    doc_charge_tax_raw = _fix6(doc_charges * (first_rate / 100.0))
    doc_allowance_tax_raw = _fix6(doc_allowances * (first_rate / 100.0))

    per_line: list[dict[str, Any]] = []
    for r in rows:
        tax_cat = r.get("Tax Category")
        rate = effective_tax_rate(tax_cat, _to_number(r.get("Tax Rate"), 0.0))
        base_qty = _to_number(r.get("Item price base quantity"), 1.0)
        gross = _to_number(r.get("Item gross price"), 0.0)
        discount = _to_number(r.get("Item price discount"), 0.0)
        qty = _to_number(r.get("Invoiced quantity"), 0.0)
        line_charge = _to_number(r.get("Invoice line charge amount"), 0.0)
        line_allowance = _to_number(r.get("Invoice line allowance amount"), 0.0)

        raw_item_net = _fix6(gross - discount)
        if base_qty > 0:
            raw_line_net = _fix6((raw_item_net * qty) / base_qty + line_charge - line_allowance)
        else:
            raw_line_net = _fix6(line_charge - line_allowance)
        raw_vat = _fix6(raw_line_net * (rate / 100.0))
        line_amount_raw = _fix6(raw_line_net + raw_vat)
        if currency_code == "AED":
            line_aed = _ceil2(line_amount_raw)
            vat_aed = _ceil2(raw_vat)
        else:
            line_aed = _ceil2(_fix6(line_amount_raw * currency_rate))
            vat_aed = _ceil2(_fix6(raw_vat * currency_rate))

        per_line.append(
            {
                "raw_line_net": raw_line_net,
                "raw_vat": raw_vat,
                "item_net": _ceil2(raw_item_net),
                "line_net": _ceil2(raw_line_net),
                "line_aed": line_aed,
                "vat_aed": None if is_exempt_from_tax_tax_category(tax_cat) else vat_aed,
            }
        )

    raw_sum_line_net = _fix6(sum(p["raw_line_net"] for p in per_line))
    raw_sum_vat = _fix6(sum(p["raw_vat"] for p in per_line))
    raw_total_without = _fix6(raw_sum_line_net + doc_charges - doc_allowances)
    raw_total_tax = _fix6(raw_sum_vat + doc_charge_tax_raw - doc_allowance_tax_raw)
    raw_total_with = _fix6(raw_total_without + raw_total_tax)

    invoice_total_without = _ceil2(raw_total_without)
    invoice_total_tax = _ceil2(raw_total_tax)
    invoice_total_with = _ceil2(raw_total_with)
    amount_due = _ceil2(_fix6(raw_total_with - paid_amount + rounding_amount))
    sum_line_net = _ceil2(raw_sum_line_net)
    tax_acct = None if currency_code == "AED" else _ceil2(_fix6(raw_total_tax * currency_rate))

    out: list[dict[str, Any]] = []
    for r, calc in zip(rows, per_line):
        row = dict(r)
        row["Item net price"] = calc["item_net"]
        row["Invoice line net amount"] = calc["line_net"]
        row["Invoice line amount in AED"] = calc["line_aed"]
        row["VAT line amount in AED"] = calc["vat_aed"]
        row["Sum of Invoice line net amount"] = sum_line_net
        row["Invoice total amount without tax"] = invoice_total_without
        row["Invoice total tax amount"] = invoice_total_tax
        row["Invoice total amount with tax"] = invoice_total_with
        row["Amount due for payment"] = amount_due
        row["Invoice total tax amount in tax accounting currency"] = tax_acct
        out.append(row)
    return out


def _clear_submit_row(ws, row_number: int) -> None:
    for col in range(1, INVOICE_SUBMIT_CLEAR_LAST_COL + 1):
        cell = ws.cell(row=row_number, column=col)
        cell.value = None
        cell.number_format = "General"


def _copy_row_values(ws, source_row: int, target_row: int, max_col: int) -> None:
    for col in range(1, max_col + 1):
        src = ws.cell(row=source_row, column=col)
        dst = ws.cell(row=target_row, column=col)
        dst.value = src.value
        dst.number_format = src.number_format


def _write_line_row(
    ws,
    header_map: dict[str, int],
    row_number: int,
    invoice_col: int,
    date_col: int,
    invoice_number: str,
    today: datetime,
    row_values: dict[str, Any],
    clear: bool,
) -> None:
    if clear:
        _clear_submit_row(ws, row_number)
    ws.cell(row=row_number, column=invoice_col).value = invoice_number
    ws.cell(row=row_number, column=date_col).value = today
    ws.cell(row=row_number, column=date_col).number_format = ISSUE_DATE_FORMAT

    for key, value in row_values.items():
        if value is None:
            continue
        _set_by_header_name(ws, HEADER_ROW, row_number, str(key), value)

    tax_cat = row_values.get("Tax Category")
    if normalize_category(tax_cat) == normalize_category(VAT_REVERSE_CHARGE_TAX_CATEGORY):
        for hdr in DOCUMENT_LEVEL_CLEARED_FOR_VAT_REVERSE_CHARGE:
            _set_by_header_name(ws, HEADER_ROW, row_number, hdr, "")
    if is_exempt_from_tax_tax_category(tax_cat):
        for header in EXEMPT_BLANK_TAX_FIELD_HEADERS:
            clear_cell_optional(ws, row_number, header_map, header)


def write_batch(
    invoice_count: int,
    file_name: str,
    base_ms: int,
    line_templates: list[dict[str, Any]],
) -> str:
    lines_per_invoice = len(line_templates)
    wb = load_workbook(TEMPLATE_PATH)
    if SHEET_NAME not in wb.sheetnames:
        raise SystemExit(f"Sheet '{SHEET_NAME}' not found in template")
    ws = wb[SHEET_NAME]
    header_map = get_header_map(ws, HEADER_ROW)
    invoice_col = get_col(header_map, "Invoice Number")
    date_col = get_col(header_map, "Invoice Issue Date")
    max_col = max(ws.max_column, INVOICE_SUBMIT_CLEAR_LAST_COL)
    today = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)

    first_inv = build_unique_submit_invoice_number(1, base_ms)

    for line_idx, row_values in enumerate(line_templates):
        row_number = DATA_ROW + line_idx
        _write_line_row(
            ws,
            header_map,
            row_number,
            invoice_col,
            date_col,
            first_inv,
            today,
            row_values,
            clear=True,
        )

    for inv_seq in range(2, invoice_count + 1):
        inv_no = build_unique_submit_invoice_number(inv_seq, base_ms)
        dest_start = DATA_ROW + (inv_seq - 1) * lines_per_invoice
        for line_idx in range(lines_per_invoice):
            src = DATA_ROW + line_idx
            dst = dest_start + line_idx
            _copy_row_values(ws, src, dst, max_col)
            ws.cell(row=dst, column=invoice_col).value = inv_no
            ws.cell(row=dst, column=date_col).value = today
            ws.cell(row=dst, column=date_col).number_format = ISSUE_DATE_FORMAT
        if inv_seq % 5_000 == 0 or inv_seq == invoice_count:
            print(f"  {file_name}: invoices through {inv_seq}/{invoice_count}...", flush=True)

    out_path = os.path.join(OUTPUT_DIR, file_name)
    wb.save(out_path)
    wb.close()
    excel_rows = invoice_count * lines_per_invoice
    print(
        f"Saved {invoice_count} invoices ({excel_rows} Excel rows, {lines_per_invoice} lines each) -> {out_path}"
    )
    print(f"  sample invoice #: {first_inv}")
    return out_path


def parse_count(raw: str, max_invoices: int) -> int:
    text = raw.strip().lower().replace(",", "").replace("_", "")
    if text in {"max", "excel_max", "excel-max"}:
        return max_invoices
    match = re.fullmatch(r"(\d+)(k)?", text)
    if not match:
        raise SystemExit(f"Invalid count '{raw}'. Use e.g. 10000, 10k, or max")
    value = int(match.group(1))
    if match.group(2):
        value *= 1_000
    if value < 1:
        raise SystemExit(f"Count must be >= 1, got {value}")
    if value > max_invoices:
        raise SystemExit(
            f"Count {value} exceeds Excel limit of {max_invoices} multiline invoices "
            f"(max row {EXCEL_MAX_ROW}, data from row {DATA_ROW})"
        )
    return value


def count_label(count: int, max_invoices: int | None = None) -> str:
    if max_invoices is not None and count == max_invoices:
        return "excel_max"
    if count % 1_000 == 0 and count >= 1_000:
        return f"{count // 1_000}k"
    return str(count)


def resolve_batches(argv: list[str], max_invoices: int) -> list[tuple[int, str]]:
    if argv:
        counts = [parse_count(a, max_invoices) for a in argv]
    else:
        counts = default_counts(max_invoices)
    if not counts:
        raise SystemExit(f"No valid batch sizes under Excel limit {max_invoices}")
    return [
        (c, f"submit_multiline_invoices_{count_label(c, max_invoices)}.xlsx")
        for c in counts
    ]


def main() -> None:
    if not os.path.isfile(TEMPLATE_PATH):
        raise SystemExit(f"Template not found: {TEMPLATE_PATH}")
    if not os.path.isfile(MULTI_ITEM_JSON):
        raise SystemExit(f"Multi-item JSON not found: {MULTI_ITEM_JSON}")

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    base_rows = load_first_multiline_case()
    line_templates = apply_multiline_totals(base_rows)
    lines_per_invoice = len(line_templates)
    if lines_per_invoice < 1:
        raise SystemExit("No line templates after totals")
    max_invoices = MAX_DATA_ROWS // lines_per_invoice
    batches = resolve_batches(sys.argv[1:], max_invoices)
    print(
        f"Excel limit: max {max_invoices} multiline invoices "
        f"({lines_per_invoice} lines each; row {DATA_ROW}..{EXCEL_MAX_ROW})"
    )
    print(
        f"Batches: {', '.join(count_label(c, max_invoices) for c, _ in batches)}"
    )

    base_ms = int(time.time() * 1000)
    for idx, (count, file_name) in enumerate(batches):
        write_batch(count, file_name, base_ms + idx * 100_000, line_templates)


if __name__ == "__main__":
    main()
