"""
Single-line Covoro submit Excel generator (local one-off).

Uses invoiceData[0]-style Standard Tax Invoice row.
Invoice #: INV-{epochMs}{worker}{seq}

Excel .xlsx hard limit: 1,048,576 rows/sheet. Template data starts at row 6,
so max invoices (1 Excel row each) = 1,048,571.

Default batches (no args): 10k, 20k, 30k, ... (step 10k) up to max, last = excel_max
Or pass counts: 10000 20000 30000  (also accepts 10k / 20k / max)

Writes to Downloads:
  submit_invoices_10k.xlsx / 20k / ... / submit_invoices_excel_max.xlsx

Run from repo root:
  python tests/kishorsubmit/scripts/generate_bulk_submit_single_line.py
  python tests/kishorsubmit/scripts/generate_bulk_submit_single_line.py 10000
  python tests/kishorsubmit/scripts/generate_bulk_submit_single_line.py 10k 20k 30k max
"""

from __future__ import annotations

import os
import re
import sys
import time
from datetime import datetime

from openpyxl import load_workbook

REPO_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), "../../.."))
UTILS_DIR = os.path.join(REPO_ROOT, "utils")
if UTILS_DIR not in sys.path:
    sys.path.insert(0, UTILS_DIR)

from invoice_excel_writer import (  # noqa: E402
    INVOICE_SUBMIT_CLEAR_LAST_COL,
    INVOICE_TEMPLATE_DATA_ROW,
    _set_by_header_name,
    apply_invoice_calculations_to_data_row,
    get_col,
    get_header_map,
)

SHEET_NAME = "E Invoice"
HEADER_ROW = 4
DATA_ROW = INVOICE_TEMPLATE_DATA_ROW
TEMPLATE_PATH = os.path.join(REPO_ROOT, "testData", "uploads", "template.xlsx")
OUTPUT_DIR = r"C:\Users\Kishor Mane\Downloads"
ISSUE_DATE_FORMAT = "yyyy-mm-dd"
WORKER = 0

# Excel .xlsx worksheet row ceiling (inclusive).
EXCEL_MAX_ROW = 1_048_576
# Usable data rows when first invoice starts at DATA_ROW (row 6).
MAX_DATA_ROWS = EXCEL_MAX_ROW - DATA_ROW + 1  # 1_048_571
# One invoice = one Excel row for this generator.
MAX_INVOICES = MAX_DATA_ROWS
BATCH_STEP = 10_000


def default_counts(max_invoices: int = MAX_INVOICES) -> list[int]:
    """10k, 20k, ... up to largest multiple of 10k <= max, then excel_max last."""
    counts = list(range(BATCH_STEP, max_invoices + 1, BATCH_STEP))
    if not counts or counts[-1] != max_invoices:
        counts.append(max_invoices)
    return counts


# First row from tests/kishorsubmit/testData/SubmitInvoice.ts (invoiceData[0]).
BASE_SUBMIT_ROW: dict[str, object] = {
    "Invoice Type Code": "Commercial Invoice",
    "Invoice Transaction Type Code": "Standard Tax Invoice",
    "Invoice Currency Code": "AED",
    "Credit note reason code": "",
    "Invoicing period start date": "01-01-2026",
    "Invoicing period end date": "31-12-2026",
    "Invoice note": "test user1",
    "Principle ID": "112345678900003",
    "Seller name": "PP Bafna Ventures LTD",
    "Seller legal registration identifier": "112345678900003",
    "Seller legal registration identifier type": "Passport",
    "Seller - Passport issuing Country code": "United Arab Emirates",
    "Seller VAT Identifier (TRN / TIN)": "100821338900003",
    "Seller electronic address Scheme": "UAE Tax Identification Number (TIN)",
    "Seller electronic address": "1008213389",
    "Seller address line 1": "Office 101, Business Bay Tower",
    "Seller city": "Dubai",
    "Seller country subdivision": "Dubai",
    "Seller country code": "United Arab Emirates",
    "Buyer name": "Prashant",
    "Buyer identifier": "1000091919v1203",
    "Scheme identifier": "UAE Tax Identification Number (TIN)",
    "Buyer legal registration identifier": "112345679000001",
    "Buyer legal registration identifier type": "Commercial/Trade License",
    "Buyer - Authority name": "Trade Authority",
    "Buyer VAT identifier": "100009191900003",
    "Buyer electronic address Scheme": "UAE Tax Identification Number (TIN)",
    "Buyer electronic address": "1200020015",
    "Beneficiary ID": "1000091919v1203",
    "Buyer address line 1": "Warehouse 12",
    "Buyer city": "Abu Dhabi",
    "Buyer country subdivision": "Abu Dhabi",
    "Buyer country code": "United Arab Emirates",
    "Deliver to party name": "",
    "Deliver to location identifier": "",
    "Delivery to location Scheme": "",
    "Actual delivery date": "",
    "Deliver to address line 1": "",
    "Deliver to address line 2": "",
    "Deliver to address line 3": "",
    "Deliver to city": "",
    "Deliver to post code": "",
    "Deliver to country sub-division": "",
    "Deliver to country code": "",
    "Invoice line identifier": "LINE-S-001",
    "Item Type": "Goods",
    "Type of goods or services subject to RCM": "",
    "Item standard identifier": "3245678",
    "Item classification - Scheme Identifier": (
        "Harmonised systemThe item number is part of, or is generated in the context of the "
        "Harmonised Commodity Description and Coding System (Harmonised System), as developed "
        "and maintained by the World Customs Organization (WCO)."
    ),
    "Item classification identifier": "AA",
    "Service Accounting code": "",
    "Item name": "Laptop Model X",
    "Item description": "Business laptop",
    "Item price base quantity": "1",
    "Item gross price": "1000",
    "Item price discount": "1",
    "Invoiced quantity": "10",
    "Invoiced quantity unit of measure code": "becquerel per kilogram",
    "Invoice line charge amount": "",
    "Invoice line allowance amount": "",
    "Tax Category": "Standard rate.",
    "Tax Rate": "5",
    "Tax exemption reason text": "",
    "Tax exemption reason code": "",
    "Charges on document level": "",
    "Vat category - charges": "",
    "Tax exemption reason - charges": "",
    "Allowances on document level": "",
    "Vat category - allowances": "",
    "Tax exemption reason - allowances": "",
    "Paid amount": "",
    "Rounding amount": "",
    "Payment means type code": "Instrument not defined",
    "Payment due date": "2026-03-21",
}


def build_unique_submit_invoice_number(seq: int, base_ms: int, worker: int = WORKER) -> str:
    suffix = f"{max(0, worker)}{str(seq).zfill(5)}"
    raw = f"INV-{base_ms}{suffix}"
    return raw if len(raw) <= 64 else raw[:64]


def _clear_submit_row(ws, row_number: int) -> None:
    for col in range(1, INVOICE_SUBMIT_CLEAR_LAST_COL + 1):
        cell = ws.cell(row=row_number, column=col)
        cell.value = None
        cell.number_format = "General"


def _copy_row_values(ws, source_row: int, target_row: int, max_col: int) -> None:
    for col in range(1, max_col + 1):
        src = ws.cell(row=source_row, column=col)
        dst = ws.cell(row=target_row, column=col)
        dst.value = src.value
        dst.number_format = src.number_format


def write_batch(invoice_count: int, file_name: str, base_ms: int) -> str:
    wb = load_workbook(TEMPLATE_PATH)
    if SHEET_NAME not in wb.sheetnames:
        raise SystemExit(f"Sheet '{SHEET_NAME}' not found in template")
    ws = wb[SHEET_NAME]
    header_map = get_header_map(ws, HEADER_ROW)
    invoice_col = get_col(header_map, "Invoice Number")
    date_col = get_col(header_map, "Invoice Issue Date")
    max_col = max(ws.max_column, INVOICE_SUBMIT_CLEAR_LAST_COL)
    today = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)

    first_inv = build_unique_submit_invoice_number(1, base_ms)
    _clear_submit_row(ws, DATA_ROW)
    ws.cell(row=DATA_ROW, column=invoice_col).value = first_inv
    ws.cell(row=DATA_ROW, column=date_col).value = today
    ws.cell(row=DATA_ROW, column=date_col).number_format = ISSUE_DATE_FORMAT

    for key, value in BASE_SUBMIT_ROW.items():
        if key == "None":
            continue
        _set_by_header_name(ws, HEADER_ROW, DATA_ROW, str(key), value)

    apply_invoice_calculations_to_data_row(ws, HEADER_ROW, DATA_ROW)

    for i in range(2, invoice_count + 1):
        row_number = DATA_ROW + (i - 1)
        _copy_row_values(ws, DATA_ROW, row_number, max_col)
        ws.cell(row=row_number, column=invoice_col).value = build_unique_submit_invoice_number(
            i, base_ms
        )
        ws.cell(row=row_number, column=date_col).value = today
        ws.cell(row=row_number, column=date_col).number_format = ISSUE_DATE_FORMAT
        if i % 10_000 == 0 or i == invoice_count:
            print(f"  {file_name}: wrote through {i}/{invoice_count}...", flush=True)

    out_path = os.path.join(OUTPUT_DIR, file_name)
    wb.save(out_path)
    wb.close()
    print(f"Saved {invoice_count} invoices -> {out_path}")
    print(f"  sample invoice #: {first_inv}")
    return out_path


def parse_count(raw: str) -> int:
    text = raw.strip().lower().replace(",", "").replace("_", "")
    if text in {"max", "excel_max", "excel-max"}:
        return MAX_INVOICES
    match = re.fullmatch(r"(\d+)(k)?", text)
    if not match:
        raise SystemExit(f"Invalid count '{raw}'. Use e.g. 10000, 10k, or max")
    value = int(match.group(1))
    if match.group(2):
        value *= 1_000
    if value < 1:
        raise SystemExit(f"Count must be >= 1, got {value}")
    if value > MAX_INVOICES:
        raise SystemExit(
            f"Count {value} exceeds Excel limit of {MAX_INVOICES} invoices "
            f"(max row {EXCEL_MAX_ROW}, data from row {DATA_ROW})"
        )
    return value


def count_label(count: int) -> str:
    if count == MAX_INVOICES:
        return "excel_max"
    if count % 1_000 == 0 and count >= 1_000:
        return f"{count // 1_000}k"
    return str(count)


def resolve_batches(argv: list[str]) -> list[tuple[int, str]]:
    counts = [parse_count(a) for a in argv] if argv else default_counts()
    return [(c, f"submit_invoices_{count_label(c)}.xlsx") for c in counts]


def main() -> None:
    if not os.path.isfile(TEMPLATE_PATH):
        raise SystemExit(f"Template not found: {TEMPLATE_PATH}")

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    batches = resolve_batches(sys.argv[1:])
    print(
        f"Excel limit: max {MAX_INVOICES} single-line invoices "
        f"(row {DATA_ROW}..{EXCEL_MAX_ROW})"
    )
    print(f"Batches: {', '.join(count_label(c) for c, _ in batches)}")
    base_ms = int(time.time() * 1000)
    for idx, (count, file_name) in enumerate(batches):
        write_batch(count, file_name, base_ms + idx * 100_000)


if __name__ == "__main__":
    main()